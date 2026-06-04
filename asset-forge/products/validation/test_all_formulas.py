#!/usr/bin/env python3
"""
ASSET-FORGE · Turnkey Startup Packs — Definition-of-Done validator
==================================================================
Hard gate for the brief's Part-C "Production-Forcing" Definition of Done. Runs
WITHOUT a spreadsheet app: it reopens every generated `.xlsx` with openpyxl and
checks the things that make a pack real rather than a mockup.

    python3 products/validation/test_all_formulas.py          # all packs
    python3 products/validation/test_all_formulas.py 05        # only files matching

Checks per workbook:
  1. opens cleanly (zip + XML well-formed)
  2. all REQUIRED sheets present
  3. NO #REF!/#DIV/0!/#VALUE! baked into any cell or formula
  4. cross-sheet formula references resolve to a real sheet (no broken links)
  5. enough LINKED formulas (not a hard-coded mockup)
  6. traffic-light conditional formatting present where required
  7. sample data present (so calcs prove out)
  8. tab names ≤ 31 chars; 0 < file size < 10 MB
Operations workbook additionally: KPI dashboard exposes ≥ 15 metrics.
SURE workbook additionally: refund calculator + eligibility gate formulas present.

Exit code 0 = all pass; 1 = any failure.
"""
from __future__ import annotations
import os, re, sys, glob
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
HOTEL_DIR = os.path.join(ROOT, "products", "industries", "hospitality", "boutique-hotel-4star")
GRANT_DIR = os.path.join(ROOT, "products", "templates", "grant-applications")

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

REQUIRED_SHEETS = {
    "00_Market_Validation.xlsx": ["Market Validation", "Assumptions"],
    "01_Business_Plan.xlsx": ["Business Plan", "Assumptions"],
    "02_Capital_Raising.xlsx": ["Capital & Funding", "Assumptions"],
    "03_Procurement.xlsx": ["Procurement", "Assumptions"],
    "04_Team_Building.xlsx": ["Team & Payroll", "Assumptions"],
    "05_Operations.xlsx": ["Daily Data Entry", "KPI Dashboard", "Assumptions"],
    "06_Launch_100Days.xlsx": ["Launch 100 Days", "Assumptions"],
}
# minimum linked-formula count expected per workbook (rough mockup guard)
MIN_FORMULAS = {
    "00_Market_Validation.xlsx": 20,
    "01_Business_Plan.xlsx": 40,
    "02_Capital_Raising.xlsx": 60,
    "03_Procurement.xlsx": 25,
    "04_Team_Building.xlsx": 25,
    "05_Operations.xlsx": 30,
    "06_Launch_100Days.xlsx": 12,
}

QUOTED_SHEET_REF = re.compile(r"'([^']+)'!")


class Result:
    def __init__(self, path):
        self.path = path
        self.name = os.path.basename(path)
        self.fails = []
        self.warns = []
        self.stats = {}

    def fail(self, msg):
        self.fails.append(msg)

    def warn(self, msg):
        self.warns.append(msg)

    @property
    def ok(self):
        return not self.fails


def iter_cells(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                yield c


def validate_file(path, required):
    res = Result(path)

    # 8. file size
    size = os.path.getsize(path)
    res.stats["size_kb"] = round(size / 1024, 1)
    if size == 0:
        res.fail("file is empty")
        return res
    if size > 10 * 1024 * 1024:
        res.fail(f"file too large: {size/1e6:.1f} MB > 10 MB")

    # 1. opens cleanly
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        res.fail(f"failed to open: {e}")
        return res

    sheetnames = wb.sheetnames

    # 2. required sheets
    for s in required:
        if s not in sheetnames:
            res.fail(f"missing required sheet: {s!r} (have {sheetnames})")

    # 8b. tab length
    for s in sheetnames:
        if len(s) > 31:
            res.fail(f"tab name >31 chars: {s!r}")

    n_formulas = 0
    n_literals_num = 0
    n_cf = 0
    bad_refs = set()
    error_cells = []

    for ws in wb.worksheets:
        # conditional formatting count
        try:
            n_cf += len(list(ws.conditional_formatting))
        except Exception:
            pass
        for c in iter_cells(ws):
            v = c.value
            if isinstance(v, str):
                # 3. error tokens baked in
                for tok in ERROR_TOKENS:
                    if tok in v:
                        error_cells.append((ws.title, c.coordinate, v))
                if v.startswith("="):
                    n_formulas += 1
                    # 4. cross-sheet reference resolution
                    for ref_sheet in QUOTED_SHEET_REF.findall(v):
                        if ref_sheet not in sheetnames:
                            bad_refs.add((ws.title, c.coordinate, ref_sheet))
            elif isinstance(v, (int, float)):
                n_literals_num += 1

    res.stats["formulas"] = n_formulas
    res.stats["num_literals"] = n_literals_num
    res.stats["cf_rules"] = n_cf

    # 3.
    if error_cells:
        res.fail(f"error tokens in {len(error_cells)} cells e.g. {error_cells[:3]}")
    # 4.
    if bad_refs:
        res.fail(f"broken cross-sheet refs: {sorted(bad_refs)[:5]}")
    # 5. linked-formula floor
    floor = MIN_FORMULAS.get(res.name, 10)
    if n_formulas < floor:
        res.fail(f"only {n_formulas} formulas (< {floor}) — looks hard-coded")
    # 6. traffic-light CF
    if n_cf == 0:
        res.fail("no conditional-formatting rules (no traffic lights)")

    # 7. sample data present (numeric inputs somewhere)
    if n_literals_num < 10:
        res.fail(f"too few numeric values ({n_literals_num}) — sample data missing")

    # --- workbook-specific ---
    if res.name == "05_Operations.xlsx" and "KPI Dashboard" in sheetnames:
        dash = wb["KPI Dashboard"]
        status_formulas = 0
        for c in iter_cells(dash):
            if isinstance(c.value, str) and "On target" in c.value and c.value.startswith("="):
                status_formulas += 1
        res.stats["dashboard_metrics"] = status_formulas
        if status_formulas < 15:
            res.fail(f"dashboard exposes {status_formulas} metrics (< 15 required)")
        # sample data in the data-entry sheet
        de = wb["Daily Data Entry"]
        de_nums = sum(1 for c in iter_cells(de) if isinstance(c.value, (int, float)))
        if de_nums < 30:
            res.fail(f"Daily Data Entry has {de_nums} sample numbers (< 30)")

    if res.name.startswith("IE_SURE"):
        joined = " ".join(
            str(c.value) for ws in wb.worksheets for c in iter_cells(ws)
            if isinstance(c.value, str))
        if "PMT" not in joined and "MIN(" not in joined:
            res.warn("SURE: expected a MIN()/refund calc formula")
        if "PASS" not in joined and "Eligible" not in joined:
            res.warn("SURE: expected an eligibility PASS/FAIL gate")

    return res


def discover(filter_str=None):
    files = []
    for d, reqmap in ((HOTEL_DIR, REQUIRED_SHEETS), (GRANT_DIR, {})):
        for p in sorted(glob.glob(os.path.join(d, "*.xlsx"))):
            name = os.path.basename(p)
            if filter_str and filter_str not in name:
                continue
            files.append((p, reqmap.get(name, ["Assumptions"] if d == HOTEL_DIR else [])))
    return files


def main():
    filter_str = sys.argv[1] if len(sys.argv) > 1 else None
    files = discover(filter_str)
    if not files:
        print("No workbooks found to validate (have they been built yet?)")
        return 1
    results = [validate_file(p, req) for p, req in files]
    print("=" * 74)
    print("ASSET-FORGE · Turnkey Pack — Definition-of-Done validation")
    print("=" * 74)
    any_fail = False
    for r in results:
        mark = "PASS" if r.ok else "FAIL"
        if not r.ok:
            any_fail = True
        extra = " · ".join(f"{k}={v}" for k, v in r.stats.items())
        print(f"[{mark}] {r.name:<34} {extra}")
        for f in r.fails:
            print(f"        ✗ {f}")
        for w in r.warns:
            print(f"        ! {w}")
    print("-" * 74)
    npass = sum(1 for r in results if r.ok)
    print(f"{npass}/{len(results)} workbooks passed")
    return 1 if any_fail else 0


if __name__ == "__main__":
    sys.exit(main())
