#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 13b — Premium Pack design system
====================================================
A reusable openpyxl styling layer so every pack looks premium *by default*.
Encodes the codified "premium look" recipe from
`research/beautification_and_competitors.md`:

  · limited 2-3 colour palette          · gridlines OFF on every sheet
  · clean sans-serif, big bold titles   · accent underline, not heavy banners
  · hairline tables, zebra striping      · KPI tiles + a dedicated dashboard tab
  · generous whitespace (margin column)  · soft input highlight (not harsh amber)

Import this from any pack builder:

    from design_system import Theme, DS
    ds = DS(Theme())                      # or DS(Theme(primary="..."), font="...")
    ds.canvas(ws, widths=[...])
    ds.title(ws, "Title", "subtitle")
    ds.kpi(ws, top=6, left=2, label="...", value="=...", fmt=Theme.EUR)

No external deps beyond openpyxl. Cross-platform-safe fonts only (Excel
substitutes gracefully if a face is missing).
"""
from __future__ import annotations
import math
from dataclasses import dataclass, field
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


def _char_w(size):
    """Approx Excel column-width units per character at a given font size."""
    return (size / 11.0) * 1.06


def _text_w(text, size):
    return len(text) * _char_w(size) + 2.6      # +breathing room so nothing sits on the edge


def _wrap_lines(text, col_width, size):
    """How many lines `text` needs when wrapped into a column of `col_width`."""
    per = max(1, int((col_width - 1) / _char_w(size)))
    total = 0
    for part in str(text).split("\n"):
        words, line = part.split(" "), ""
        n = 1
        for w in words:
            cand = (line + " " + w).strip()
            if len(cand) <= per or not line:
                line = cand
            else:
                n += 1; line = w
        total += n
    return total


@dataclass
class Theme:
    """A pack's colour + format palette. Override per vertical for brand variety."""
    # core 3-colour palette
    ink: str = "1A2B45"        # near-black navy — primary text / titles
    primary: str = "2D6CDF"    # clean blue — table headers, links, KPI accents
    accent: str = "15A38C"     # teal-green — secondary accent / underline
    # neutrals (the whitespace that reads "premium")
    paper: str = "FFFFFF"
    mist: str = "F4F7FB"       # very light blue-grey — zebra / tile body
    band: str = "E9F0FA"       # soft band — section fills
    line: str = "DCE3ED"       # hairline borders
    muted: str = "6B7280"      # grey secondary text
    # input + semantic (soft, not the harsh 90s amber/green/red)
    input: str = "FFF7E6"
    input_line: str = "EBD49B"
    good: str = "1E9E6A"; good_bg: str = "E6F4EC"
    warn: str = "B26B00"; warn_bg: str = "FBEFD6"
    bad:  str = "C0392B"; bad_bg:  str = "F8E1DD"
    # number formats (EU conventions)
    EUR: str = '#,##0.00\\ "€"'
    EUR0: str = '#,##0\\ "€"'
    PCT: str = "0.0%"
    DATE: str = "DD/MM/YYYY"


class DS:
    """Design-system helpers bound to a Theme + a font family."""

    def __init__(self, theme: Theme | None = None, font: str = "Calibri",
                 font_light: str = "Calibri Light"):
        self.t = theme or Theme()
        self.font_name = font
        self.font_light = font_light

    # ---- atoms -----------------------------------------------------------
    def font(self, size=11, bold=False, color=None, light=False, italic=False):
        return Font(name=self.font_light if light else self.font_name,
                    size=size, bold=bold, italic=italic,
                    color=color or self.t.ink)

    def fill(self, color):
        return PatternFill("solid", fgColor=color)

    def _side(self, color=None):
        return Side(style="thin", color=color or self.t.line)

    def hairline_bottom(self, color=None):
        return Border(bottom=self._side(color))

    def box(self, color=None):
        s = self._side(color)
        return Border(left=s, right=s, top=s, bottom=s)

    # ---- canvas ----------------------------------------------------------
    def canvas(self, ws, widths, margin=2.4, tab=None):
        """Gridlines off + a thin margin column A + content column widths."""
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = margin           # breathing room
        for i, w in enumerate(widths, start=2):            # content starts at col B
            ws.column_dimensions[get_column_letter(i)].width = w
        if tab:
            ws.sheet_properties.tabColor = tab

    def title(self, ws, title, subtitle="", row=1, left=2, span=8):
        """Big dark title on paper + a thin accent underline (modern, not a navy slab)."""
        last = left + span - 1
        ws.merge_cells(start_row=row, start_column=left, end_row=row, end_column=last)
        c = ws.cell(row, left, title)
        c.font = self.font(22, bold=True, color=self.t.ink, light=True)
        c.alignment = Alignment("left", "center")
        ws.row_dimensions[row].height = 38
        if subtitle:
            ws.merge_cells(start_row=row + 1, start_column=left, end_row=row + 1, end_column=last)
            s = ws.cell(row + 1, left, subtitle)
            s.font = self.font(11, color=self.t.muted)
            s.alignment = Alignment("left", "center")
            ws.row_dimensions[row + 1].height = 20
            under = row + 2
        else:
            under = row + 1
        # accent underline (a thin filled row)
        for col in range(left, last + 1):
            ws.cell(under, col).fill = self.fill(self.t.accent)
        ws.row_dimensions[under].height = 3
        return under + 1  # next free row

    def section(self, ws, row, text, left=2, span=8):
        """Uppercase section label, primary colour, hairline under."""
        last = left + span - 1
        ws.merge_cells(start_row=row, start_column=left, end_row=row, end_column=last)
        c = ws.cell(row, left, text.upper())
        c.font = self.font(10.5, bold=True, color=self.t.primary)
        c.alignment = Alignment("left", "center")
        for col in range(left, last + 1):
            ws.cell(row, col).border = self.hairline_bottom(self.t.primary)
        ws.row_dimensions[row].height = 24
        return row + 1

    def note(self, ws, row, text, left=2, span=8, tone="info"):
        """Soft info/help band."""
        bg = {"info": self.t.band, "good": self.t.good_bg,
              "warn": self.t.warn_bg, "bad": self.t.bad_bg}[tone]
        fg = {"info": self.t.ink, "good": self.t.good,
              "warn": self.t.warn, "bad": self.t.bad}[tone]
        last = left + span - 1
        ws.merge_cells(start_row=row, start_column=left, end_row=row, end_column=last)
        c = ws.cell(row, left, f"  {text}")
        c.font = self.font(9.5, italic=True, color=fg)
        c.fill = self.fill(bg)
        c.alignment = Alignment("left", "center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 30
        return row + 1

    # ---- tables ----------------------------------------------------------
    def thead(self, ws, row, headers, left=2, height=30):
        """Primary-fill header row, white bold, accent bottom border."""
        for j, h in enumerate(headers, start=left):
            c = ws.cell(row, j, h)
            c.font = self.font(10, bold=True, color="FFFFFF")
            c.fill = self.fill(self.t.primary)
            c.alignment = Alignment("center", "center", wrap_text=True)
            c.border = Border(bottom=self._side(self.t.accent))
        ws.row_dimensions[row].height = height
        return row + 1

    def trow(self, ws, row, ncols, left=2, zebra_on=False, height=19,
             align="center", first_left=True):
        """Style one data row (hairline bottom + optional zebra). Returns nothing."""
        for j in range(left, left + ncols):
            c = ws.cell(row, j)
            c.font = self.font(9.5, color=self.t.ink)
            c.border = self.hairline_bottom()
            a = "left" if (first_left and j == left) else align
            c.alignment = Alignment(a, "center", wrap_text=False,
                                    indent=1 if a == "left" else 0)
            if zebra_on:
                c.fill = self.fill(self.t.mist)
        ws.row_dimensions[row].height = height

    def input_cell(self, ws, row, col, fmt=None):
        c = ws.cell(row, col)
        c.fill = self.fill(self.t.input)
        c.border = self.box(self.t.input_line)
        if fmt:
            c.number_format = fmt
        return c

    def calc_cell(self, ws, row, col, fmt=None, bold=False):
        c = ws.cell(row, col)
        c.fill = self.fill(self.t.band)
        c.font = self.font(9.5, bold=bold, color=self.t.ink)
        c.border = self.hairline_bottom()
        if fmt:
            c.number_format = fmt
        return c

    # ---- KPI tile (the dashboard hero) -----------------------------------
    def kpi(self, ws, top, left, label, value, fmt=None, width=3, accent=None,
            value_size=20):
        """A merged KPI tile: accent strip + label + big value. Returns value cell."""
        accent = accent or self.t.primary
        last_col = left + width - 1
        # accent strip (row top)
        ws.merge_cells(start_row=top, start_column=left, end_row=top, end_column=last_col)
        strip = ws.cell(top, left, "")
        strip.fill = self.fill(accent)
        ws.row_dimensions[top].height = 4
        # label (row top+1)
        ws.merge_cells(start_row=top + 1, start_column=left, end_row=top + 1, end_column=last_col)
        lab = ws.cell(top + 1, left, label.upper())
        lab.font = self.font(9, bold=True, color=self.t.muted)
        lab.fill = self.fill(self.t.paper)
        lab.alignment = Alignment("left", "center", indent=1)
        ws.row_dimensions[top + 1].height = 18
        # value (rows top+2..top+3)
        ws.merge_cells(start_row=top + 2, start_column=left, end_row=top + 3, end_column=last_col)
        val = ws.cell(top + 2, left, value)
        val.font = self.font(value_size, bold=True, color=accent)
        val.fill = self.fill(self.t.paper)
        val.alignment = Alignment("left", "center", indent=1)
        if fmt:
            val.number_format = fmt
        ws.row_dimensions[top + 2].height = 22
        ws.row_dimensions[top + 3].height = 14
        # box border around the whole tile
        for r in range(top, top + 4):
            for c in range(left, last_col + 1):
                cell = ws.cell(r, c)
                sides = {}
                if r == top:
                    sides["top"] = self._side(self.t.line)
                if r == top + 3:
                    sides["bottom"] = self._side(self.t.line)
                if c == left:
                    sides["left"] = self._side(self.t.line)
                if c == last_col:
                    sides["right"] = self._side(self.t.line)
                if sides:
                    cell.border = Border(**sides)
        return val

    def footer(self, ws, row, text, left=2, span=10):
        last = left + span - 1
        ws.merge_cells(start_row=row, start_column=left, end_row=row, end_column=last)
        c = ws.cell(row, left, text)
        c.font = self.font(8, italic=True, color="9AA3B0")
        c.alignment = Alignment("left", "center")
        ws.row_dimensions[row].height = 22

    # ---- legibility guarantee (run last on every sheet) ------------------
    def fit(self, ws, min_w=8.5, max_w=46, line_h=15.0, max_lines=4):
        """Auto-fit so EVERY label/title/header is fully readable with no manual
        column-width or wrap fiddling: widen columns to their content, and grow
        row heights so wrapped text shows in full. Only ever increases sizes."""
        merged_tl, covered = {}, set()
        for m in ws.merged_cells.ranges:
            c1, r1, c2, r2 = range_boundaries(str(m))
            merged_tl[(r1, c1)] = (r2, c2)
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    if (rr, cc) != (r1, c1):
                        covered.add((rr, cc))

        def is_formula(v):
            return isinstance(v, str) and v.startswith("=")

        # ---- pass 1: column widths ----
        colw = {}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or is_formula(v):
                    continue
                coord = (cell.row, cell.column)
                if coord in covered:
                    continue
                text, size = str(v), (cell.font.size or 11)
                wrapped = bool(cell.alignment and cell.alignment.wrap_text)
                indent = (cell.alignment.indent or 0) if cell.alignment else 0
                if coord in merged_tl:                      # spans columns — usually fits
                    if not wrapped:
                        r2, c2 = merged_tl[coord]
                        span_w = sum((ws.column_dimensions[get_column_letter(cc)].width or min_w)
                                     for cc in range(cell.column, c2 + 1))
                        need = _text_w(text, size) + indent
                        if need > span_w:                   # widen last col to absorb overflow
                            colw[c2] = max(colw.get(c2, 0),
                                           (ws.column_dimensions[get_column_letter(c2)].width or min_w)
                                           + (need - span_w))
                    continue
                if wrapped:                                  # ensure longest WORD never clips
                    longest = max((len(w) for w in text.replace("\n", " ").split(" ")), default=1)
                    colw[cell.column] = max(colw.get(cell.column, 0), _text_w("x" * longest, size))
                else:                                        # fit the whole label
                    colw[cell.column] = max(colw.get(cell.column, 0), _text_w(text, size) + indent)
        for c, w in colw.items():
            col = get_column_letter(c)
            cur = ws.column_dimensions[col].width or min_w
            ws.column_dimensions[col].width = round(min(max_w, max(cur, min_w, w)), 1)

        # ---- pass 2: row heights for wrapped cells (using final widths) ----
        rowh = {}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or is_formula(v):
                    continue
                if not (cell.alignment and cell.alignment.wrap_text):
                    continue
                coord = (cell.row, cell.column)
                if coord in covered:
                    continue
                size = cell.font.size or 11
                if coord in merged_tl:
                    r2, c2 = merged_tl[coord]
                    width = sum((ws.column_dimensions[get_column_letter(cc)].width or min_w)
                                for cc in range(cell.column, c2 + 1))
                    span = r2 - cell.row + 1
                else:
                    width = ws.column_dimensions[get_column_letter(cell.column)].width or min_w
                    span = 1
                lines = min(max_lines, _wrap_lines(str(v), width, size))
                need = (lines * line_h) / span
                for rr in range(cell.row, cell.row + span):
                    rowh[rr] = max(rowh.get(rr, 0), need)
        for r, h in rowh.items():
            cur = ws.row_dimensions[r].height or 15
            ws.row_dimensions[r].height = round(max(cur, h), 1)
