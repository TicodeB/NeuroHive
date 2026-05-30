#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 12 — Flagship compliance pack (BONUS track)
P13 — Compliance Gap-Analysis & Mock-Audit (Lite)   [FREE lead magnet]

Bundles the two flagship compliance assets:
  · CA 1 — Clause-by-clause Gap-Analysis Tool
  · CA 2 — Mock-Audit / Readiness Self-Assessment ("will I pass?")

into one bilingual (EN / SK) .xlsx with REAL logic:
  - clause-by-clause % conformance roll-up (generic Annex SL + food spine)
  - per-section conformance breakdown (AVERAGEIFS)
  - auto-prioritised gap flagging (HIGH / MEDIUM)
  - mock-audit readiness score
  - combined readiness % + RAG verdict band ("will I pass?")
  - next-steps routing to the matching paid kit (P14–P18)

"Lite" = one generic clause set (no per-standard depth). The full per-standard
version becomes the scoring engine inside every paid kit (P14–P18). FREE on
Lemon Squeezy with email capture → upsell.

EU conventions: metric units, DD/MM/YYYY dates, comma thousands separators.
Framing rule (brief §15): positioned as "audit-ready for a TÜV-style certifier
auditing you to ISO / HACCP / BRCGS / IFS" — never "TÜV templates".

Run:  python3 scripts/build_p13_gap_analysis_pack.py
Out:  products/P13_Compliance_Gap_Analysis_Mock_Audit.xlsx          (full, free download)
      products/P13_DEMO_Compliance_Gap_Analysis_Mock_Audit.xlsx     (watermarked preview)
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PRODUCTS = os.path.normpath(os.path.join(HERE, "..", "products"))
os.makedirs(PRODUCTS, exist_ok=True)

VERSION = "v1.1"  # v1.1: fixed Excel corruption (31-char tab limit + demo insert_rows)
BUILD_DATE = "30/05/2026"  # DD/MM/YYYY (EU)

# ---- palette (matches P1/P2 flagships) ---------------------------------------
NAVY = "1F3A5F"
TEAL = "2A7D7B"
SAND = "F4EEE2"
LBLUE = "DCE6F1"
LGREY = "F2F2F2"
AMBER = "FFF2CC"
GREEN = "E2EFDA"
RED = "F8CBAD"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- sheet titles (referenced in cross-sheet formulas) -----------------------
# NOTE: Excel hard-limits sheet (tab) names to 31 characters. Exceeding it
# silently corrupts the workbook ("we found a problem with content"). Keep every
# title <= 31 chars — build_workbook() asserts this. Bilingual titles live in
# row 1 of each sheet, so short tab names lose nothing.
S_COVER = "00 · Start Here · Začnite tu"
S_GAP = "01 · Gap Analysis · Analýza"
S_MOCK = "02 · Mock Audit · Skúška"
S_SCORE = "03 · Readiness · Pripravenosť"
S_NEXT = "04 · Next Steps · Ďalšie kroky"


def f(sz=11, b=False, color="000000", italic=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=italic)


def fill(c):
    return PatternFill("solid", fgColor=c)


def style_header(cell, bg=NAVY, fg=WHITE, sz=11):
    cell.font = f(sz, True, fg)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def title_block(ws, en, sk, sub_en, sub_sk, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(1, 1, f"{en}  |  {sk}")
    t.font = f(15, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(2, 1, f"{sub_en}  ·  {sub_sk}")
    s.font = f(10, False, NAVY, italic=True)
    s.fill = fill(SAND)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 26


def biz_field(ws, row, label_en, label_sk, span=8):
    c = ws.cell(row, 1, f"{label_en} / {label_sk}:")
    c.font = f(10, True, NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span)
    v = ws.cell(row, 2, "")
    v.fill = fill(AMBER)
    v.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def note(ws, row, text_en, text_sk, span=8, bg=LBLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, f"ℹ  {text_en}\n    {text_sk}")
    c.font = f(9, False, NAVY, italic=True)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 32


# =============================================================================
# Generic clause set (Annex SL high-level structure + food-safety spine).
# Each section -> list of (clause_ref, requirement_en, requirement_sk).
# "8F · Food Safety" rows are marked N/A by non-food businesses.
# =============================================================================
SECTIONS = [
    ("4 · Context / Kontext", [
        ("4.1", "Internal & external issues affecting the business are identified",
         "Interné a externé faktory ovplyvňujúce prevádzku sú identifikované"),
        ("4.2", "Needs & expectations of interested parties (customers, regulators) recorded",
         "Potreby a očakávania zainteresovaných strán sú zaznamenané"),
        ("4.3/4.4", "Scope of the management system is defined and key processes are mapped",
         "Rozsah systému riadenia je definovaný a kľúčové procesy zmapované"),
    ]),
    ("5 · Leadership / Vedenie", [
        ("5.1", "Top management demonstrably commits to and resources the system",
         "Vrcholový manažment preukázateľne podporuje a zdrojuje systém"),
        ("5.2", "A policy is documented, communicated and available to staff",
         "Politika je zdokumentovaná, komunikovaná a dostupná zamestnancom"),
        ("5.3", "Roles, responsibilities and authorities are assigned and understood",
         "Úlohy, zodpovednosti a právomoci sú pridelené a pochopené"),
    ]),
    ("6 · Planning / Plánovanie", [
        ("6.1", "Risks & opportunities are assessed (incl. climate, Amd 1:2024)",
         "Riziká a príležitosti sú posúdené (vrátane klímy, dodatok 1:2024)"),
        ("6.2", "Objectives are set, measurable, with plans to achieve them",
         "Ciele sú stanovené, merateľné a s plánom na ich dosiahnutie"),
        ("6.3", "Changes to the system are planned and controlled",
         "Zmeny systému sú plánované a riadené"),
    ]),
    ("7 · Support / Podpora", [
        ("7.1", "Resources, infrastructure and work environment are provided",
         "Zdroje, infraštruktúra a pracovné prostredie sú zabezpečené"),
        ("7.2", "Competence: training needs identified and training records kept",
         "Kompetencie: potreby školenia určené a záznamy o školení vedené"),
        ("7.3/7.4", "Staff are aware of the policy; communication is defined",
         "Zamestnanci poznajú politiku; komunikácia je definovaná"),
        ("7.5", "Documented information is version-controlled, approved and retrievable",
         "Dokumentácia má riadené verzie, je schválená a dohľadateľná"),
    ]),
    ("8 · Operation / Prevádzka", [
        ("8.1", "Operational processes are planned and controlled",
         "Prevádzkové procesy sú plánované a riadené"),
        ("8.2", "Work instructions / SOPs exist and are followed in practice",
         "Pracovné postupy (SOP) existujú a v praxi sa dodržiavajú"),
        ("8.4", "Externally provided products/services controlled (suppliers approved)",
         "Externé produkty/služby sú riadené (dodávatelia schválení)"),
    ]),
    ("8F · Food Safety / Bezpečnosť potravín", [
        ("HACCP", "HACCP / PRP plan documented with CCPs and critical limits",
         "HACCP / PRP plán s kritickými bodmi (CCP) a kritickými limitmi"),
        ("FIC", "Allergen management & labelling in place (Reg. 1169/2011)",
         "Riadenie alergénov a označovanie zavedené (nar. 1169/2011)"),
        ("Trace", "Traceability 'one step back/forward' + recall procedure",
         "Vysledovateľnosť „o krok späť/vpred“ + postup stiahnutia produktu"),
        ("852", "Hygiene & temperature monitoring records kept (Reg. 852/2004)",
         "Záznamy hygieny a monitorovania teploty vedené (nar. 852/2004)"),
    ]),
    ("9 · Performance / Hodnotenie", [
        ("9.1", "Monitoring, measurement and analysis are performed and recorded",
         "Monitorovanie, meranie a analýza sa vykonávajú a zaznamenávajú"),
        ("9.2", "An internal audit programme is planned and carried out",
         "Program interných auditov je naplánovaný a vykonávaný"),
        ("9.3", "Management review meetings are held and minuted",
         "Preskúmania manažmentom sa konajú a vedú sa z nich zápisy"),
    ]),
    ("10 · Improvement / Zlepšovanie", [
        ("10.1", "Nonconformities are recorded and corrected",
         "Nezhody sú zaznamenané a odstránené"),
        ("10.2", "Corrective action with root-cause analysis (CAPA log) is used",
         "Nápravné opatrenia s analýzou príčin (CAPA) sa používajú"),
        ("10.3", "Continual improvement is evidenced over time",
         "Trvalé zlepšovanie je preukázané v čase"),
    ]),
]

STATUS_LIST = '"Conform / Zhoda,Partial / Čiastočne,Not in place / Chýba,N/A / Neaplikuje sa"'
ANSWER_LIST = '"Yes / Áno,Partly / Čiastočne,No / Nie"'

# captured at build time so the dashboard can reference exact ranges
META: dict[str, object] = {}


# =============================================================================
# SHEET 0 — Cover / Start here
# =============================================================================
def build_cover(wb):
    ws = wb.active
    ws.title = S_COVER
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 30, 30, 30, 30, 30, 14, 14])

    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, "COMPLIANCE GAP-ANALYSIS & MOCK-AUDIT")
    t.font = f(20, True, WHITE); t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    t2 = ws.cell(2, 1, "Analýza nedostatkov a skúšobný audit — „Prejdem auditom?“")
    t2.font = f(13, True, NAVY); t2.fill = fill(SAND)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    ws.merge_cells("A3:H3")
    t3 = ws.cell(3, 1, "Score your readiness in 15 minutes · Ohodnoťte svoju pripravenosť za 15 minút")
    t3.font = f(10, False, TEAL, italic=True)
    t3.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ("", ""),
        ("WHAT'S INSIDE / ČO OBSAHUJE", "header"),
        ("01 · Clause-by-clause Gap Analysis", "Analýza nedostatkov po jednotlivých bodoch"),
        ("02 · Mock-Audit Readiness Self-Assessment", "Skúšobný audit — sebahodnotenie pripravenosti"),
        ("03 · Readiness Score & Verdict (RAG)", "Skóre pripravenosti a verdikt (RAG)"),
        ("04 · Next Steps — which kit closes your gaps", "Ďalšie kroky — ktorý balík vyrieši vaše medzery"),
        ("", ""),
        ("WHAT IT IS / ČO TO JE", "header"),
        ("A generic readiness check across the common management-system spine",
         "Všeobecná kontrola pripravenosti naprieč spoločnou kostrou systému riadenia"),
        ("(ISO 9001 · ISO 22000 · HACCP · BRCGS · IFS · FSSC 22000 structure)",
         "(štruktúra ISO 9001 · ISO 22000 · HACCP · BRCGS · IFS · FSSC 22000)"),
        ("Audit-ready for a TÜV-style certifier auditing you to these standards —",
         "Pripravené pre certifikačný orgán (štýl TÜV), ktorý vás audituje podľa týchto noriem —"),
        ("NOT a 'TÜV template'. TÜV is a certifier, not a standard.",
         "NIE je to „TÜV šablóna“. TÜV je certifikačný orgán, nie norma."),
        ("", ""),
        ("HOW TO USE / AKO POUŽÍVAŤ", "header"),
        ("1. On Sheet 01, set a Status for each clause (dropdown).",
         "1. Na hárku 01 nastavte stav pre každý bod (rozbaľovací zoznam)."),
        ("2. On Sheet 02, answer the 20 mock-audit questions.",
         "2. Na hárku 02 odpovedzte na 20 otázok skúšobného auditu."),
        ("3. Sheet 03 auto-calculates your readiness % and verdict.",
         "3. Hárok 03 automaticky vypočíta % pripravenosti a verdikt."),
        ("4. Sheet 04 routes you to the kit that closes the gaps.",
         "4. Hárok 04 vás nasmeruje na balík, ktorý medzery vyrieši."),
        ("Amber = you fill · Green/Red = auto-scored · Grey = guidance.",
         "Oranžová = vyplníte · Zelená/červená = automaticky · Sivá = pokyny."),
    ]
    r = 5
    for en, sk in rows:
        if sk == "header":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = ws.cell(r, 1, en)
            c.font = f(12, True, WHITE); c.fill = fill(TEAL)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 24
        elif en == "" and sk == "":
            ws.row_dimensions[r].height = 6
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
            ce = ws.cell(r, 1, en); ce.font = f(10, True, NAVY)
            ce.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            cs = ws.cell(r, 5, sk); cs.font = f(10, False, "404040", italic=True)
            cs.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    fcell = ws.cell(r, 1, f"ASSET-FORGE · {VERSION} · {BUILD_DATE} · EU (Ireland) · FREE lead magnet · "
                          "Template only — not legal advice nor a guarantee of certification. · "
                          "Šablóna — nie je právne poradenstvo ani záruka certifikácie.")
    fcell.font = f(8, False, "808080", italic=True)
    fcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 28
    ws.sheet_properties.tabColor = NAVY


# =============================================================================
# SHEET 1 — Gap Analysis
# =============================================================================
def build_gap(wb):
    ws = wb.create_sheet(S_GAP)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 9, 46, 22, 12, 26, 26, 12])
    title_block(ws, "Clause-by-clause Gap Analysis", "Analýza nedostatkov po jednotlivých bodoch",
                "Current state vs each requirement — % conformance auto-rolls up to Sheet 03",
                "Aktuálny stav vs každá požiadavka — % zhody sa prenáša na hárok 03")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Standard / scheme assessed", "Posudzovaná norma / schéma")
    biz_field(ws, 5, "Date assessed (DD/MM/YYYY)", "Dátum posúdenia")
    note(ws, 6,
         "Set a Status per clause: Conform (100%) · Partial (50%) · Not in place (0%) · N/A (excluded). "
         "Conformance % and Priority fill automatically. Non-food businesses set the 8F food rows to N/A.",
         "Nastavte stav: Zhoda (100 %) · Čiastočne (50 %) · Chýba (0 %) · Neaplikuje (vylúčené). "
         "% zhody a priorita sa vyplnia automaticky. Nepotravinárske prevádzky nastavia riadky 8F na Neaplikuje.")

    headers = [
        "Section\nSekcia", "Clause\nBod", "Requirement\nPožiadavka", "Status\nStav",
        "Conf. %\nZhoda %", "Evidence / where\nDôkaz / kde", "Action to close\nOpatrenie", "Priority\nPriorita",
    ]
    HROW = 8
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(HROW, j, h))
    ws.row_dimensions[HROW].height = 38

    status_dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True)
    ws.add_data_validation(status_dv)

    r = HROW + 1
    first_data = r
    section_codes = []  # (code, label) for dashboard AVERAGEIFS
    for sec_label, clauses in SECTIONS:
        code = sec_label.split(" · ")[0]  # e.g. "4", "8F"
        section_codes.append((code, sec_label))
        for ref, en, sk in clauses:
            ws.cell(r, 1, sec_label).font = f(8, False, "404040")
            ws.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(r, 2, ref).font = f(9, True, NAVY)
            ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
            rq = ws.cell(r, 3, f"{en}\n{sk}")
            rq.font = f(9); rq.alignment = Alignment(vertical="center", wrap_text=True)
            st = ws.cell(r, 4, ""); st.fill = fill(AMBER); status_dv.add(st)
            st.alignment = Alignment(horizontal="center", vertical="center")
            # Conformance %: Conform->1, Partial->0.5, Not->0, N/A/blank-> "" (excluded)
            ws.cell(r, 5).value = (f'=IF(D{r}="","",IF(LEFT(D{r},3)="Con",1,'
                                   f'IF(LEFT(D{r},3)="Par",0.5,IF(LEFT(D{r},3)="Not",0,""))))')
            ws.cell(r, 5).number_format = "0%"
            ws.cell(r, 5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 6, ""); ws.cell(r, 6).fill = fill(AMBER)
            ws.cell(r, 7, ""); ws.cell(r, 7).fill = fill(AMBER)
            # Priority: Not in place -> HIGH, Partial -> MEDIUM, else blank
            ws.cell(r, 8).value = (f'=IF(D{r}="","",IF(LEFT(D{r},3)="Not","HIGH",'
                                   f'IF(LEFT(D{r},3)="Par","MEDIUM","")))')
            ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 8).font = f(9, True)
            for j in range(1, 9):
                ws.cell(r, j).border = BORDER
            ws.row_dimensions[r].height = 30
            r += 1
    last_data = r - 1

    # conditional formatting
    ws.conditional_formatting.add(f"E{first_data}:E{last_data}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.99"], fill=fill(GREEN)))
    ws.conditional_formatting.add(f"E{first_data}:E{last_data}",
        CellIsRule(operator="lessThan", formula=["0.5"], fill=fill(RED)))
    ws.conditional_formatting.add(f"E{first_data}:E{last_data}",
        CellIsRule(operator="between", formula=["0.5", "0.99"], fill=fill(AMBER)))
    ws.conditional_formatting.add(f"H{first_data}:H{last_data}",
        CellIsRule(operator="equal", formula=['"HIGH"'], fill=fill(RED), font=f(9, True, "9C0006")))
    ws.conditional_formatting.add(f"H{first_data}:H{last_data}",
        CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=fill(AMBER)))

    # overall conformance line at the bottom
    tot = last_data + 1
    ws.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=4)
    c = ws.cell(tot, 1, "OVERALL CONFORMANCE / CELKOVÁ ZHODA →")
    c.font = f(11, True, WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ov = ws.cell(tot, 5, f'=IF(COUNT(E{first_data}:E{last_data})=0,"",AVERAGE(E{first_data}:E{last_data}))')
    ov.number_format = "0%"; ov.font = f(12, True, NAVY); ov.fill = fill(SAND)
    ov.alignment = Alignment(horizontal="center", vertical="center"); ov.border = BORDER
    ws.merge_cells(start_row=tot, start_column=6, end_row=tot, end_column=8)
    hg = ws.cell(tot, 6, '=CONCATENATE("HIGH-priority gaps / kritické medzery: ",'
                         f'COUNTIF(H{first_data}:H{last_data},"HIGH"))')
    hg.font = f(10, True, "9C0006"); hg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[tot].height = 24

    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = f"A{HROW+1}"

    META["gap_e_range"] = f"'{S_GAP}'!E{first_data}:E{last_data}"
    META["gap_h_range"] = f"'{S_GAP}'!H{first_data}:H{last_data}"
    META["gap_sec_range"] = f"'{S_GAP}'!A{first_data}:A{last_data}"
    META["gap_overall"] = f"'{S_GAP}'!E{tot}"
    META["sections"] = section_codes


# =============================================================================
# SHEET 2 — Mock-Audit / Readiness Self-Assessment
# =============================================================================
QUESTIONS = [
    ("Could you show a current policy, signed and dated by management?",
     "Viete predložiť aktuálnu politiku podpísanú a datovanú manažmentom?"),
    ("Could you produce an up-to-date scope / process map of the business?",
     "Viete predložiť aktuálny rozsah / mapu procesov prevádzky?"),
    ("Could you show this year's measurable objectives and progress?",
     "Viete ukázať tohtoročné merateľné ciele a pokrok?"),
    ("Has your risk register been reviewed in the last 12 months?",
     "Bol register rizík preskúmaný za posledných 12 mesiacov?"),
    ("Could you show training records for every member of staff?",
     "Viete predložiť záznamy o školení pre každého zamestnanca?"),
    ("Would staff, asked at random, know the policy and their duties?",
     "Poznali by náhodne oslovení zamestnanci politiku a svoje povinnosti?"),
    ("Is your documented information version-controlled (no undated drafts)?",
     "Má vaša dokumentácia riadené verzie (žiadne nedatované koncepty)?"),
    ("Could you produce an approved-supplier list with evidence of approval?",
     "Viete predložiť zoznam schválených dodávateľov s dôkazom schválenia?"),
    ("(Food) Could you show a HACCP plan with CCPs and critical limits?",
     "(Potraviny) Viete ukázať HACCP plán s CCP a kritickými limitmi?"),
    ("(Food) Could you give allergen information for every item/product?",
     "(Potraviny) Viete poskytnúť informácie o alergénoch pre každú položku?"),
    ("Could you trace a batch one step back and one step forward in minutes?",
     "Viete vystopovať šaržu o krok späť a o krok vpred v priebehu minút?"),
    ("Could you show monitoring records (temperatures/checks) for last month?",
     "Viete ukázať záznamy monitorovania (teploty/kontroly) za minulý mesiac?"),
    ("Have you completed at least one internal audit in the last 12 months?",
     "Vykonali ste aspoň jeden interný audit za posledných 12 mesiacov?"),
    ("Could you produce minutes of a management review meeting?",
     "Viete predložiť zápis z preskúmania manažmentom?"),
    ("Do you keep a nonconformity / corrective-action log that is actually used?",
     "Vediete register nezhôd / nápravných opatrení, ktorý sa naozaj používa?"),
    ("Could you show a closed-out corrective action with root-cause analysis?",
     "Viete ukázať uzavreté nápravné opatrenie s analýzou koreňovej príčiny?"),
    ("Could you show calibration / verification records for monitoring equipment?",
     "Viete ukázať záznamy o kalibrácii / overení meradiel?"),
    ("Can you evidence continual improvement (a clear before/after)?",
     "Viete preukázať trvalé zlepšovanie (jasný stav pred/po)?"),
    ("Are records retained for the required period (≥ 12 months) and retrievable?",
     "Sú záznamy uchované požadovaný čas (≥ 12 mesiacov) a dohľadateľné?"),
    ("If an auditor arrived unannounced today, would your records be in order?",
     "Keby dnes prišiel audítor bez ohlásenia, boli by vaše záznamy v poriadku?"),
]


def build_mock(wb):
    ws = wb.create_sheet(S_MOCK)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [5, 64, 20, 12, 36])
    title_block(ws, "Mock-Audit · Readiness Self-Assessment", "Skúšobný audit · Sebahodnotenie pripravenosti",
                "Answer honestly as if the auditor is in the room — score auto-rolls up to Sheet 03",
                "Odpovedajte úprimne, akoby bol audítor v miestnosti — skóre sa prenáša na hárok 03", span=5)
    biz_field(ws, 3, "Assessed by", "Posúdil", span=5)
    biz_field(ws, 4, "Date (DD/MM/YYYY)", "Dátum", span=5)
    note(ws, 5,
         "Yes = you could evidence it now (100%) · Partly = some evidence/gaps (50%) · No = not in place (0%). "
         "Be honest — the point is to find gaps before the auditor does.",
         "Áno = viete to doložiť teraz (100 %) · Čiastočne = časť dôkazov (50 %) · Nie = chýba (0 %). "
         "Buďte úprimní — cieľom je nájsť medzery skôr ako audítor.", span=5)

    headers = ["#", "Question / Otázka", "Answer\nOdpoveď", "Score\nSkóre", "Note / Poznámka"]
    HROW = 7
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(HROW, j, h))
    ws.row_dimensions[HROW].height = 32

    ans_dv = DataValidation(type="list", formula1=ANSWER_LIST, allow_blank=True)
    ws.add_data_validation(ans_dv)

    r = HROW + 1
    first = r
    for i, (en, sk) in enumerate(QUESTIONS, start=1):
        ws.cell(r, 1, i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 1).font = f(9, True, NAVY)
        q = ws.cell(r, 2, f"{en}\n{sk}")
        q.font = f(9); q.alignment = Alignment(vertical="center", wrap_text=True)
        a = ws.cell(r, 3, ""); a.fill = fill(AMBER); ans_dv.add(a)
        a.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 4).value = (f'=IF(C{r}="","",IF(LEFT(C{r},3)="Yes",1,'
                               f'IF(LEFT(C{r},3)="Par",0.5,0)))')
        ws.cell(r, 4).number_format = "0%"
        ws.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 5, ""); ws.cell(r, 5).fill = fill(LGREY)
        for j in range(1, 6):
            ws.cell(r, j).border = BORDER
        if i % 2 == 0:
            for j in (1, 5):
                ws.cell(r, j).fill = fill(LGREY)
        ws.row_dimensions[r].height = 28
        r += 1
    last = r - 1

    ws.conditional_formatting.add(f"D{first}:D{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.99"], fill=fill(GREEN)))
    ws.conditional_formatting.add(f"D{first}:D{last}",
        CellIsRule(operator="lessThan", formula=["0.5"], fill=fill(RED)))
    ws.conditional_formatting.add(f"D{first}:D{last}",
        CellIsRule(operator="between", formula=["0.5", "0.99"], fill=fill(AMBER)))

    tot = last + 1
    ws.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=3)
    c = ws.cell(tot, 1, "MOCK-AUDIT READINESS / PRIPRAVENOSŤ →")
    c.font = f(11, True, WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    sc = ws.cell(tot, 4, f'=IF(COUNT(D{first}:D{last})=0,"",AVERAGE(D{first}:D{last}))')
    sc.number_format = "0%"; sc.font = f(12, True, NAVY); sc.fill = fill(SAND)
    sc.alignment = Alignment(horizontal="center", vertical="center"); sc.border = BORDER
    ws.cell(tot, 5, "").border = BORDER
    ws.row_dimensions[tot].height = 24

    ws.sheet_properties.tabColor = "2E75B6"
    ws.freeze_panes = f"A{HROW+1}"

    META["mock_score"] = f"'{S_MOCK}'!D{tot}"


# =============================================================================
# SHEET 3 — Readiness Score & Verdict (dashboard)
# =============================================================================
def build_score(wb):
    ws = wb.create_sheet(S_SCORE)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [34, 18, 14, 30, 16, 16])
    title_block(ws, "Readiness Score & Verdict", "Skóre pripravenosti a verdikt",
                "Auto-calculated from Sheets 01 & 02 — the answer to 'will I pass?'",
                "Automaticky z hárkov 01 a 02 — odpoveď na otázku „prejdem?“", span=6)

    gap = META["gap_overall"]
    mock = META["mock_score"]

    # KPI block
    r = 4
    def kpi(row, label_en, label_sk, formula, big=False):
        l = ws.cell(row, 1, f"{label_en} / {label_sk}")
        l.font = f(11, True, NAVY); l.alignment = Alignment(vertical="center", wrap_text=True)
        l.fill = fill(LBLUE); l.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        v = ws.cell(row, 4, formula)
        v.number_format = "0%"; v.border = BORDER
        v.font = f(16 if big else 12, True, NAVY if not big else TEAL)
        v.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 30 if not big else 40

    kpi(r, "Gap-analysis conformance", "Zhoda z analýzy nedostatkov", f'=IF({gap}="","",{gap})')
    kpi(r + 1, "Mock-audit readiness", "Pripravenosť zo skúšobného auditu", f'=IF({mock}="","",{mock})')
    # overall = average of the two if both present, else whichever exists
    overall_f = (f'=IF(AND({gap}="",{mock}=""),"",'
                 f'IF({gap}="",{mock},IF({mock}="",{gap},({gap}+{mock})/2)))')
    kpi(r + 2, "OVERALL READINESS", "CELKOVÁ PRIPRAVENOSŤ", overall_f, big=True)
    overall_cell = f"D{r+2}"

    # verdict band
    vr = r + 4
    ws.merge_cells(start_row=vr, start_column=1, end_row=vr, end_column=6)
    vh = ws.cell(vr, 1, "VERDICT — WILL I PASS? / VERDIKT — PREJDEM?")
    style_header(vh, NAVY); ws.row_dimensions[vr].height = 22
    ws.merge_cells(start_row=vr + 1, start_column=1, end_row=vr + 2, end_column=6)
    verdict = ws.cell(vr + 1, 1, (
        f'=IF({overall_cell}="","Complete Sheets 01 & 02 to see your verdict / Vyplňte hárky 01 a 02",'
        f'IF({overall_cell}>=0.85,"🟢 LIKELY TO PASS — tidy the ambers, then book. / PRAVDEPODOBNE PREJDETE — doriešte oranžové a objednajte audit.",'
        f'IF({overall_cell}>=0.6,"🟠 AT RISK — close the red (HIGH) gaps before booking. / RIZIKO — pred objednaním uzavrite červené (kritické) medzery.",'
        f'"🔴 NOT READY — significant gaps; work through Sheet 04 first. / NEPRIPRAVENÍ — vážne medzery; najprv prejdite hárok 04.")))'))
    verdict.font = f(12, True, NAVY)
    verdict.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    verdict.fill = fill(SAND); verdict.border = BORDER
    ws.row_dimensions[vr + 1].height = 24
    ws.row_dimensions[vr + 2].height = 24

    # RAG band legend
    lr = vr + 4
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=6)
    style_header(ws.cell(lr, 1, "RAG BANDS / PÁSMA RAG"), TEAL); ws.row_dimensions[lr].height = 20
    bands = [
        ("🟢 ≥ 85%", "Likely to pass — minor tidy-ups", "Pravdepodobne prejdete — drobné úpravy", GREEN),
        ("🟠 60–84%", "At risk — close HIGH gaps first", "Riziko — najprv kritické medzery", AMBER),
        ("🔴 < 60%", "Not ready — significant work needed", "Nepripravení — vážna práca", RED),
    ]
    br = lr + 1
    for band, en, sk, col in bands:
        ws.cell(br, 1, band).fill = fill(col); ws.cell(br, 1).font = f(10, True)
        ws.cell(br, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=br, start_column=2, end_row=br, end_column=4)
        ws.cell(br, 2, en).font = f(9); ws.cell(br, 2).alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=br, start_column=5, end_row=br, end_column=6)
        ws.cell(br, 5, sk).font = f(9, italic=True, color="404040")
        ws.cell(br, 5).alignment = Alignment(vertical="center", indent=1)
        for j in range(1, 7):
            ws.cell(br, j).border = BORDER
        ws.row_dimensions[br].height = 20
        br += 1

    # gap counters
    gc = br + 1
    ws.merge_cells(start_row=gc, start_column=1, end_row=gc, end_column=6)
    style_header(ws.cell(gc, 1, "OPEN GAPS / OTVORENÉ MEDZERY"), TEAL); ws.row_dimensions[gc].height = 20
    ws.merge_cells(start_row=gc + 1, start_column=1, end_row=gc + 1, end_column=4)
    ws.cell(gc + 1, 1, "HIGH priority (Not in place) / kritické (chýba)").font = f(10, True, "9C0006")
    ws.cell(gc + 1, 1).alignment = Alignment(vertical="center", indent=1); ws.cell(gc + 1, 1).border = BORDER
    ws.merge_cells(start_row=gc + 1, start_column=5, end_row=gc + 1, end_column=6)
    hc = ws.cell(gc + 1, 5, f'=COUNTIF({META["gap_h_range"]},"HIGH")')
    hc.font = f(12, True, "9C0006"); hc.alignment = Alignment(horizontal="center", vertical="center"); hc.border = BORDER
    ws.merge_cells(start_row=gc + 2, start_column=1, end_row=gc + 2, end_column=4)
    ws.cell(gc + 2, 1, "MEDIUM priority (Partial) / stredné (čiastočne)").font = f(10, True, "BF8F00")
    ws.cell(gc + 2, 1).alignment = Alignment(vertical="center", indent=1); ws.cell(gc + 2, 1).border = BORDER
    ws.merge_cells(start_row=gc + 2, start_column=5, end_row=gc + 2, end_column=6)
    mc = ws.cell(gc + 2, 5, f'=COUNTIF({META["gap_h_range"]},"MEDIUM")')
    mc.font = f(12, True, "BF8F00"); mc.alignment = Alignment(horizontal="center", vertical="center"); mc.border = BORDER
    ws.row_dimensions[gc + 1].height = 20; ws.row_dimensions[gc + 2].height = 20

    # section breakdown
    sb = gc + 4
    ws.merge_cells(start_row=sb, start_column=1, end_row=sb, end_column=6)
    style_header(ws.cell(sb, 1, "SECTION BREAKDOWN / ROZPIS PO SEKCIÁCH"), TEAL); ws.row_dimensions[sb].height = 20
    hdr = sb + 1
    style_header(ws.cell(hdr, 1, "Section / Sekcia"))
    ws.merge_cells(start_row=hdr, start_column=1, end_row=hdr, end_column=4)
    style_header(ws.cell(hdr, 5, "Conf. % / Zhoda %"))
    ws.merge_cells(start_row=hdr, start_column=5, end_row=hdr, end_column=6)
    rr = hdr + 1
    for code, label in META["sections"]:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
        ws.cell(rr, 1, label).font = f(9); ws.cell(rr, 1).alignment = Alignment(vertical="center", indent=1)
        ws.cell(rr, 1).border = BORDER
        ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=6)
        v = ws.cell(rr, 5, f'=IFERROR(AVERAGEIFS({META["gap_e_range"]},{META["gap_sec_range"]},"{label}"),"—")')
        v.number_format = "0%"; v.alignment = Alignment(horizontal="center", vertical="center"); v.border = BORDER
        v.font = f(9)
        ws.conditional_formatting.add(f"E{rr}",
            CellIsRule(operator="greaterThanOrEqual", formula=["0.85"], fill=fill(GREEN)))
        ws.conditional_formatting.add(f"E{rr}",
            CellIsRule(operator="lessThan", formula=["0.6"], fill=fill(RED)))
        ws.conditional_formatting.add(f"E{rr}",
            CellIsRule(operator="between", formula=["0.6", "0.85"], fill=fill(AMBER)))
        ws.row_dimensions[rr].height = 18
        rr += 1

    ws.sheet_properties.tabColor = "548235"


# =============================================================================
# SHEET 4 — Next Steps / upsell routing
# =============================================================================
def build_next(wb):
    ws = wb.create_sheet(S_NEXT)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [30, 36, 14, 44])
    title_block(ws, "Next Steps — Close Your Gaps", "Ďalšie kroky — vyriešte svoje medzery",
                "Pick the readiness kit that matches your sector & standard",
                "Vyberte balík pripravenosti podľa vášho odvetvia a normy", span=4)
    note(ws, 3,
         "This free tool finds the gaps. The matching paid kit gives you the ready-made templates, logs and "
         "procedures to close them — built once, audit-ready. Prices indicative; see the listing.",
         "Tento bezplatný nástroj nájde medzery. Príslušný platený balík vám dá hotové šablóny, záznamy a "
         "postupy na ich vyriešenie — pripravené na audit. Ceny orientačné; pozri ponuku.", span=4)

    headers = ["Your sector / standard\nVaše odvetvie / norma", "Recommended kit\nOdporúčaný balík",
               "From €\nOd €", "What it closes\nČo vyrieši"]
    HROW = 5
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(HROW, j, h))
    ws.row_dimensions[HROW].height = 34

    rows = [
        ("Café · restaurant · bar · B&B · hotel\nHACCP · 852/2004 · 1169/2011",
         "P14 — HACCP Readiness Pack\n(Cafés & Restaurants)", "49",
         "HACCP plan, CCP/PRP logs, allergen matrix, traceability, training, CAPA"),
        ("Food manufacturing (GFSI)\nISO 22000 · FSSC 22000 v7",
         "P15 — ISO 22000 / FSSC 22000\nFood Safety Management Kit", "99",
         "Full FSMS spine: doc-control, internal audit, mgmt review, CAPA, calibration"),
        ("Food manufacturing\nBRCGS · IFS Food",
         "P16 — BRCGS / IFS Document-Control\n& Audit-Readiness Suite", "89",
         "Master document list, supplier approval, internal audit — the #1 NC area"),
        ("Non-food manufacturing\nISO 9001",
         "P17 — ISO 9001 Quality-Management\nAudit-Readiness Pack", "79",
         "ISO 9001 clause 4–10 spine for B2B customer-required certification"),
        ("Already certified FSSC v6\nupgrade to v7 by Apr 2028",
         "P18 — FSSC 22000 V7\nTransition Pack", "49",
         "v6→v7 delta gap-analysis, internal audit & management-review updates"),
    ]
    r = HROW + 1
    for sec, kit, price, closes in rows:
        ws.cell(r, 1, sec).font = f(9, True, NAVY)
        ws.cell(r, 2, kit).font = f(9, True, TEAL)
        ws.cell(r, 3, price).font = f(11, True, NAVY)
        ws.cell(r, 4, closes).font = f(9)
        for j in range(1, 5):
            ws.cell(r, j).border = BORDER
            ws.cell(r, j).alignment = Alignment(
                wrap_text=True, vertical="center",
                horizontal="center" if j == 3 else "left", indent=0 if j == 3 else 1)
        if (r - HROW) % 2 == 0:
            for j in (1, 2, 4):
                ws.cell(r, j).fill = fill(LGREY)
        ws.row_dimensions[r].height = 46
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, "Auditors & consultants / Audítori a konzultanti → P19 Auditor Edition · "
                      "P20 Multi-Client Console (from €149)")
    c.font = f(10, True, WHITE); c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 26

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    fc = ws.cell(r, 1, "Positioning: audit-ready packs that satisfy a TÜV-style certifier auditing you to "
                       "ISO / HACCP / BRCGS / IFS — not 'TÜV templates'. · "
                       "Balíky pripravené na audit certifikačným orgánom (štýl TÜV) podľa ISO / HACCP / BRCGS / IFS.")
    fc.font = f(8, False, "808080", italic=True)
    fc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 28

    ws.sheet_properties.tabColor = "BF8F00"


def build_workbook():
    META.clear()
    wb = Workbook()
    build_cover(wb)
    build_gap(wb)
    build_mock(wb)
    build_score(wb)
    build_next(wb)
    # Excel corrupts on tab names > 31 chars — fail loudly if we ever regress.
    for ws in wb.worksheets:
        assert len(ws.title) <= 31, f"Sheet name too long ({len(ws.title)}): {ws.title!r}"
    wb.properties.title = "Compliance Gap-Analysis & Mock-Audit (Lite) — EN/SK"
    wb.properties.creator = "ASSET-FORGE"
    wb.properties.subject = ("Free compliance readiness check — generic ISO/HACCP/BRCGS/IFS/FSSC spine; "
                             "gap-analysis + mock-audit + readiness verdict")
    return wb


def add_demo_watermark(wb):
    """Watermarked, locked preview for the listing image.

    IMPORTANT: do NOT use insert_rows() here. openpyxl's insert_rows shifts cell
    values but not merged-cell ranges, conditional formatting or data-validation
    ranges — on sheets full of merges that produces overlapping ranges and Excel
    reports "we found a problem with content". Instead we prepend a dedicated,
    non-destructive DEMO notice sheet and lock every sheet read-only.
    """
    notice = wb.create_sheet("DEMO Preview · Ukážka", 0)
    notice.sheet_view.showGridLines = False
    set_widths(notice, [4, 28, 28, 28, 28])
    notice.merge_cells("A1:E1")
    b = notice.cell(1, 1, "DEMO — PREVIEW ONLY · NOT FOR RESALE")
    b.font = f(16, True, "FFFFFF"); b.fill = fill("C00000")
    b.alignment = Alignment(horizontal="center", vertical="center")
    notice.row_dimensions[1].height = 36
    notice.merge_cells("A2:E2")
    b2 = notice.cell(2, 1, "Ukážka — len na náhľad · nie na ďalší predaj")
    b2.font = f(12, True, "C00000"); b2.fill = fill(SAND)
    b2.alignment = Alignment(horizontal="center", vertical="center")
    notice.row_dimensions[2].height = 24
    lines = [
        "",
        "This is a locked, read-only preview of the workbook.",
        "Toto je uzamknutý náhľad zošita len na čítanie.",
        "",
        "Browse the tabs to see the Gap Analysis, Mock Audit, Readiness",
        "verdict and Next Steps. Buy the full pack to unlock editing.",
        "Prehliadnite si hárky. Kúpou plnej verzie odomknete úpravy.",
    ]
    r = 4
    for ln in lines:
        notice.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = notice.cell(r, 1, ln)
        c.font = f(11, False, NAVY, italic=(ln.startswith("Toto") or ln.startswith("Prehliad")))
        c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    notice.sheet_properties.tabColor = "C00000"

    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = "demo"
    wb.properties.title = "Compliance Gap-Analysis & Mock-Audit (Lite) — DEMO (EN/SK)"
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    full = os.path.join(PRODUCTS, "P13_Compliance_Gap_Analysis_Mock_Audit.xlsx")
    wb.save(full)
    print(f"✓ Built flagship: {full}  ({len(wb.sheetnames)} sheets)")

    demo = add_demo_watermark(build_workbook())
    demo_path = os.path.join(PRODUCTS, "P13_DEMO_Compliance_Gap_Analysis_Mock_Audit.xlsx")
    demo.save(demo_path)
    print(f"✓ Built watermarked demo: {demo_path}")
    print("  Sheets:", " | ".join(wb.sheetnames))
