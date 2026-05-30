#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 9 — Flagship build
P1 — Café / Restaurant Compliance Pack  (HOSPITALITY-FIRST flagship)

Builds the bilingual (EN / SK) EHO-ready compliance workbook bundling the 7
MUST assets (DB ids 1,2,3,4,5,16,17) into one linked .xlsx, plus a watermarked
read-only DEMO for the listing preview.

Legal floor satisfied: Reg. 852/2004 (food hygiene) · Reg. 1169/2011 (allergens)
· Fire Services Acts 1981/2003 · Safety, Health & Welfare at Work Act 2005.

EU conventions: metric units, DD/MM/YYYY dates, comma thousands separators.
Run:  python3 scripts/build_p1_compliance_pack.py
Out:  products/P1_Cafe_Restaurant_Compliance_Pack.xlsx
      products/P1_DEMO_Cafe_Restaurant_Compliance_Pack.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PRODUCTS = os.path.normpath(os.path.join(HERE, "..", "products"))
os.makedirs(PRODUCTS, exist_ok=True)

VERSION = "v1.0"
BUILD_DATE = "30/05/2026"  # DD/MM/YYYY (EU)

# ---- palette (calm, print-friendly) -----------------------------------------
NAVY = "1F3A5F"
TEAL = "2A7D7B"
SAND = "F4EEE2"
LBLUE = "DCE6F1"
LGREY = "F2F2F2"
AMBER = "FFF2CC"
GREEN = "E2EFDA"
RED = "F8CBAD"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=11, b=False, color="000000", italic=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=italic)

def fill(c):
    return PatternFill("solid", fgColor=c)

def style_header(cell, bg=NAVY, fg=WHITE, sz=11):
    cell.font = f(sz, True, fg)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def title_block(ws, en, sk, sub_en, sub_sk, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(1, 1, f"{en}  |  {sk}")
    t.font = f(15, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(2, 1, f"{sub_en}  ·  {sub_sk}")
    s.font = f(10, False, NAVY, italic=True)
    s.fill = fill(SAND)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 26

def biz_field(ws, row, label_en, label_sk, span=8):
    """A 'fill me in' business-detail row."""
    c = ws.cell(row, 1, f"{label_en} / {label_sk}:")
    c.font = f(10, True, NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span)
    v = ws.cell(row, 2, "")
    v.fill = fill(AMBER)
    v.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def header_row(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(row, j, h))
    ws.row_dimensions[row].height = 40

def blank_rows(ws, start, n, ncols, zebra=True, date_col=None):
    for r in range(start, start + n):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c, "")
            cell.border = BORDER
            if zebra and (r - start) % 2 == 1:
                cell.fill = fill(LGREY)
            if date_col and c == date_col:
                cell.number_format = "DD/MM/YYYY"
        ws.row_dimensions[r].height = 18

def note(ws, row, text_en, text_sk, span=8, bg=LBLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, f"ℹ  {text_en}\n    {text_sk}")
    c.font = f(9, False, NAVY, italic=True)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 30


# =============================================================================
# SHEET 0 — Cover / Návod (how to use)
# =============================================================================
def build_cover(wb):
    ws = wb.active
    ws.title = "00 · Start Here · Začnite tu"
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 30, 30, 30, 30, 30, 14, 14])

    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, "CAFÉ / RESTAURANT COMPLIANCE PACK")
    t.font = f(20, True, WHITE); t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    t2 = ws.cell(2, 1, "Súbor pre súlad pre kaviarne a reštaurácie")
    t2.font = f(13, True, NAVY); t2.fill = fill(SAND)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    ws.merge_cells("A3:H3")
    t3 = ws.cell(3, 1, "One folder that answers an EHO inspection · Jeden súbor, ktorý obstojí pri kontrole hygieny")
    t3.font = f(10, False, TEAL, italic=True)
    t3.alignment = Alignment(horizontal="center", vertical="center")

    # What's inside
    rows = [
        ("", ""),
        ("WHAT'S INSIDE / ČO OBSAHUJE", "header"),
        ("01 · HACCP Food Safety Management System", "Systém riadenia bezpečnosti potravín HACCP"),
        ("02 · Allergen Matrix & Menu Declaration", "Matica alergénov a označovanie jedálneho lístka"),
        ("03 · Temperature Monitoring Log", "Záznamník monitorovania teploty"),
        ("04 · Cleaning & Sanitation Schedule", "Harmonogram čistenia a sanitácie"),
        ("05 · Supplier & Delivery Traceability Log", "Záznam vysledovateľnosti dodávateľov a dodávok"),
        ("06 · H&S Risk Assessment & Safety Statement", "Posúdenie rizík BOZP a bezpečnostné vyhlásenie"),
        ("07 · Fire Safety Register & Checks Log", "Register požiarnej ochrany a záznam kontrol"),
        ("", ""),
        ("LEGAL BASIS / ZÁKONNÝ ZÁKLAD", "header"),
        ("Reg. (EC) 852/2004 — food hygiene / hygiena potravín", "HACCP, temperature, cleaning, traceability"),
        ("Reg. (EU) 1169/2011 — allergens / alergény (FIC)", "14 allergen declaration / označovanie 14 alergénov"),
        ("Fire Services Acts 1981 & 2003 / zákon o požiarnej ochrane", "Fire register & drills / register a cvičenia"),
        ("Safety, Health & Welfare at Work Act 2005 / zákon o BOZP", "Risk assessment & safety statement"),
        ("", ""),
        ("HOW TO USE / AKO POUŽÍVAŤ", "header"),
        ("1. Fill the amber business-detail cells on each sheet.",
         "Vyplňte oranžové bunky s údajmi o prevádzke na každom hárku."),
        ("2. Print logs or complete on screen daily/weekly as dated.",
         "Záznamy tlačte alebo vypĺňajte denne/týždenne podľa dátumu."),
        ("3. Keep at least 12 months of records for inspection.",
         "Uchovávajte záznamy minimálne 12 mesiacov pre kontrolu."),
        ("4. Amber = you fill · Grey = guidance · White = log entries.",
         "Oranžová = vyplníte · Sivá = pokyny · Biela = záznamy."),
    ]
    r = 5
    for en, sk in rows:
        if sk == "header":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = ws.cell(r, 1, en)
            c.font = f(12, True, WHITE); c.fill = fill(TEAL)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 24
        elif en == "" and sk == "":
            ws.row_dimensions[r].height = 6
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
            ce = ws.cell(r, 1, en); ce.font = f(10, True, NAVY)
            ce.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            cs = ws.cell(r, 5, sk); cs.font = f(10, False, "404040", italic=True)
            cs.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            ws.row_dimensions[r].height = 18
        r += 1

    # footer
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    fcell = ws.cell(r, 1, f"ASSET-FORGE · {VERSION} · {BUILD_DATE} · EU (Ireland) · "
                          "Template only — not legal advice. Adapt to your premises. · "
                          "Šablóna — nie je právne poradenstvo. Prispôsobte svojej prevádzke.")
    fcell.font = f(8, False, "808080", italic=True)
    fcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 26
    ws.sheet_properties.tabColor = NAVY


# =============================================================================
# SHEET 1 — HACCP FSMS
# =============================================================================
def build_haccp(wb):
    ws = wb.create_sheet("01 · HACCP")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 24, 20, 16, 18, 22, 16, 16])
    title_block(ws, "HACCP Food Safety Management System",
                "Systém riadenia bezpečnosti potravín HACCP",
                "Reg. 852/2004 — documented food safety system every food business must keep",
                "Reg. 852/2004 — dokumentovaný systém bezpečnosti potravín")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Person in charge (FSMS owner)", "Zodpovedná osoba (FSMS)")
    biz_field(ws, 5, "Date prepared / reviewed", "Dátum vypracovania / revízie")

    note(ws, 7,
         "CCP = Critical Control Point. List each step where a hazard must be controlled, the critical limit, how it's monitored, and the corrective action if it fails.",
         "CCP = kritický kontrolný bod. Uveďte krok, kritický limit, spôsob monitorovania a nápravné opatrenie pri prekročení.")

    header_row(ws, 8, [
        "CCP / Step\nKrok",
        "Hazard (B/C/P)\nNebezpečenstvo",
        "Critical limit\nKritický limit",
        "Monitoring\nMonitorovanie",
        "Frequency\nČetnosť",
        "Corrective action\nNápravné opatrenie",
        "Record (sheet)\nZáznam",
        "Verified by\nOveril",
    ])
    seed = [
        ("Chilled storage / Chladené skladovanie", "B — pathogen growth", "≤ 5 °C", "Probe / fridge display", "Daily / Denne", "Reject/relocate stock; call engineer", "Sheet 03", ""),
        ("Cooking / Varenie", "B — survival", "≥ 75 °C core (or 70 °C/2 min)", "Probe core temp", "Each batch / Každá dávka", "Continue cooking; re-check", "Sheet 03", ""),
        ("Hot holding / Udržiavanie tepla", "B — growth/toxin", "≥ 63 °C", "Probe hot-hold", "Every 2 h / Každé 2 h", "Reheat ≥75 °C or discard", "Sheet 03", ""),
        ("Cooling / Chladenie", "B — growth", "≤ 5 °C within 2 h / do 2 h", "Probe + time", "Each batch", "Discard if out of limit", "Sheet 03", ""),
        ("Delivery intake / Príjem tovaru", "B/P — temp & condition", "Chilled ≤5 °C, frozen ≤-18 °C", "Probe + visual", "Each delivery / Každá dodávka", "Reject delivery; record", "Sheet 05", ""),
        ("Allergen control / Kontrola alergénov", "C — cross-contact", "Correct declaration & separation", "Matrix check", "Menu change / Zmena menu", "Re-train; correct menu", "Sheet 02", ""),
    ]
    r = 9
    for row in seed:
        for j, val in enumerate(row, start=1):
            c = ws.cell(r, j, val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.font = f(9)
            if j in (7,):
                c.fill = fill(LBLUE)
        ws.row_dimensions[r].height = 30
        r += 1
    blank_rows(ws, r, 4, 8)

    # PRP checklist block
    r2 = r + 5
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
    h = ws.cell(r2, 1, "Prerequisite Programmes (PRPs) / Predpoklady správnej praxe")
    style_header(h, TEAL); ws.row_dimensions[r2].height = 22
    header_row(ws, r2+1, ["PRP / Predpoklad", "In place?\nZavedené?", "Evidence sheet\nDôkaz", "Owner\nZodpovedný",
                          "Notes\nPoznámky", "", "", ""])
    prps = [
        "Cleaning & disinfection / Čistenie a dezinfekcia (Sheet 04)",
        "Pest control / Ochrana proti škodcom",
        "Supplier approval & traceability / Schvaľovanie dodávateľov (Sheet 05)",
        "Personal hygiene & training / Hygiena a školenie personálu",
        "Waste management / Nakladanie s odpadom",
        "Water safety / Bezpečnosť vody",
    ]
    rr = r2 + 2
    yn = DataValidation(type="list", formula1='"Yes / Áno,No / Nie,N/A"', allow_blank=True)
    ws.add_data_validation(yn)
    for p in prps:
        ws.cell(rr, 1, p).font = f(9)
        ws.cell(rr, 1).alignment = Alignment(vertical="center", wrap_text=True)
        for j in range(1, 6):
            ws.cell(rr, j).border = BORDER
        ws.cell(rr, 2).fill = fill(AMBER)
        yn.add(ws.cell(rr, 2))
        ws.row_dimensions[rr].height = 20
        rr += 1
    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = "A9"


# =============================================================================
# SHEET 2 — Allergen Matrix (14 EU allergens × menu items)
# =============================================================================
def build_allergens(wb):
    ws = wb.create_sheet("02 · Allergens")
    ws.sheet_view.showGridLines = False
    allergens = [
        ("Cereals w/ gluten", "Obilniny s lepkom"),
        ("Crustaceans", "Kôrovce"),
        ("Eggs", "Vajcia"),
        ("Fish", "Ryby"),
        ("Peanuts", "Arašidy"),
        ("Soybeans", "Sója"),
        ("Milk", "Mlieko"),
        ("Tree nuts", "Škrupinové orechy"),
        ("Celery", "Zeler"),
        ("Mustard", "Horčica"),
        ("Sesame", "Sezam"),
        ("Sulphites (SO₂)", "Siričitany (SO₂)"),
        ("Lupin", "Vlčí bôb"),
        ("Molluscs", "Mäkkýše"),
    ]
    set_widths(ws, [28] + [10] * 14 + [22])
    span = 1 + 14 + 1
    title_block(ws, "Allergen Matrix & Menu Declaration",
                "Matica alergénov a označovanie jedálneho lístka",
                "Reg. 1169/2011 — 14 statutory allergens mapped to every menu item",
                "Reg. 1169/2011 — 14 zákonných alergénov pre každé jedlo", span=span)
    biz_field(ws, 3, "Business name", "Názov prevádzky", span=span)
    biz_field(ws, 4, "Menu version / date", "Verzia menu / dátum", span=span)
    note(ws, 6,
         "Mark 'Y' if the dish CONTAINS the allergen, 'T' if there is a risk of cross-contact (traces). Update on every recipe or menu change. Cell turns red when marked Y.",
         "Označte 'Y' ak jedlo OBSAHUJE alergén, 'T' pri riziku krížovej kontaminácie (stopy). Aktualizujte pri každej zmene receptúry alebo menu.",
         span=span)

    # header
    style_header(ws.cell(8, 1, "Menu item\nJedlo"))
    for j, (en, sk) in enumerate(allergens, start=2):
        style_header(ws.cell(8, j, f"{en}\n{sk}"), bg=TEAL, sz=8)
    style_header(ws.cell(8, 16, "Notes\nPoznámky"))
    ws.row_dimensions[8].height = 46

    sample = [
        "Brown soda bread / Hnedý chlieb",
        "Seafood chowder / Rybacia polievka",
        "Caesar salad / Caesar šalát",
        "Beef burger & bun / Hovädzí burger",
        "Veg stir-fry / Zeleninové wok",
        "Cheesecake / Tvarohový koláč",
    ]
    yt = DataValidation(type="list", formula1='"Y,T,-"', allow_blank=True)
    ws.add_data_validation(yt)
    r = 9
    NROWS = 24
    for i in range(NROWS):
        name = sample[i] if i < len(sample) else ""
        nc = ws.cell(r, 1, name)
        nc.border = BORDER; nc.font = f(9)
        nc.alignment = Alignment(vertical="center", wrap_text=True)
        nc.fill = fill(WHITE if name else AMBER)
        for j in range(2, 16):
            c = ws.cell(r, j, "")
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            yt.add(c)
        nt = ws.cell(r, 16, ""); nt.border = BORDER; nt.fill = fill(LGREY)
        ws.row_dimensions[r].height = 20
        r += 1

    # conditional formatting: contains (Y) -> red, traces (T) -> amber
    from openpyxl.formatting.rule import CellIsRule
    rng = f"B9:O{8+NROWS}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Y"'], fill=fill(RED)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"T"'], fill=fill(AMBER)))

    note(ws, r + 1,
         "Customer declaration: 'Full allergen information is available — please ask a member of staff.' Display this and keep this matrix accessible at all times.",
         "Vyhlásenie pre hostí: 'Úplné informácie o alergénoch sú k dispozícii — opýtajte sa personálu.' Zverejnite a maticu majte vždy dostupnú.",
         span=span, bg=GREEN)
    ws.sheet_properties.tabColor = "C00000"
    ws.freeze_panes = "B9"


# =============================================================================
# SHEET 3 — Temperature Log
# =============================================================================
def build_temp(wb):
    ws = wb.create_sheet("03 · Temperature")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 18, 12, 12, 12, 14, 22, 12])
    title_block(ws, "Temperature Monitoring Log", "Záznamník monitorovania teploty",
                "Daily fridge / freezer / hot-hold / cook / delivery temperatures (°C) — HACCP CCP evidence",
                "Denné teploty (°C) — dôkaz HACCP CCP")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Month / Mesiac (MM/YYYY)", "Mesiac")
    note(ws, 6,
         "Targets: Fridge ≤5 °C · Freezer ≤-18 °C · Hot-hold ≥63 °C · Cook core ≥75 °C. 'Pass?' auto-flags OK/CHECK from the reading vs target.",
         "Ciele: Chladnička ≤5 °C · Mrazák ≤-18 °C · Udržiavanie tepla ≥63 °C · Jadro ≥75 °C. 'Vyhovuje?' sa vyhodnotí automaticky.")
    header_row(ws, 7, [
        "Date\nDátum", "Unit / point\nZariadenie", "Type\nTyp", "Reading °C\nNameraná °C",
        "Target °C\nCieľ °C", "Pass?\nVyhovuje?", "Corrective action\nNápravné opatrenie", "Init.\nPodpis",
    ])
    typ = DataValidation(type="list", formula1='"Fridge/Chladnička,Freezer/Mrazák,Hot-hold/Teplo,Cook/Varenie,Delivery/Dodávka"', allow_blank=True)
    ws.add_data_validation(typ)
    NROWS = 31
    r = 8
    for i in range(NROWS):
        for j in range(1, 9):
            c = ws.cell(r, j, "")
            c.border = BORDER
            if i % 2: c.fill = fill(LGREY)
            c.font = f(9)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 1).number_format = "DD/MM/YYYY"
        typ.add(ws.cell(r, 3))
        # Pass? formula: OK if within target depending on direction.
        # Simple robust rule: cold types (Fridge/Freezer/Delivery) reading<=target -> OK; hot types reading>=target -> OK.
        formula = (f'=IF(OR(D{r}="",E{r}=""),"",'
                   f'IF(OR(LEFT(C{r},4)="Frid",LEFT(C{r},4)="Free",LEFT(C{r},4)="Deli"),'
                   f'IF(D{r}<=E{r},"OK","CHECK"),IF(D{r}>=E{r},"OK","CHECK")))')
        ws.cell(r, 6).value = formula
        ws.cell(r, 4).fill = fill(AMBER)
        ws.cell(r, 5).fill = fill(AMBER)
        ws.row_dimensions[r].height = 18
        r += 1
    from openpyxl.formatting.rule import CellIsRule
    rng = f"F8:F{7+NROWS}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"CHECK"'], fill=fill(RED), font=f(9, True, "9C0006")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=fill(GREEN)))
    ws.sheet_properties.tabColor = "2E75B6"
    ws.freeze_panes = "A8"


# =============================================================================
# SHEET 4 — Cleaning & Sanitation Schedule
# =============================================================================
def build_cleaning(wb):
    ws = wb.create_sheet("04 · Cleaning")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 16, 18, 16, 12, 12, 12, 18])
    title_block(ws, "Cleaning & Sanitation Schedule", "Harmonogram čistenia a sanitácie",
                "Daily / weekly / periodic cleaning with sign-off — good-hygiene-practice evidence",
                "Denné / týždenné / pravidelné čistenie s podpisom")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Week commencing (DD/MM/YYYY)", "Týždeň od")
    note(ws, 6,
         "List each item/area, who cleans it, the product & method, and tick the day completed with initials. 'Done' counts completed days in the week.",
         "Uveďte oblasť, kto čistí, prostriedok a metódu; označte deň s podpisom. 'Hotovo' spočíta dni v týždni.")
    header_row(ws, 7, [
        "Item / area\nOblasť", "Frequency\nČetnosť", "Product & method\nProstriedok a metóda",
        "Responsible\nZodpovedný", "Mon–Sun ticks\nPo–Ne", "", "", "Done /7\nHotovo",
    ])
    # merge the Mon-Sun span visually
    ws.merge_cells("E7:G7")
    freq = DataValidation(type="list", formula1='"Daily/Denne,Weekly/Týždenne,Periodic/Pravidelne"', allow_blank=True)
    ws.add_data_validation(freq)
    seed = [
        ("Food prep surfaces / Pracovné plochy", "Daily/Denne", "Sanitiser, BS EN 1276", ""),
        ("Floors (kitchen) / Podlahy (kuchyňa)", "Daily/Denne", "Degreaser + mop", ""),
        ("Fridges & seals / Chladničky a tesnenia", "Weekly/Týždenne", "Sanitiser wipe", ""),
        ("Extraction & filters / Odsávanie a filtre", "Periodic/Pravidelne", "Degrease / specialist", ""),
        ("Toilets / Toalety", "Daily/Denne", "Disinfectant", ""),
        ("Bins & waste area / Odpadové nádoby", "Daily/Denne", "Disinfect + reline", ""),
        ("Ice machine / Výrobník ľadu", "Periodic/Pravidelne", "Descale + sanitise", ""),
    ]
    r = 8
    NROWS = 16
    for i in range(NROWS):
        row = seed[i] if i < len(seed) else ("", "", "", "")
        ws.cell(r, 1, row[0]).fill = fill(WHITE if row[0] else AMBER)
        ws.cell(r, 2, row[1]); freq.add(ws.cell(r, 2))
        ws.cell(r, 3, row[2]); ws.cell(r, 4, row[3])
        for j in (1, 2, 3, 4):
            ws.cell(r, j).border = BORDER; ws.cell(r, j).font = f(9)
            ws.cell(r, j).alignment = Alignment(vertical="center", wrap_text=True)
        for j in (5, 6, 7):
            ws.cell(r, j).border = BORDER
        ws.cell(r, 8).value = f'=COUNTA(E{r}:G{r})'
        ws.cell(r, 8).border = BORDER
        ws.cell(r, 8).alignment = Alignment(horizontal="center", vertical="center")
        if i % 2:
            for j in range(1,9):
                if not ws.cell(r,j).fill.fgColor.rgb or ws.cell(r,j).fill.fgColor.rgb=="00000000":
                    ws.cell(r,j).fill = fill(LGREY)
        ws.row_dimensions[r].height = 22
        r += 1
    ws.sheet_properties.tabColor = "548235"
    ws.freeze_panes = "A8"


# =============================================================================
# SHEET 5 — Supplier & Delivery Traceability Log
# =============================================================================
def build_traceability(wb):
    ws = wb.create_sheet("05 · Traceability")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 22, 22, 14, 14, 12, 14, 18])
    title_block(ws, "Supplier & Delivery Traceability Log", "Záznam vysledovateľnosti dodávateľov a dodávok",
                "Incoming-goods checks + batch/lot 'one step back' traceability — recall readiness",
                "Príjem tovaru a šarže — pripravenosť na stiahnutie produktu")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    note(ws, 6,
         "Record every delivery: supplier, product, batch/lot, quantity, temperature on arrival (chilled ≤5 °C / frozen ≤-18 °C) and accept/reject. This is your recall trail.",
         "Zaznamenajte každú dodávku: dodávateľ, produkt, šarža, množstvo, teplota pri príjme a prijatie/odmietnutie. Toto je vaša stopa pre stiahnutie.")
    header_row(ws, 7, [
        "Date\nDátum", "Supplier\nDodávateľ", "Product\nProdukt", "Batch / lot\nŠarža",
        "Qty (kg/units)\nMnožstvo", "Temp °C\nTeplota", "Accept?\nPrijaté?", "Checked by\nSkontroloval",
    ])
    acc = DataValidation(type="list", formula1='"Accept/Prijaté,Reject/Odmietnuté"', allow_blank=True)
    ws.add_data_validation(acc)
    NROWS = 26
    r = 8
    for i in range(NROWS):
        for j in range(1, 9):
            c = ws.cell(r, j, "")
            c.border = BORDER; c.font = f(9)
            if i % 2: c.fill = fill(LGREY)
            c.alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(r, 1).number_format = "DD/MM/YYYY"
        ws.cell(r, 5).number_format = "#,##0.00"  # comma thousands (EU)
        acc.add(ws.cell(r, 7))
        ws.row_dimensions[r].height = 18
        r += 1
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(f"G8:G{7+NROWS}",
        CellIsRule(operator="containsText", formula=['"Reject"'], fill=fill(RED)))
    ws.sheet_properties.tabColor = "BF8F00"
    ws.freeze_panes = "A8"


# =============================================================================
# SHEET 6 — H&S Risk Assessment & Safety Statement
# =============================================================================
def build_hs(wb):
    ws = wb.create_sheet("06 · Health & Safety")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 22, 10, 10, 12, 26, 14, 14])
    title_block(ws, "H&S Risk Assessment & Safety Statement", "Posúdenie rizík BOZP a bezpečnostné vyhlásenie",
                "Safety, Health & Welfare at Work Act 2005 (S.19/S.20) — risk assessment, statement & accident log",
                "Zákon o BOZP 2005 — posúdenie rizík, vyhlásenie a kniha úrazov")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Safety Statement prepared by", "Vyhlásenie vypracoval")
    biz_field(ws, 5, "Date / review date", "Dátum / revízia")
    note(ws, 7,
         "Risk rating = Likelihood (1–5) × Severity (1–5). 1–6 Low · 8–12 Medium · 15–25 High. Higher ratings need controls before work continues.",
         "Miera rizika = Pravdepodobnosť (1–5) × Závažnosť (1–5). 1–6 Nízke · 8–12 Stredné · 15–25 Vysoké.")
    header_row(ws, 8, [
        "Hazard\nNebezpečenstvo", "Who's at risk\nKoho ohrozuje", "L\nP", "S\nZ", "Risk (L×S)\nRiziko",
        "Controls in place\nOpatrenia", "Action by\nVykoná", "Done\nHotovo",
    ])
    seed = [
        ("Slips/trips (wet floors) / Pošmyknutie", "Staff, customers", 3, 4, "Wet-floor signs, non-slip mats, spill routine"),
        ("Burns/scalds (hot oil, steam) / Popáleniny", "Kitchen staff", 3, 4, "Training, gloves, no overfilling fryers"),
        ("Knives / sharps / Porezanie", "Kitchen staff", 3, 3, "Cut gloves, correct technique, sheaths"),
        ("Manual handling / Manipulácia s bremenami", "Staff", 3, 3, "Training, trolleys, weight limits"),
        ("Electrical (appliances) / Elektrika", "All", 2, 4, "PAT testing, RCD, no daisy-chaining"),
        ("Fire (kitchen) / Požiar", "All", 2, 5, "See Sheet 07 fire register & drills"),
        ("Chemicals (cleaning) / Chemikálie", "Staff", 2, 3, "SDS sheets, COSHH, ventilation"),
    ]
    r = 9
    NROWS = 14
    for i in range(NROWS):
        row = seed[i] if i < len(seed) else ("", "", "", "", "")
        ws.cell(r, 1, row[0]); ws.cell(r, 2, row[1])
        ws.cell(r, 3, row[2] if row[2] else ""); ws.cell(r, 4, row[3] if row[3] else "")
        ws.cell(r, 5).value = f'=IF(OR(C{r}="",D{r}=""),"",C{r}*D{r})'
        ws.cell(r, 6, row[4])
        for j in range(1, 9):
            c = ws.cell(r, j); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment(vertical="center", wrap_text=True,
                                    horizontal="center" if j in (3,4,5,8) else "left")
        ws.cell(r, 1).fill = fill(WHITE if row[0] else AMBER)
        ws.cell(r, 3).fill = fill(AMBER); ws.cell(r, 4).fill = fill(AMBER)
        if i % 2 == 1:
            for j in (2, 6, 7):
                ws.cell(r, j).fill = fill(LGREY)
        ws.row_dimensions[r].height = 26
        r += 1
    from openpyxl.formatting.rule import CellIsRule
    rng = f"E9:E{8+NROWS}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["15"], fill=fill(RED)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["8", "12"], fill=fill(AMBER)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["1", "6"], fill=fill(GREEN)))

    # Accident log block
    r2 = r + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
    h = ws.cell(r2, 1, "Accident / Incident Log — kniha úrazov a incidentov")
    style_header(h, TEAL); ws.row_dimensions[r2].height = 22
    header_row(ws, r2+1, ["Date\nDátum", "Person\nOsoba", "What happened\nČo sa stalo", "Injury\nZranenie",
                          "First aid / action\nPrvá pomoc", "Reported (HSA?)\nNahlásené", "By whom\nKto", "Closed\nUzavreté"])
    rr = r2 + 2
    for i in range(6):
        for j in range(1, 9):
            c = ws.cell(rr, j, ""); c.border = BORDER; c.font = f(9)
            if i % 2: c.fill = fill(LGREY)
        ws.cell(rr, 1).number_format = "DD/MM/YYYY"
        ws.row_dimensions[rr].height = 20
        rr += 1
    ws.sheet_properties.tabColor = "C55A11"
    ws.freeze_panes = "A9"


# =============================================================================
# SHEET 7 — Fire Safety Register & Checks Log
# =============================================================================
def build_fire(wb):
    ws = wb.create_sheet("07 · Fire Safety")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 16, 16, 16, 16, 14, 16, 16])
    title_block(ws, "Fire Safety Register & Checks Log", "Register požiarnej ochrany a záznam kontrol",
                "Fire Services Acts 1981 & 2003 — equipment register, routine checks, drills & emergency lighting",
                "Zákon o požiarnej ochrane — register zariadení, kontroly, cvičenia a núdzové osvetlenie")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Responsible person", "Zodpovedná osoba")

    # Register block
    note(ws, 6,
         "Fire-equipment register: list each item, location, last service and next-due date. 'Status' flags OVERDUE when next-due has passed today (30/05/2026).",
         "Register zariadení: položka, umiestnenie, posledný servis a ďalší termín. 'Stav' upozorní na PO TERMÍNE.")
    header_row(ws, 7, [
        "Equipment / item\nZariadenie", "Location\nUmiestnenie", "Type\nTyp", "Last service\nPosledný servis",
        "Next due\nĎalší termín", "Status\nStav", "Provider\nDodávateľ", "Notes\nPoznámky",
    ])
    seed = [
        ("Fire extinguisher (CO₂) / Hasiaci prístroj", "Kitchen", "Annual"),
        ("Fire extinguisher (foam) / Hasiaci prístroj", "Front-of-house", "Annual"),
        ("Fire blanket / Hasiaca deka", "Kitchen", "Visual"),
        ("Fire alarm panel / Požiarny panel", "Hallway", "Quarterly"),
        ("Emergency lighting / Núdzové osvetlenie", "Exits", "Monthly"),
        ("Fire exit / route / Únikový východ", "All", "Weekly"),
    ]
    r = 8
    for i in range(12):
        row = seed[i] if i < len(seed) else ("", "", "")
        ws.cell(r, 1, row[0]); ws.cell(r, 2, row[1]); ws.cell(r, 3, row[2])
        ws.cell(r, 6).value = (f'=IF(E{r}="","",IF(E{r}<DATE(2026,5,30),"OVERDUE",'
                               f'IF(E{r}<=DATE(2026,6,29),"DUE SOON","OK")))')
        for j in range(1, 9):
            c = ws.cell(r, j); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 and j not in (4, 5): c.fill = fill(LGREY)
        ws.cell(r, 1).fill = fill(WHITE if row[0] else AMBER)
        ws.cell(r, 4).fill = fill(AMBER); ws.cell(r, 4).number_format = "DD/MM/YYYY"
        ws.cell(r, 5).fill = fill(AMBER); ws.cell(r, 5).number_format = "DD/MM/YYYY"
        ws.row_dimensions[r].height = 22
        r += 1
    from openpyxl.formatting.rule import CellIsRule
    rng = f"F8:F{r-1}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OVERDUE"'], fill=fill(RED), font=f(9, True, "9C0006")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=fill(AMBER)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=fill(GREEN)))

    # Drill log
    r2 = r + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
    h = ws.cell(r2, 1, "Fire Drill & Evacuation Log — záznam cvičení a evakuácie")
    style_header(h, TEAL); ws.row_dimensions[r2].height = 22
    header_row(ws, r2+1, ["Date\nDátum", "Type (drill/real)\nTyp", "Evac time (min)\nČas evakuácie",
                          "Persons\nOsoby", "Issues found\nNedostatky", "Action\nOpatrenie", "By whom\nKto", "Signed\nPodpis"])
    rr = r2 + 2
    for i in range(6):
        for j in range(1, 9):
            c = ws.cell(rr, j, ""); c.border = BORDER; c.font = f(9)
            if i % 2: c.fill = fill(LGREY)
        ws.cell(rr, 1).number_format = "DD/MM/YYYY"
        ws.cell(rr, 3).number_format = "0.0"  # minutes, metric/decimal
        ws.row_dimensions[rr].height = 20
        rr += 1
    ws.sheet_properties.tabColor = "C00000"
    ws.freeze_panes = "A8"


def build_workbook():
    wb = Workbook()
    build_cover(wb)
    build_haccp(wb)
    build_allergens(wb)
    build_temp(wb)
    build_cleaning(wb)
    build_traceability(wb)
    build_hs(wb)
    build_fire(wb)
    wb.properties.title = "Café / Restaurant Compliance Pack (EN/SK)"
    wb.properties.creator = "ASSET-FORGE"
    wb.properties.subject = "Hospitality EHO compliance — Reg 852/2004, 1169/2011, Fire, SHWWA 2005"
    return wb


def add_demo_watermark(wb):
    """Watermarked, locked preview for the listing image.

    Non-destructive: prepend a dedicated DEMO notice sheet and lock every sheet
    read-only. Do NOT use insert_rows() — openpyxl shifts cell values but not
    merged-cell / conditional-format / data-validation ranges, which corrupts
    merge-heavy sheets (Excel: "we found a problem with content").
    """
    notice = wb.create_sheet("DEMO Preview · Ukážka", 0)
    notice.sheet_view.showGridLines = False
    set_widths(notice, [4, 28, 28, 28, 28])
    notice.merge_cells("A1:E1")
    b = notice.cell(1, 1, "DEMO — PREVIEW ONLY · NOT FOR RESALE")
    b.font = f(16, True, "FFFFFF"); b.fill = fill("C00000")
    b.alignment = Alignment(horizontal="center", vertical="center")
    notice.row_dimensions[1].height = 36
    notice.merge_cells("A2:E2")
    b2 = notice.cell(2, 1, "Ukážka — len na náhľad · nie na ďalší predaj")
    b2.font = f(12, True, "C00000"); b2.fill = fill(SAND)
    b2.alignment = Alignment(horizontal="center", vertical="center")
    notice.row_dimensions[2].height = 24
    notice.merge_cells("A4:E4")
    m = notice.cell(4, 1, "Browse the tabs to preview the pack. Buy the full version to unlock editing.")
    m.font = f(11, False, NAVY); m.alignment = Alignment(horizontal="center", vertical="center")
    notice.merge_cells("A5:E5")
    m2 = notice.cell(5, 1, "Prehliadnite si hárky. Kúpou plnej verzie odomknete úpravy.")
    m2.font = f(11, False, NAVY, italic=True); m2.alignment = Alignment(horizontal="center", vertical="center")
    notice.sheet_properties.tabColor = "C00000"
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = "demo"
    wb.properties.title = "Café / Restaurant Compliance Pack — DEMO (EN/SK)"
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    full = os.path.join(PRODUCTS, "P1_Cafe_Restaurant_Compliance_Pack.xlsx")
    wb.save(full)
    print(f"✓ Built flagship: {full}  ({len(wb.sheetnames)} sheets)")

    demo = add_demo_watermark(build_workbook())
    demo_path = os.path.join(PRODUCTS, "P1_DEMO_Cafe_Restaurant_Compliance_Pack.xlsx")
    demo.save(demo_path)
    print(f"✓ Built watermarked demo: {demo_path}")
    print("  Sheets:", " | ".join(wb.sheetnames))
