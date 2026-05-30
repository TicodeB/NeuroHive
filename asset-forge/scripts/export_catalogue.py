#!/usr/bin/env python3
"""
export_catalogue.py — ASSET-FORGE Phase 6 deliverable.

Flat-exports the full intelligence.db catalogue to
`deliverables/asset_catalogue.xlsx` for skim-reading (Section [4]/[12]).

One workbook, several sheets:
  - Overview        : counts, tier mix, generation stamp
  - Asset_Map       : the full 442-row asset x business-type matrix (raw axes,
                      score, tier, buyer, evidence) — the heart of the export
  - Digital_Assets  : the 54-asset master list
  - Universal_Core  : assets that are MUST across >=3 business types
  - MUST_Haves      : every MUST row (v_must_haves)
  - Pain_Points     : owner-voice pain catalogue
  - Tier_Summary    : tier x vertical pivot

EU formatting: dates DD/MM/YYYY, metric units already baked into the data.
No external network needed; depends only on `openpyxl` (pip install openpyxl).

Usage:
    python3 scripts/export_catalogue.py
    python3 scripts/export_catalogue.py --db path/to.db --out path/to.xlsx
"""
from __future__ import annotations

import argparse
import datetime as _dt
import os
import sqlite3
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit(
        "openpyxl is required for the xlsx export.\n"
        "  pip install openpyxl\n"
        "Then re-run: python3 scripts/export_catalogue.py"
    )

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)  # asset-forge/
DEFAULT_DB = os.path.join(ROOT, "intelligence.db")
DEFAULT_OUT = os.path.join(ROOT, "deliverables", "asset_catalogue.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TIER_FILL = {
    "MUST": PatternFill("solid", fgColor="C6EFCE"),
    "SHOULD": PatternFill("solid", fgColor="FFEB9C"),
    "COULD": PatternFill("solid", fgColor="FCE4D6"),
    "WON'T": PatternFill("solid", fgColor="F2F2F2"),
}


def _q(cur: sqlite3.Cursor, sql: str) -> tuple[list[str], list[tuple]]:
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    return cols, cur.fetchall()


def _write_sheet(ws, headers, rows, tier_col: str | None = None) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    tier_idx = headers.index(tier_col) if tier_col and tier_col in headers else None
    for r in rows:
        ws.append(list(r))
        if tier_idx is not None:
            val = r[tier_idx]
            fill = TIER_FILL.get(val)
            if fill:
                ws.cell(row=ws.max_row, column=tier_idx + 1).fill = fill
    # auto width (capped)
    for i, _ in enumerate(headers, start=1):
        letter = get_column_letter(i)
        longest = max(
            [len(str(headers[i - 1]))] + [len(str(row[i - 1])) for row in rows]
            if rows else [len(str(headers[i - 1]))]
        )
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 60)
    ws.auto_filter.ref = ws.dimensions


def build(db_path: str, out_path: str) -> None:
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    wb = Workbook()

    # --- Overview ---
    ov = wb.active
    ov.title = "Overview"
    stamp = _dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    n_bt = cur.execute("SELECT COUNT(*) FROM business_types").fetchone()[0]
    n_assets = cur.execute("SELECT COUNT(*) FROM digital_assets").fetchone()[0]
    n_map = cur.execute("SELECT COUNT(*) FROM asset_map").fetchone()[0]
    tiers = dict(cur.execute("SELECT tier, COUNT(*) FROM asset_map GROUP BY tier").fetchall())
    meta = [
        ("ASSET-FORGE — Asset Catalogue", ""),
        ("Generated", stamp),
        ("Source DB", os.path.basename(db_path)),
        ("", ""),
        ("Verticals", cur.execute("SELECT COUNT(*) FROM verticals").fetchone()[0]),
        ("Business types", n_bt),
        ("Digital assets (deduped by function)", n_assets),
        ("Asset x business-type rows", n_map),
        ("", ""),
        ("Tier — MUST", tiers.get("MUST", 0)),
        ("Tier — SHOULD", tiers.get("SHOULD", 0)),
        ("Tier — COULD", tiers.get("COULD", 0)),
        ("Tier — WON'T", tiers.get("WON'T", 0)),
    ]
    for k, v in meta:
        ov.append([k, v])
    ov["A1"].font = Font(bold=True, size=14)
    ov.column_dimensions["A"].width = 40
    ov.column_dimensions["B"].width = 28

    # --- Asset_Map (the flat export) ---
    headers, rows = _q(
        cur,
        """
        SELECT v.name        AS vertical,
               bt.name       AS business_type,
               bt.work_context AS work_context,
               d.name        AS department,
               da.name       AS asset,
               da.asset_type AS asset_type,
               am.buyer      AS buyer,
               am.legal      AS legal,
               am.revenue    AS revenue,
               am.pain       AS pain,
               am.frequency  AS frequency,
               am.score      AS score,
               am.tier       AS tier,
               am.evidence_url AS evidence_url,
               am.notes      AS notes
        FROM asset_map am
        JOIN business_types bt ON bt.id = am.business_type_id
        JOIN verticals v       ON v.id  = bt.vertical_id
        JOIN departments d     ON d.id  = am.department_id
        JOIN digital_assets da ON da.id = am.asset_id
        ORDER BY v.id, bt.id, am.score DESC, da.name
        """,
    )
    _write_sheet(wb.create_sheet("Asset_Map"), headers, rows, tier_col="tier")

    # --- Digital_Assets ---
    headers, rows = _q(
        cur, "SELECT id, name, asset_type, description FROM digital_assets ORDER BY id"
    )
    _write_sheet(wb.create_sheet("Digital_Assets"), headers, rows)

    # --- Universal_Core ---
    headers, rows = _q(
        cur,
        "SELECT asset_id, asset, must_in_n_business_types FROM v_universal_core",
    )
    _write_sheet(wb.create_sheet("Universal_Core"), headers, rows)

    # --- MUST_Haves ---
    headers, rows = _q(
        cur,
        "SELECT vertical, business_type, asset, buyer, score, tier, evidence_url "
        "FROM v_must_haves",
    )
    _write_sheet(wb.create_sheet("MUST_Haves"), headers, rows, tier_col="tier")

    # --- Pain_Points ---
    headers, rows = _q(
        cur,
        """
        SELECT v.name AS vertical, bt.name AS business_type,
               pp.description, pp.severity, pp.source_url
        FROM pain_points pp
        JOIN business_types bt ON bt.id = pp.business_type_id
        JOIN verticals v       ON v.id  = bt.vertical_id
        ORDER BY v.id, bt.id
        """,
    )
    _write_sheet(wb.create_sheet("Pain_Points"), headers, rows)

    # --- Tier_Summary (tier x vertical) ---
    headers, rows = _q(
        cur,
        """
        SELECT v.name AS vertical,
               SUM(am.tier='MUST')   AS must,
               SUM(am.tier='SHOULD') AS should,
               SUM(am.tier='COULD')  AS could,
               COUNT(*)              AS total
        FROM asset_map am
        JOIN business_types bt ON bt.id = am.business_type_id
        JOIN verticals v       ON v.id  = bt.vertical_id
        GROUP BY v.id ORDER BY v.id
        """,
    )
    _write_sheet(wb.create_sheet("Tier_Summary"), headers, rows)

    con.close()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  {n_map} asset_map rows | {n_assets} assets | {n_bt} business types")
    print(f"  tiers: MUST {tiers.get('MUST',0)} | SHOULD {tiers.get('SHOULD',0)} "
          f"| COULD {tiers.get('COULD',0)}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Export intelligence.db to xlsx")
    ap.add_argument("--db", default=DEFAULT_DB)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()
    if not os.path.exists(args.db):
        sys.exit(f"DB not found: {args.db}")
    build(args.db, args.out)


if __name__ == "__main__":
    main()
