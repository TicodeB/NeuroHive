#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 13b — Premium Pack pilot (re-skin proof)
============================================================
P2 re-skinned: **Hospitality Operations & GP Bundle**, but as the new premium,
SLOVAK-ONLY edition built on `design_system.py`. This is the visible proof that
we've beaten the "looks like the 90s" problem (Samuel, 30/05).

What changed vs the old build_p2_operations_bundle.py:
  · single language (clean Slovak) — no EN|SK cramming
  · the design system: gridlines off, hairline tables, KPI tiles, accent titles
  · NEW "02 · Prehľad" dashboard tab — KPI tiles + formula-driven INSIGHTS that
    pull live from the operational sheets ("reads like a tiny analyst")
  · NEW "01 · Denný plán" daily planner (Samuel: method + planner built in)

Slovak text here is DRAFT-grade for the build proof — route through a native
editor before any public listing (standing rule).

Run:  python3 scripts/build_p2_sk_premium.py
Out:  products/P2_SK_Hospitality_Premium.xlsx
"""
from __future__ import annotations
import os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from design_system import Theme, DS

PRODUCTS = os.path.normpath(os.path.join(HERE, "..", "products"))
os.makedirs(PRODUCTS, exist_ok=True)
VERSION, BUILD_DATE = "v1.0", "31/05/2026"

# Hospitality theme: warm teal accent over the default navy/blue
ds = DS(Theme(primary="2D6CDF", accent="15A38C"))

# Sheet names (kept ≤31 chars; quoted in cross-sheet formulas)
S_METHOD = "00 · Metóda"
S_PLAN   = "01 · Denný plán"
S_DASH   = "02 · Prehľad"
S_CASH   = "03 · Cash flow"
S_GP     = "04 · Marža"
S_STOCK  = "05 · Zásoby"
S_ROTA   = "06 · Zmeny"
S_TILL   = "07 · Tržby"
S_TRAIN  = "08 · Školenia"

def ref(sheet, cell):  # safe cross-sheet reference
    return f"'{sheet}'!{cell}"

# Anchor cells the dashboard reads (must match the layouts below exactly)
A_REV    = ref(S_CASH, "O13")   # ročné tržby
A_GP     = ref(S_CASH, "O18")   # hrubá marža €/rok
A_NP     = ref(S_CASH, "O26")   # čistý zisk €/rok
A_CLOSE  = ref(S_CASH, "N29")   # koncová hotovosť (dec)
A_AVGGP  = ref(S_GP,   "H30")   # priemerná reálna marža %
A_LOWGP  = ref(S_GP,   "H31")   # počet položiek pod cieľom
A_WASTE  = ref(S_STOCK,"I27")   # hodnota strát €
A_LABPCT = ref(S_ROTA, "I31")   # podiel miezd %
A_TILLVAR= ref(S_TILL, "H40")   # rozdiel pokladne (rok)


# ----------------------------------------------------------------- helpers
def label_input(ws, row, label, col_label=2, col_in=3, in_span=2, fmt=None):
    c = ws.cell(row, col_label, label)
    c.font = ds.font(10, bold=True, color=ds.t.ink)
    c.alignment = Alignment("left", "center")
    last = col_in + in_span - 1
    if in_span > 1:
        ws.merge_cells(start_row=row, start_column=col_in, end_row=row, end_column=last)
    ds.input_cell(ws, row, col_in, fmt)
    ws.row_dimensions[row].height = 20


# ----------------------------------------------------------------- 00 Metóda
def build_method(wb):
    ws = wb.create_sheet(S_METHOD)
    ds.canvas(ws, [3, 52, 30, 14, 14, 14, 14], tab=ds.t.ink)
    ds.title(ws, "Prevádzka a marža pre gastro",
             "Balík nástrojov, ktorý vám ukáže maržu, mzdy a hotovosť skôr, než zabolia")
    r = 5
    r = ds.section(ws, r, "Čo balík obsahuje")
    sheets = [
        ("01 · Denný plán", "priority dňa, časový rozvrh, otvorenie/zatvorenie"),
        ("02 · Prehľad", "automatický dashboard — KPI a postrehy z vašich čísel"),
        ("03 · Cash flow", "12-mesačný cash flow a výkaz ziskov a strát"),
        ("04 · Marža", "kalkulácia nákladov a hrubej marže receptúr"),
        ("05 · Zásoby", "sledovač zásob a hodnoty strát"),
        ("06 · Zmeny", "plánovač zmien a podielu mzdových nákladov"),
        ("07 · Tržby", "denné tržby a uzávierka pokladne"),
        ("08 · Školenia", "matica školení a zaškolenia — dôkaz pred zmenou"),
    ]
    for name, desc in sheets:
        a = ws.cell(r, 2, "  " + name); a.font = ds.font(10.5, bold=True, color=ds.t.primary)
        a.alignment = Alignment("left", "center", indent=1)
        b = ws.cell(r, 3, desc); b.font = ds.font(10, color=ds.t.muted)
        b.alignment = Alignment("left", "center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 20; r += 1
    r += 1
    r = ds.section(ws, r, "Metóda — ako to používať každý deň")
    steps = [
        "RÁNO (5 min): otvorte „01 · Denný plán“ — zapíšte 3 priority a prejdite otvorenie.",
        "POČAS DŇA: zapisujte tržby do „07 · Tržby“ a straty do „05 · Zásoby“.",
        "VEČER (5 min): uzávierka pokladne v „07 · Tržby“, kontrola rozdielu.",
        "TÝŽDENNE: skontrolujte „06 · Zmeny“ (podiel miezd) a „04 · Marža“.",
        "MESAČNE: vyplňte „03 · Cash flow“ a pozrite „02 · Prehľad“ — uvidíte celý obraz.",
    ]
    for s in steps:
        c = ws.cell(r, 2, "•  " + s); c.font = ds.font(10, color=ds.t.ink)
        c.alignment = Alignment("left", "center", wrap_text=True, indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22; r += 1
    r += 1
    r = ds.section(ws, r, "Farby")
    r = ds.note(ws, r, "Oranžová = vyplníte vy    ·    Modrá = vypočíta sa automaticky    ·    Biela/sivá = záznamy", tone="info")
    r += 1
    ds.footer(ws, r, f"ASSET-FORGE · {VERSION} · {BUILD_DATE} · EU (Írsko) · "
                     "Šablóna — nie je finančné ani právne poradenstvo.")


# ----------------------------------------------------------------- 01 Denný plán
def build_planner(wb):
    ws = wb.create_sheet(S_PLAN)
    ds.canvas(ws, [3, 12, 44, 16, 12], tab=ds.t.accent)
    ds.title(ws, "Denný plán", "Začnite deň s tromi prioritami a jasným rozvrhom")
    label_input(ws, 5, "Dátum", col_in=3, in_span=1, fmt=ds.t.DATE)
    label_input(ws, 6, "Deň", col_in=3, in_span=2)
    r = 8
    r = ds.section(ws, r, "Dnešné 3 priority", span=4)
    for i in range(1, 4):
        c = ws.cell(r, 2, str(i)); c.font = ds.font(11, bold=True, color=ds.t.primary)
        c.alignment = Alignment("center", "center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ds.input_cell(ws, r, 3); ws.row_dimensions[r].height = 22; r += 1
    r += 1
    r = ds.section(ws, r, "Rozvrh dňa", span=4)
    pr = DataValidation(type="list", formula1='"Vysoká,Stredná,Nízka"', allow_blank=True)
    done = DataValidation(type="list", formula1='"Áno,Nie"', allow_blank=True)
    ws.add_data_validation(pr); ws.add_data_validation(done)
    r = ds.thead(ws, r, ["Čas", "Úloha", "Priorita", "Hotovo"])
    slots = [f"{h:02d}:00" for h in range(7, 23)]
    for i, slot in enumerate(slots):
        ds.trow(ws, r, 4, zebra_on=(i % 2 == 1))
        t = ws.cell(r, 2, slot); t.alignment = Alignment("center", "center")
        t.font = ds.font(9.5, bold=True, color=ds.t.muted)
        ds.input_cell(ws, r, 3); ws.cell(r, 3).alignment = Alignment("left", "center", indent=1)
        ds.input_cell(ws, r, 4); pr.add(ws.cell(r, 4))
        ds.input_cell(ws, r, 5); done.add(ws.cell(r, 5))
        r += 1
    r += 1
    r = ds.section(ws, r, "Otvorenie / Zatvorenie", span=4)
    checks = ["Otvorenie: teploty chladničiek skontrolované",
              "Otvorenie: hotovosť v pokladni spočítaná",
              "Zatvorenie: tržby zapísané do hárku 07",
              "Zatvorenie: straty zapísané do hárku 05",
              "Zatvorenie: spotrebiče vypnuté, prevádzka zabezpečená"]
    for i, chk in enumerate(checks):
        ds.trow(ws, r, 4, zebra_on=(i % 2 == 1))
        c = ws.cell(r, 2, "  " + chk); c.alignment = Alignment("left", "center", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ds.input_cell(ws, r, 5); done.add(ws.cell(r, 5))
        r += 1


# ----------------------------------------------------------------- 02 Prehľad
def build_dashboard(wb):
    ws = wb.create_sheet(S_DASH)
    ds.canvas(ws, [3] + [13] * 11, tab=ds.t.primary)
    ds.title(ws, "Prehľad", "Automatický dashboard — ťahá živé čísla z ostatných hárkov")
    # KPI tiles: 3 per row, each 3 cols wide (B-D, F-H, J-L), gap cols E,I
    tiles = [
        ("Ročné tržby", A_REV, ds.t.EUR0, ds.t.primary),
        ("Hrubá marža (rok)", A_GP, ds.t.EUR0, ds.t.accent),
        ("Čistý zisk (rok)", A_NP, ds.t.EUR0, ds.t.ink),
        ("Koncová hotovosť", A_CLOSE, ds.t.EUR0, ds.t.primary),
        ("Priemerná marža", A_AVGGP, ds.t.PCT, ds.t.accent),
        ("Podiel miezd", A_LABPCT, ds.t.PCT, ds.t.ink),
    ]
    positions = [2, 6, 10]  # left columns for the 3 tiles per row
    top = 5
    for idx, (label, anchor, fmt, accent) in enumerate(tiles):
        row_block = top + (idx // 3) * 5
        left = positions[idx % 3]
        val = ds.kpi(ws, row_block, left, label, f"=IFERROR({anchor},0)", fmt=fmt, accent=accent)
        # RAG accents on the two risk KPIs
        if label == "Čistý zisk (rok)":
            ws.conditional_formatting.add(val.coordinate,
                CellIsRule(operator="lessThan", formula=["0"], font=ds.font(20, True, ds.t.bad)))
        if label == "Podiel miezd":
            ws.conditional_formatting.add(val.coordinate,
                CellIsRule(operator="greaterThan", formula=["0.35"], font=ds.font(20, True, ds.t.bad)))
    r = top + 10
    r = ds.section(ws, r, "Postrehy", span=11)
    insights = [
        f'=IFERROR("Čistá marža: "&TEXT({A_NP}/{A_REV},"0.0%")&IF({A_NP}/{A_REV}>=0.1," — zdravé."," — pozor, je nízka."),"Čistá marža: zatiaľ bez dát.")',
        f'=IFERROR("Podiel miezd: "&TEXT({A_LABPCT},"0.0%")&IF({A_LABPCT}>0.35," — NAD cieľom 35 %, skontrolujte zmeny."," — v poriadku."),"Podiel miezd: zatiaľ bez dát.")',
        f'=IFERROR("Hodnota strát na zásobách: "&TEXT({A_WASTE},"#,##0.00 €")&".","Straty: zatiaľ bez dát.")',
        f'=IFERROR("Položky pod cieľovou maržou: "&TEXT({A_LOWGP},"0")&" — zvážte cenu alebo recept.","Marža: zatiaľ bez dát.")',
        f'=IFERROR("Rozdiel v pokladni (rok): "&TEXT({A_TILLVAR},"#,##0.00 €")&IF(ABS({A_TILLVAR})>1," — preverte uzávierky."," — sedí."),"Pokladňa: zatiaľ bez dát.")',
    ]
    for i, formula in enumerate(insights):
        ds.trow(ws, r, 11, zebra_on=(i % 2 == 1), align="left")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        c = ws.cell(r, 2, formula)
        c.font = ds.font(10.5, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1)
        ws.row_dimensions[r].height = 24; r += 1
    r += 1
    ds.note(ws, r, "Tieto čísla sa aktualizujú automaticky, keď vyplníte hárky 03–07.", span=11, tone="info")


# ----------------------------------------------------------------- 03 Cash flow
def build_cash(wb):
    ws = wb.create_sheet(S_CASH)
    ds.canvas(ws, [26] + [9.5] * 12 + [12], tab=ds.t.primary)
    ds.title(ws, "Cash flow a výkaz ziskov a strát",
             "12-mesačný prehľad hotovosti a mesačný zisk/strata", span=14)
    label_input(ws, 5, "Názov prevádzky", col_in=3, in_span=3)
    label_input(ws, 6, "Rok", col_in=3, in_span=1)
    ds.note(ws, 7, "Vyplňte oranžové bunky. Súčty, marža, zisk a koncová hotovosť sa počítajú automaticky.", span=14)
    months = ["Jan", "Feb", "Mar", "Apr", "Máj", "Jún", "Júl", "Aug", "Sep", "Okt", "Nov", "Dec"]
    ds.thead(ws, 8, ["Položka"] + months + ["Rok"])

    def section(r, text):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=15)
        c = ws.cell(r, 2, text); c.font = ds.font(9.5, bold=True, color=ds.t.primary)
        c.fill = ds.fill(ds.t.band); c.alignment = Alignment("left", "center", indent=1)
        ws.row_dimensions[r].height = 18

    def line(r, label, calc=None, bold=False):
        c = ws.cell(r, 2, label); c.font = ds.font(9.5, bold=bold, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1); c.border = ds.hairline_bottom()
        for j in range(3, 15):  # 12 months
            col = get_column_letter(j)
            if calc:
                cell = ds.calc_cell(ws, r, j, ds.t.EUR); cell.value = calc(col)
            else:
                ds.input_cell(ws, r, j, ds.t.EUR)
        tot = ds.calc_cell(ws, r, 15, ds.t.EUR, bold=True)
        tot.value = f"=SUM(C{r}:N{r})"
        ws.row_dimensions[r].height = 16

    section(9, "TRŽBY")
    line(10, "Tržby z jedál"); line(11, "Tržby z nápojov"); line(12, "Ostatné príjmy")
    line(13, "Tržby spolu", calc=lambda c: f"=SUM({c}10:{c}12)", bold=True)         # O13
    section(14, "NÁKLADY NA PREDAJ")
    line(15, "Náklady na jedlo"); line(16, "Náklady na nápoje")
    line(17, "Náklady spolu", calc=lambda c: f"=SUM({c}15:{c}16)", bold=True)
    line(18, "HRUBÁ MARŽA", calc=lambda c: f"={c}13-{c}17", bold=True)               # O18
    section(19, "RÉŽIE")
    for i, name in enumerate(["Mzdy a personál", "Nájom a poplatky", "Energie", "Marketing", "Ostatné réžie"]):
        line(20 + i, name)
    line(25, "Réžie spolu", calc=lambda c: f"=SUM({c}20:{c}24)", bold=True)
    line(26, "ČISTÝ ZISK", calc=lambda c: f"={c}18-{c}25", bold=True)               # O26
    section(27, "HOTOVOSŤ")
    # opening: Jan input, others = prior closing
    ws.cell(28, 2, "Počiatočná hotovosť").font = ds.font(9.5, color=ds.t.ink)
    ws.cell(28, 2).alignment = Alignment("left", "center", indent=1); ws.cell(28, 2).border = ds.hairline_bottom()
    ds.input_cell(ws, 28, 3, ds.t.EUR)
    for j in range(4, 15):
        prev = get_column_letter(j - 1)
        cell = ds.calc_cell(ws, 28, j, ds.t.EUR); cell.value = f"={prev}29"
    ds.calc_cell(ws, 28, 15, ds.t.EUR).value = "=C28"
    ws.cell(29, 2, "Koncová hotovosť").font = ds.font(9.5, bold=True, color=ds.t.ink)
    ws.cell(29, 2).alignment = Alignment("left", "center", indent=1); ws.cell(29, 2).border = ds.hairline_bottom()
    for j in range(3, 15):
        col = get_column_letter(j)
        cell = ds.calc_cell(ws, 29, j, ds.t.EUR, bold=True); cell.value = f"={col}28+{col}26"
    ds.calc_cell(ws, 29, 15, ds.t.EUR, bold=True).value = "=N29"                     # O15 unused
    ws.conditional_formatting.add("C29:N29",
        CellIsRule(operator="lessThan", formula=["0"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.freeze_panes = "C9"


# ----------------------------------------------------------------- 04 Marža
def build_gp(wb):
    ws = wb.create_sheet(S_GP)
    ds.canvas(ws, [30, 13, 13, 13, 15, 14, 12, 14], tab=ds.t.accent)
    ds.title(ws, "Kalkulácia marže receptúr",
             "Náklady na porciu, hrubá marža a odporúčaná cena", span=8)
    label_input(ws, 5, "Názov prevádzky", col_in=3, in_span=3)
    ds.note(ws, 6, "Zadajte náklady a cieľovú maržu. Odporúčaná cena (bez DPH) sa vypočíta; vedľa vidíte reálnu maržu.", span=8)
    ds.thead(ws, 7, ["Jedlo / nápoj", "Veľkosť", "Náklady €", "Cieľ marža %",
                     "Odpor. cena €", "Cena v menu €", "Reálna marža %", "Stav"])
    seed = [
        ("Kôš s domácim chlebom", "120 g", 0.45, 0.70),
        ("Rybacia polievka", "350 ml", 2.10, 0.68),
        ("Hovädzí burger s hranolkami", "300 g", 3.40, 0.68),
        ("Cappuccino", "240 ml", 0.55, 0.85),
        ("Čapované pivo", "500 ml", 1.95, 0.55),
    ]
    N = 22
    for i in range(N):
        r = 8 + i
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        row = seed[i] if i < len(seed) else ("", "", "", "")
        ws.cell(r, 2, row[0]); ws.cell(r, 3, row[1])
        ds.input_cell(ws, r, 4, ds.t.EUR).value = row[2] if row[2] != "" else None
        ds.input_cell(ws, r, 5, ds.t.PCT).value = row[3] if row[3] != "" else None
        ds.calc_cell(ws, r, 6, ds.t.EUR).value = f'=IF(OR(D{r}="",E{r}=""),"",D{r}/(1-E{r}))'
        ds.input_cell(ws, r, 7, ds.t.EUR)
        ds.calc_cell(ws, r, 8, ds.t.PCT).value = f'=IF(OR(G{r}="",D{r}=""),"",(G{r}-D{r})/G{r})'
        ws.cell(r, 9, f'=IF(OR(H{r}="",E{r}=""),"",IF(H{r}<E{r},"NÍZKA","OK"))')
        ws.cell(r, 9).alignment = Alignment("center", "center"); ws.cell(r, 9).font = ds.font(9.5)
        ws.cell(r, 9).border = ds.hairline_bottom()
    last = 7 + N
    ws.conditional_formatting.add(f"I8:I{last}",
        CellIsRule(operator="equal", formula=['"NÍZKA"'], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"I8:I{last}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    # summary anchors for the dashboard
    s1 = ds.calc_cell(ws, 30, 8, ds.t.PCT, bold=True); s1.value = f'=IFERROR(AVERAGEIF(H8:H{last},"<>",H8:H{last}),0)'
    ws.cell(30, 2, "Priemerná reálna marža").font = ds.font(9.5, bold=True, color=ds.t.primary)
    s2 = ds.calc_cell(ws, 31, 8, bold=True); s2.value = f'=COUNTIF(I8:I{last},"NÍZKA")'
    ws.cell(31, 2, "Položky pod cieľom").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.freeze_panes = "B8"


# ----------------------------------------------------------------- 05 Zásoby
def build_stock(wb):
    ws = wb.create_sheet(S_STOCK)
    ds.canvas(ws, [26, 11, 13, 13, 13, 14, 13, 15], tab=ds.t.primary)
    ds.title(ws, "Zásoby a straty",
             "Počiatočné + nákup − koncové = spotreba; hodnota strát ukáže únik marže", span=8)
    label_input(ws, 5, "Názov prevádzky", col_in=3, in_span=3)
    label_input(ws, 6, "Obdobie", col_in=3, in_span=2, fmt=ds.t.DATE)
    ds.note(ws, 7, "Hodnota strát sa počíta automaticky. Vysoké straty sa zvýraznia — konajte pri úniku a prelievaní.", span=8)
    ds.thead(ws, 8, ["Položka", "Jedn.", "Počiatočné", "Nákup", "Koncové",
                     "Straty (množ.)", "Cena/jedn. €", "Hodnota strát €"])
    seed = [
        ("Čapované pivo", "L", 80, 220, 60, 4),
        ("Domáce červené víno", "fľaša", 24, 60, 30, 1),
        ("Čerstvé ryby", "kg", 5, 30, 2, 1.5),
        ("Rib-eye steak", "kg", 8, 25, 6, 0.5),
        ("Mlieko", "L", 20, 120, 25, 3),
    ]
    N = 18
    for i in range(N):
        r = 9 + i
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        row = seed[i] if i < len(seed) else ("", "", "", "", "", "")
        ws.cell(r, 2, row[0]); ws.cell(r, 3, row[1])
        for idx, col in [(2, 4), (3, 5), (4, 6), (5, 7)]:
            ds.input_cell(ws, r, col).value = row[idx] if (len(row) > idx and row[idx] != "") else None
        ds.input_cell(ws, r, 8, ds.t.EUR).value = row[5] if (len(row) > 5 and row[5] != "") else None
        ds.calc_cell(ws, r, 9, ds.t.EUR).value = f'=IF(OR(G{r}="",H{r}=""),"",G{r}*H{r})'
    last = 8 + N
    ws.conditional_formatting.add(f"I9:I{last}",
        CellIsRule(operator="greaterThan", formula=["50"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ds.calc_cell(ws, 27, 9, ds.t.EUR, bold=True).value = f"=SUM(I9:I{last})"   # I27 anchor
    ws.cell(27, 2, "Hodnota strát spolu").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.freeze_panes = "B9"


# ----------------------------------------------------------------- 06 Zmeny
def build_rota(wb):
    ws = wb.create_sheet(S_ROTA)
    ds.canvas(ws, [22, 16, 11, 10, 10, 12, 14, 14], tab=ds.t.accent)
    ds.title(ws, "Zmeny a mzdové náklady",
             "Naplánujte zmeny a hneď vidíte podiel mzdových nákladov na tržbách", span=8)
    label_input(ws, 5, "Názov prevádzky", col_in=3, in_span=3)
    label_input(ws, 6, "Týždeň od", col_in=3, in_span=2, fmt=ds.t.DATE)
    label_input(ws, 7, "Plán tržieb €", col_in=3, in_span=1, fmt=ds.t.EUR)   # C7
    ds.note(ws, 8, "Zadajte hodiny a sadzbu. Náklad = hodiny × sadzba. Cieľ pre gastro býva 25–35 % tržieb.", span=8)
    ds.thead(ws, 9, ["Zamestnanec", "Pozícia", "Deň", "Od", "Do", "Hodiny", "Sadzba €/h", "Náklad €"])
    days = DataValidation(type="list", formula1='"Po,Ut,St,Št,Pi,So,Ne"', allow_blank=True)
    ws.add_data_validation(days)
    N = 20
    for i in range(N):
        r = 10 + i
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        ds.input_cell(ws, r, 2); ds.input_cell(ws, r, 3)
        ds.input_cell(ws, r, 4); days.add(ws.cell(r, 4))
        ds.input_cell(ws, r, 5); ds.input_cell(ws, r, 6)
        ds.input_cell(ws, r, 7); ds.input_cell(ws, r, 8, ds.t.EUR)
        ds.calc_cell(ws, r, 9, ds.t.EUR).value = f'=IF(OR(G{r}="",H{r}=""),"",G{r}*H{r})'
    last = 9 + N
    ds.calc_cell(ws, 30, 9, ds.t.EUR, bold=True).value = f"=SUM(I10:I{last})"   # I30
    ws.cell(30, 2, "Mzdové náklady spolu").font = ds.font(9.5, bold=True, color=ds.t.primary)
    lp = ds.calc_cell(ws, 31, 9, ds.t.PCT, bold=True); lp.value = '=IF(C7="","",I30/C7)'  # I31
    ws.cell(31, 2, "Podiel miezd na tržbách").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.conditional_formatting.add("I31",
        CellIsRule(operator="greaterThan", formula=["0.35"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.freeze_panes = "B10"


# ----------------------------------------------------------------- 07 Tržby
def build_till(wb):
    ws = wb.create_sheet(S_TILL)
    ds.canvas(ws, [13, 14, 13, 13, 14, 15, 13, 16], tab=ds.t.primary)
    ds.title(ws, "Denné tržby a uzávierka pokladne",
             "Porovná uzávierku (Z) s hotovosťou a kartami a upozorní na rozdiel", span=8)
    label_input(ws, 5, "Názov prevádzky", col_in=3, in_span=3)
    label_input(ws, 6, "Mesiac", col_in=3, in_span=2)
    ds.note(ws, 7, "Rozdiel = (hotovosť + karty) − uzávierka. Hodnoty mimo nuly sa zvýraznia — preverte ich.", span=8)
    ds.thead(ws, 8, ["Dátum", "Uzávierka € (Z)", "Hotovosť €", "Karty €",
                     "Vklad €", "Spolu spočítané €", "Rozdiel €", "Kontroloval"])
    N = 31
    for i in range(N):
        r = 9 + i
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        ds.input_cell(ws, r, 2, ds.t.DATE)
        for col in (3, 4, 5, 6):
            ds.input_cell(ws, r, col, ds.t.EUR)
        ds.calc_cell(ws, r, 7, ds.t.EUR).value = f'=IF(AND(D{r}="",E{r}=""),"",D{r}+E{r})'
        ds.calc_cell(ws, r, 8, ds.t.EUR).value = f'=IF(OR(C{r}="",G{r}=""),"",G{r}-C{r})'
        ds.input_cell(ws, r, 9)
    last = 8 + N
    ws.conditional_formatting.add(f"H9:H{last}",
        CellIsRule(operator="greaterThan", formula=["1"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"H9:H{last}",
        CellIsRule(operator="lessThan", formula=["-1"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ds.calc_cell(ws, 40, 3, ds.t.EUR, bold=True).value = f"=SUM(C9:C{last})"    # C40
    ws.cell(40, 2, "Spolu (rok)").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ds.calc_cell(ws, 40, 8, ds.t.EUR, bold=True).value = f"=SUM(H9:H{last})"    # H40 anchor
    ws.freeze_panes = "B9"


# ----------------------------------------------------------------- 08 Školenia
def build_training(wb):
    ws = wb.create_sheet(S_TRAIN)
    topics = ["Zaškolenie", "Manipulácia s bremenami", "Hygiena potravín",
              "Alergény", "Požiarna ochrana", "BOZP / prvá pomoc"]
    ncols = 2 + len(topics) + 1
    ds.canvas(ws, [22, 14] + [13] * len(topics) + [14], tab=ds.t.ink)
    ds.title(ws, "Školenia a zaškolenie",
             "Dôkaz, že pracovník bol školený PRED zmenou — vrátane manipulácie a zaškolenia", span=ncols)
    label_input(ws, 5, "Názov prevádzky", col_in=3, in_span=3)
    ds.note(ws, 6, "Zadajte dátum absolvovania (DD/MM/RRRR). Prázdna bunka = chýbajúce školenie pred prácou.", span=ncols)
    ds.thead(ws, 7, ["Zamestnanec", "Nástup"] + topics + ["Preškolenie"], height=34)
    N = 16
    for i in range(N):
        r = 8 + i
        ds.trow(ws, r, ncols, zebra_on=(i % 2 == 1))
        ds.input_cell(ws, r, 2)
        for j in range(3, 3 + ncols - 2):
            ds.input_cell(ws, r, j, ds.t.DATE)
    ds.note(ws, 8 + N + 1,
            "Pozn.: podľa zákona o BOZP musia byť zamestnanci školení na svoju prácu — maticu udržiavajte aktuálnu a s dátumami.",
            span=ncols, tone="good")
    ws.freeze_panes = "C8"


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    build_method(wb)
    build_planner(wb)
    build_dashboard(wb)
    build_cash(wb)
    build_gp(wb)
    build_stock(wb)
    build_rota(wb)
    build_till(wb)
    build_training(wb)
    wb.properties.title = "Prevádzka a marža pre gastro (SK)"
    wb.properties.creator = "ASSET-FORGE"
    wb.properties.subject = "Gastro prevádzka — cash flow, marža, zásoby, zmeny, tržby, školenia"
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    # guard: Excel hard limit on sheet-name length
    for nm in wb.sheetnames:
        assert len(nm) <= 31, f"tab name too long: {nm!r} ({len(nm)})"
    out = os.path.join(PRODUCTS, "P2_SK_Hospitality_Premium.xlsx")
    wb.save(out)
    print(f"✓ Built SK premium pilot: {out}  ({len(wb.sheetnames)} sheets)")
    print("  Sheets:", " | ".join(wb.sheetnames))
