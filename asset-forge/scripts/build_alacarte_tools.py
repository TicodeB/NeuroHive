#!/usr/bin/env python3
"""
ASSET-FORGE · Launch extension — À-la-carte tools (value-ladder Rung 1)

Composes 4 standalone single-tool workbooks (full + watermarked DEMO) by reusing
the already-validated sheet builders from the P1/P2 flagship scripts:

  P3  — H&S Risk Assessment & Safety Statement Builder   (€19)  <- p1.build_hs
  P4  — Cashflow & P&L Tracker                           (€24)  <- p2.build_cashflow
  P5  — Fire Safety Register & Checks Log                (€15)  <- p1.build_fire
  P12 — Staff Training & Induction Matrix                (€15)  <- p2.build_training

Bilingual EN/SK comes from the source builders. Each workbook gets its own
Start-Here sheet and the safe DEMO pattern (notice sheet + sheet protection —
never insert_rows, see P13 v1.1 postmortem).

Run:  python3 scripts/build_alacarte_tools.py
Out:  products/P{3,4,5,12}_*.xlsx + DEMO variants
"""
from __future__ import annotations
import os
import sys
import zipfile
import xml.etree.ElementTree as ET

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import build_p1_compliance_pack as p1
import build_p2_operations_bundle as p2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

PRODUCTS = p1.PRODUCTS
VERSION = "v1.0"
BUILD_DATE = "09/06/2026"  # DD/MM/YYYY (EU)


def mini_cover(wb, name_en, name_sk, what_en, what_sk, legal, steps_en, steps_sk):
    """Standalone Start-Here sheet, built with the P1 style helpers."""
    ws = wb.active
    ws.title = "00 · Start Here · Začnite tu"
    ws.sheet_view.showGridLines = False
    p1.set_widths(ws, [4, 30, 30, 30, 30])
    p1.title_block(ws, name_en, name_sk, what_en, what_sk, span=5)

    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, f"Leanta · leanta.ie · {VERSION} · {BUILD_DATE} · single-business licence / licencia pre jednu prevádzku")
    c.font = p1.f(9, False, p1.TEAL, italic=True)
    c.alignment = Alignment(horizontal="center")

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    h = ws.cell(r, 1, "How to use · Ako používať")
    h.font = p1.f(12, True, p1.NAVY)
    for i, (en, sk) in enumerate(zip(steps_en, steps_sk), start=1):
        r += 1
        ws.cell(r, 1, str(i)).font = p1.f(11, True, p1.TEAL)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(r, 2, en).font = p1.f(10)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        e = ws.cell(r, 4, sk)
        e.font = p1.f(10, italic=True)
        for col in range(1, 6):
            ws.cell(r, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 26

    r += 2
    p1.note(ws, r,
            f"Legal basis: {legal}. Template only — not legal advice; adapt to your premises.",
            "Právny základ uvedený vyššie. Šablóna — nie je právne poradenstvo; prispôsobte svojej prevádzke.",
            span=5)
    return ws


def add_demo(wb, title):
    """Safe DEMO: dedicated notice sheet first + read-only protection on all sheets."""
    notice = wb.create_sheet("DEMO Preview · Ukážka", 0)
    notice.sheet_view.showGridLines = False
    p1.set_widths(notice, [4, 28, 28, 28, 28])
    notice.merge_cells("A1:E1")
    b = notice.cell(1, 1, "DEMO — PREVIEW ONLY · NOT FOR RESALE")
    b.font = p1.f(16, True, "FFFFFF"); b.fill = p1.fill("C00000")
    b.alignment = Alignment(horizontal="center", vertical="center")
    notice.row_dimensions[1].height = 36
    notice.merge_cells("A2:E2")
    b2 = notice.cell(2, 1, "Ukážka — len na náhľad · nie na ďalší predaj")
    b2.font = p1.f(12, True, "C00000"); b2.fill = p1.fill(p1.SAND)
    b2.alignment = Alignment(horizontal="center", vertical="center")
    notice.row_dimensions[2].height = 24
    notice.merge_cells("A4:E4")
    m = notice.cell(4, 1, "Browse the tabs to preview. Buy the full version at leanta.ie to unlock editing.")
    m.font = p1.f(11, False, p1.NAVY); m.alignment = Alignment(horizontal="center", vertical="center")
    notice.sheet_properties.tabColor = "C00000"
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = "demo"
    wb.properties.title = title + " — DEMO (EN/SK)"
    return wb


TOOLS = [
    {
        "id": "P3",
        "file": "P3_HS_Safety_Statement_Builder",
        "build": p1.build_hs,
        "rename": ("06 · Health & Safety", "01 · Health & Safety"),
        "name_en": "H&S Risk Assessment & Safety Statement Builder",
        "name_sk": "Posúdenie rizík BOZP a bezpečnostné vyhlásenie",
        "what_en": "Risk register with auto Likelihood × Severity rating, Safety Statement scaffold and accident log",
        "what_sk": "Register rizík s automatickým hodnotením, bezpečnostné vyhlásenie a kniha úrazov",
        "legal": "Safety, Health & Welfare at Work Act 2005 (S.19/S.20)",
        "steps_en": ["Fill the business fields at the top.",
                     "List each hazard; pick Likelihood and Severity — the risk band colours itself.",
                     "Record controls and the person responsible.",
                     "Review yearly or after any change; log accidents as they happen."],
        "steps_sk": ["Vyplňte údaje o prevádzke.",
                     "Zapíšte každé riziko; zvoľte pravdepodobnosť a závažnosť — pásmo sa zafarbí samo.",
                     "Zaznamenajte opatrenia a zodpovednú osobu.",
                     "Revidujte ročne alebo po každej zmene; úrazy zapisujte priebežne."],
    },
    {
        "id": "P4",
        "file": "P4_Cashflow_PL_Tracker",
        "build": p2.build_cashflow,
        "rename": None,  # already "01 · Cashflow & P&L"
        "name_en": "Cashflow & P&L Tracker",
        "name_sk": "Sledovač cash flow a ziskov a strát",
        "what_en": "12-month rolling cash position + monthly P&L: revenue → GP → net profit → closing cash, auto-calculated",
        "what_sk": "12-mesačný cash flow a mesačný výkaz: tržby → hrubá marža → zisk → konečná hotovosť",
        "legal": "Good practice — management accounts (not a statutory record)",
        "steps_en": ["Enter opening cash once.",
                     "Each month, fill revenue and cost lines — totals and closing cash calculate themselves.",
                     "Negative closing cash turns red months ahead — act early.",
                     "Use the P&L block to see margin trends month by month."],
        "steps_sk": ["Zadajte počiatočnú hotovosť.",
                     "Každý mesiac vyplňte tržby a náklady — súčty sa vypočítajú samé.",
                     "Záporná hotovosť sa zafarbí načerveno s predstihom — konajte včas.",
                     "Blok ziskov a strát ukáže vývoj marže po mesiacoch."],
    },
    {
        "id": "P5",
        "file": "P5_Fire_Safety_Register",
        "build": p1.build_fire,
        "rename": ("07 · Fire Safety", "01 · Fire Safety"),
        "name_en": "Fire Safety Register & Checks Log",
        "name_sk": "Register požiarnej ochrany a záznam kontrol",
        "what_en": "Equipment register with auto OVERDUE / DUE SOON flags, routine checks, drills and emergency lighting log",
        "what_sk": "Register zariadení s automatickými upozorneniami na termíny, kontroly, cvičenia a núdzové osvetlenie",
        "legal": "Fire Services Acts 1981 & 2003",
        "steps_en": ["List every extinguisher, alarm point, exit and light.",
                     "Set the check frequency — the next-due date flags itself OVERDUE / DUE SOON / OK.",
                     "Tick off routine checks with date and initials.",
                     "Log drills and keep the register where an inspector can see it."],
        "steps_sk": ["Zapíšte každý hasiaci prístroj, hlásič, východ a svetlo.",
                     "Nastavte frekvenciu kontrol — termíny sa strážia samé.",
                     "Kontroly potvrdzujte dátumom a podpisom.",
                     "Zaznamenávajte cvičenia; register majte pripravený pre kontrolu."],
    },
    {
        "id": "P12",
        "file": "P12_Staff_Training_Matrix",
        "build": p2.build_training,
        "rename": ("06 · Training & Induction", "01 · Training & Induction"),
        "name_en": "Staff Training & Induction Matrix",
        "name_sk": "Matica školení a zaškolenia zamestnancov",
        "what_en": "Dated proof of training per employee — induction, manual handling, hygiene, allergens, fire, first aid — with refresher-due flags",
        "what_sk": "Datovaný dôkaz školení pre každého zamestnanca — zaškolenie, bremená, hygiena, alergény, požiarna ochrana, prvá pomoc",
        "legal": "SHWWA 2005 + Reg. 852/2004 (hygiene training duty)",
        "steps_en": ["Add each employee as a row.",
                     "Enter the date each training was completed — gaps stay visibly empty.",
                     "Set refresher dates; due items flag themselves.",
                     "Print or show on screen when an inspector asks for proof."],
        "steps_sk": ["Pridajte každého zamestnanca ako riadok.",
                     "Zapíšte dátum absolvovania každého školenia — medzery ostanú viditeľné.",
                     "Nastavte termíny opakovania; blížiace sa termíny sa zvýraznia.",
                     "Pri kontrole ukážte vytlačené alebo na obrazovke."],
    },
]


def merged_overlap_check(path):
    """No two merged ranges on a sheet may overlap (Excel corruption guard)."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ranges = list(ws.merged_cells.ranges)
        for i, a in enumerate(ranges):
            for b in ranges[i + 1:]:
                if (a.min_row <= b.max_row and b.min_row <= a.max_row and
                        a.min_col <= b.max_col and b.min_col <= a.max_col):
                    raise SystemExit(f"✗ {path}: overlapping merges {a} / {b} on '{ws.title}'")


def validate(path):
    with zipfile.ZipFile(path) as z:
        bad = z.testzip()
        if bad:
            raise SystemExit(f"✗ {path}: corrupt zip member {bad}")
        for name in z.namelist():
            if name.endswith(".xml"):
                ET.fromstring(z.read(name))
    wb = load_workbook(path)
    for name in wb.sheetnames:
        assert len(name) <= 31, f"✗ {path}: tab '{name}' > 31 chars"
    merged_overlap_check(path)
    return len(wb.sheetnames)


def build_tool(spec):
    def assemble():
        wb = Workbook()
        mini_cover(wb, spec["name_en"], spec["name_sk"], spec["what_en"], spec["what_sk"],
                   spec["legal"], spec["steps_en"], spec["steps_sk"])
        spec["build"](wb)
        if spec["rename"]:
            old, new = spec["rename"]
            assert len(new) <= 31
            wb[old].title = new
        wb.properties.title = spec["name_en"] + " (EN/SK)"
        wb.properties.creator = "Leanta"
        return wb

    full_path = os.path.join(PRODUCTS, spec["file"] + ".xlsx")
    assemble().save(full_path)
    n = validate(full_path)
    print(f"✓ {spec['id']} full: {full_path} ({n} sheets)")

    demo_path = os.path.join(PRODUCTS, spec["file"].replace(spec["id"] + "_", spec["id"] + "_DEMO_", 1) + ".xlsx")
    add_demo(assemble(), spec["name_en"]).save(demo_path)
    n = validate(demo_path)
    print(f"✓ {spec['id']} demo: {demo_path} ({n} sheets)")
    return full_path, demo_path


if __name__ == "__main__":
    for spec in TOOLS:
        build_tool(spec)
    print("All à-la-carte tools built + validated (zip/XML, tab length, merge overlaps).")
