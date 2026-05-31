#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 13d — generic PackSpec-driven builder
=========================================================
ONE builder that turns any `PackSpec` (from `scripts/pack_spec.py`) into a
premium, single-language workbook on `scripts/design_system.py`. The skeleton +
every formula are fixed here; the vertical only supplies terminology + palette.

  python3 scripts/build_pack.py            # builds every spec in the REGISTRY
  python3 scripts/build_pack.py baker_sk   # builds one

Out:  products/pack_<key>.xlsx  (+ preview PNGs via render_preview)

Phase-13d proof: `hospitality_sk` reproduces the hand-built pilot, and
`baker_sk` is the first net-new vertical with ZERO new layout code.
"""
from __future__ import annotations
import os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from design_system import Theme, DS
from pack_spec import REGISTRY, FIXED, validate
from i18n import get_strings

PRODUCTS = os.path.normpath(os.path.join(HERE, "..", "products"))
os.makedirs(PRODUCTS, exist_ok=True)
VERSION, BUILD_DATE = "v1.0", "31/05/2026"

EUR_METRICS = {"revenue_total", "gross", "net", "cash_close", "cost", "loss_value",
               "variance_total", "rev_plan", "rev_var", "net_var"}
PCT_METRICS = {"avg", "pct", "rev_var_pct"}


def refc(name, cell):
    return f"'{name}'!{cell}"


def label_input(ds, ws, row, label, col_in=3, in_span=2, fmt=None):
    c = ws.cell(row, 2, label); c.font = ds.font(10, bold=True, color=ds.t.ink)
    c.alignment = Alignment("left", "center")
    if in_span > 1:
        ws.merge_cells(start_row=row, start_column=col_in, end_row=row, end_column=col_in + in_span - 1)
    ds.input_cell(ws, row, col_in, fmt); ws.row_dimensions[row].height = 20


# --------------------------------------------------------------- fixed spine
def build_method(ws, ds, t, spec, S):
    ds.canvas(ws, [40, 30, 14, 14, 14, 14], tab=ds.t.ink)
    ds.title(ws, spec.vertical, t["subtitle"])
    r = 5
    r = ds.section(ws, r, S["method_contains"])
    for m in spec.modules:
        if m.type in FIXED:
            continue
        a = ws.cell(r, 2, "  " + m.terms["title"]); a.font = ds.font(10.5, bold=True, color=ds.t.primary)
        a.alignment = Alignment("left", "center", indent=1)
        b = ws.cell(r, 3, m.terms.get("subtitle", "")); b.font = ds.font(10, color=ds.t.muted)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 20; r += 1
    r += 1
    r = ds.section(ws, r, S["method_how"])
    for s in t["steps"]:
        c = ws.cell(r, 2, "•  " + s); c.font = ds.font(10, color=ds.t.ink)
        c.alignment = Alignment("left", "center", wrap_text=True, indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22; r += 1
    r += 1
    r = ds.section(ws, r, S["method_colors"])
    r = ds.note(ws, r, S["method_color_note"])
    ds.footer(ws, r + 1, f"ASSET-FORGE · {VERSION} · {BUILD_DATE} · {S['footer_region']} · "
                         f"{S['footer_tail']}")


def build_planner(ws, ds, t, S):
    ds.canvas(ws, [12, 44, 16, 12], tab=ds.t.accent)
    ds.title(ws, t["title"], t["subtitle"], span=4)
    label_input(ds, ws, 5, S["planner_date"], in_span=1, fmt=ds.t.DATE)
    label_input(ds, ws, 6, S["planner_day"], in_span=2)
    r = ds.section(ws, 8, S["planner_priorities"], span=4)
    for i in range(1, 4):
        c = ws.cell(r, 2, str(i)); c.font = ds.font(11, bold=True, color=ds.t.primary)
        c.alignment = Alignment("center", "center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ds.input_cell(ws, r, 3); ws.row_dimensions[r].height = 22; r += 1
    r += 1
    r = ds.section(ws, r, S["planner_schedule"], span=4)
    pr = DataValidation(type="list", formula1=f'"{S["dv_priority"]}"', allow_blank=True)
    done = DataValidation(type="list", formula1=f'"{S["dv_yesno"]}"', allow_blank=True)
    ws.add_data_validation(pr); ws.add_data_validation(done)
    r = ds.thead(ws, r, S["planner_head"])
    for i, h in enumerate(range(7, 23)):
        ds.trow(ws, r, 4, zebra_on=(i % 2 == 1))
        tcell = ws.cell(r, 2, f"{h:02d}:00"); tcell.alignment = Alignment("center", "center")
        tcell.font = ds.font(9.5, bold=True, color=ds.t.muted)
        ds.input_cell(ws, r, 3); ws.cell(r, 3).alignment = Alignment("left", "center", indent=1)
        ds.input_cell(ws, r, 4); pr.add(ws.cell(r, 4))
        ds.input_cell(ws, r, 5); done.add(ws.cell(r, 5)); r += 1
    r += 1
    r = ds.section(ws, r, S["planner_openclose"], span=4)
    for i, chk in enumerate(t["checklist"]):
        ds.trow(ws, r, 4, zebra_on=(i % 2 == 1))
        c = ws.cell(r, 2, "  " + chk); c.alignment = Alignment("left", "center", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ds.input_cell(ws, r, 5); done.add(ws.cell(r, 5)); r += 1


# --------------------------------------------------------------- operational
def build_ledger(ws, ds, t, name, S):
    # Datarails-style FP&A spine (Phase 13i): actuals + Plán + Odchýlka €/% so the
    # ledger compares plan vs reality, the way every FP&A view does. Monthly cells
    # are actuals; Plán is the annual budget per line; variance is computed.
    plan_label = t.get("plan_label", S["ledger_plan_default"])
    ds.canvas(ws, [26] + [9.5] * 12 + [12, 12, 12, 11], tab=ds.t.primary)
    ds.title(ws, t["title"], t["subtitle"], span=17)
    label_input(ds, ws, 5, S["business_name"], in_span=3)
    label_input(ds, ws, 6, S["year"], in_span=1)
    ds.note(ws, 7, S["ledger_note"], span=17)
    months = S["months"]
    ds.thead(ws, 8, [S["ledger_item_head"]] + months + [S["ledger_head_year"], plan_label,
                     S["ledger_head_var_eur"], S["ledger_head_var_pct"]])

    def section(r, text):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
        c = ws.cell(r, 2, text); c.font = ds.font(9.5, bold=True, color=ds.t.primary)
        c.fill = ds.fill(ds.t.band); c.alignment = Alignment("left", "center", indent=1)
        ws.row_dimensions[r].height = 18

    def line(r, label, calc=None, bold=False, variance=False):
        c = ws.cell(r, 2, label); c.font = ds.font(9.5, bold=bold, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1); c.border = ds.hairline_bottom()
        for j in range(3, 15):
            col = get_column_letter(j)
            if calc:
                ds.calc_cell(ws, r, j, ds.t.EUR).value = calc(col)
            else:
                ds.input_cell(ws, r, j, ds.t.EUR)
        # O = actual annual (sum of months)
        ds.calc_cell(ws, r, 15, ds.t.EUR, bold=True).value = f"=SUM(C{r}:N{r})"
        # P = plan (rolls up on total lines, else an input); Q/R = variance € / %
        if calc:
            ds.calc_cell(ws, r, 16, ds.t.EUR, bold=bold).value = calc("P")
        else:
            ds.input_cell(ws, r, 16, ds.t.EUR)
        ds.calc_cell(ws, r, 17, ds.t.EUR, bold=bold).value = f'=IF(OR(O{r}="",P{r}=""),"",O{r}-P{r})'
        ds.calc_cell(ws, r, 18, ds.t.PCT, bold=bold).value = f'=IF(OR(P{r}="",P{r}=0,O{r}=""),"",(O{r}-P{r})/P{r})'
        ws.row_dimensions[r].height = 16

    r = 9
    section(r, S["ledger_sec_revenue"]); r += 1
    rev0 = r
    for nm in t["revenue_lines"]: line(r, nm); r += 1
    rev_total = r; line(r, S["ledger_revenue_total"], calc=lambda c: f"=SUM({c}{rev0}:{c}{rev_total-1})", bold=True); r += 1
    section(r, S["ledger_sec_cos"]); r += 1
    cos0 = r
    for nm in t["cos_lines"]: line(r, nm); r += 1
    cos_total = r; line(r, S["ledger_cos_total"], calc=lambda c: f"=SUM({c}{cos0}:{c}{cos_total-1})", bold=True); r += 1
    gross = r; line(r, S["ledger_gross"], calc=lambda c: f"={c}{rev_total}-{c}{cos_total}", bold=True); r += 1
    section(r, S["ledger_sec_overhead"]); r += 1
    oh0 = r
    for nm in t["overhead_lines"]: line(r, nm); r += 1
    oh_total = r; line(r, S["ledger_overhead_total"], calc=lambda c: f"=SUM({c}{oh0}:{c}{oh_total-1})", bold=True); r += 1
    net = r; line(r, S["ledger_net"], calc=lambda c: f"={c}{gross}-{c}{oh_total}", bold=True); r += 1
    # variance RAG on the "more-is-better" bottom lines: under plan = red, over = green
    for vr in (rev_total, gross, net):
        ws.conditional_formatting.add(f"Q{vr}",
            CellIsRule(operator="lessThan", formula=["0"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
        ws.conditional_formatting.add(f"Q{vr}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    section(r, S["ledger_sec_cash"]); r += 1
    ob = r
    ws.cell(ob, 2, S["ledger_cash_open"]).font = ds.font(9.5, color=ds.t.ink)
    ws.cell(ob, 2).alignment = Alignment("left", "center", indent=1); ws.cell(ob, 2).border = ds.hairline_bottom()
    ds.input_cell(ws, ob, 3, ds.t.EUR)
    cb = ob + 1
    for j in range(4, 15):
        ds.calc_cell(ws, ob, j, ds.t.EUR).value = f"={get_column_letter(j-1)}{cb}"
    ds.calc_cell(ws, ob, 15, ds.t.EUR).value = "=C" + str(ob)
    ws.cell(cb, 2, S["ledger_cash_close"]).font = ds.font(9.5, bold=True, color=ds.t.ink)
    ws.cell(cb, 2).alignment = Alignment("left", "center", indent=1); ws.cell(cb, 2).border = ds.hairline_bottom()
    for j in range(3, 15):
        col = get_column_letter(j)
        ds.calc_cell(ws, cb, j, ds.t.EUR, bold=True).value = f"={col}{ob}+{col}{net}"
    ds.calc_cell(ws, cb, 15, ds.t.EUR, bold=True).value = f"=N{cb}"
    ws.conditional_formatting.add(f"C{cb}:N{cb}",
        CellIsRule(operator="lessThan", formula=["0"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.freeze_panes = "C9"
    return {"revenue_total": f"O{rev_total}", "gross": f"O{gross}", "net": f"O{net}",
            "cash_close": f"N{cb}", "rev_plan": f"P{rev_total}", "rev_var": f"Q{rev_total}",
            "rev_var_pct": f"R{rev_total}", "net_var": f"Q{net}"}


def build_margin(ws, ds, t, name, S):
    low, okk = S["margin_low"], S["margin_ok"]
    ds.canvas(ws, [30, 13, 13, 13, 15, 14, 12, 14], tab=ds.t.accent)
    ds.title(ws, t["title"], t["subtitle"], span=8)
    label_input(ds, ws, 5, S["business_name"], in_span=3)
    ds.note(ws, 6, S["margin_note"], span=8)
    ds.thead(ws, 7, [t["unit_label"]] + S["margin_head"])
    N = 22
    for i in range(N):
        r = 8 + i; ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        if i < len(t["seed_items"]): ws.cell(r, 2, t["seed_items"][i])
        ds.input_cell(ws, r, 4, ds.t.EUR); ds.input_cell(ws, r, 5, ds.t.PCT)
        ds.calc_cell(ws, r, 6, ds.t.EUR).value = f'=IF(OR(D{r}="",E{r}=""),"",D{r}/(1-E{r}))'
        ds.input_cell(ws, r, 7, ds.t.EUR)
        ds.calc_cell(ws, r, 8, ds.t.PCT).value = f'=IF(OR(G{r}="",D{r}=""),"",(G{r}-D{r})/G{r})'
        sc = ws.cell(r, 9, f'=IF(OR(H{r}="",E{r}=""),"",IF(H{r}<E{r},"{low}","{okk}"))')
        sc.alignment = Alignment("center", "center"); sc.font = ds.font(9.5); sc.border = ds.hairline_bottom()
    last = 7 + N
    ws.conditional_formatting.add(f"I8:I{last}", CellIsRule(operator="equal", formula=[f'"{low}"'], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"I8:I{last}", CellIsRule(operator="equal", formula=[f'"{okk}"'], fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    ds.calc_cell(ws, 30, 8, ds.t.PCT, bold=True).value = f'=IFERROR(AVERAGEIF(H8:H{last},"<>",H8:H{last}),0)'
    ws.cell(30, 2, S["margin_avg"]).font = ds.font(9.5, bold=True, color=ds.t.primary)
    ds.calc_cell(ws, 31, 8, bold=True).value = f'=COUNTIF(I8:I{last},"{low}")'
    ws.cell(31, 2, S["margin_below"]).font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.freeze_panes = "B8"
    return {"avg": "H30", "below_target": "H31"}


def build_stock(ws, ds, t, name, S):
    ds.canvas(ws, [26, 11, 13, 13, 13, 14, 13, 15], tab=ds.t.primary)
    ds.title(ws, t["title"], t["subtitle"], span=8)
    label_input(ds, ws, 5, S["business_name"], in_span=3)
    label_input(ds, ws, 6, S["period"], in_span=2, fmt=ds.t.DATE)
    ds.note(ws, 7, S["stock_note"], span=8)
    ds.thead(ws, 8, [t["item_label"], S["stock_head_unit"], S["stock_head_open"], S["stock_head_buy"], S["stock_head_close"],
                     f"{t['loss_label']} {S['stock_qty_suffix']}", S["stock_unit_price"], S["stock_value_tmpl"].format(loss=t['loss_label'].lower())])
    N = 18
    for i in range(N):
        r = 9 + i; ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        if i < len(t["seed_items"]): ws.cell(r, 2, t["seed_items"][i])
        for col in (4, 5, 6, 7): ds.input_cell(ws, r, col)
        ds.input_cell(ws, r, 8, ds.t.EUR)
        ds.calc_cell(ws, r, 9, ds.t.EUR).value = f'=IF(OR(G{r}="",H{r}=""),"",G{r}*H{r})'
    last = 8 + N
    ws.conditional_formatting.add(f"I9:I{last}", CellIsRule(operator="greaterThan", formula=["50"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ds.calc_cell(ws, 27, 9, ds.t.EUR, bold=True).value = f"=SUM(I9:I{last})"
    ws.cell(27, 2, S["stock_total_tmpl"].format(loss=t['loss_label'])).font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.freeze_panes = "B9"
    return {"loss_value": "I27"}


def build_labour(ws, ds, t, name, S):
    ds.canvas(ws, [22, 16, 11, 10, 10, 12, 14, 14], tab=ds.t.accent)
    ds.title(ws, t["title"], t["subtitle"], span=8)
    label_input(ds, ws, 5, S["business_name"], in_span=3)
    label_input(ds, ws, 6, S["labour_week"], in_span=2, fmt=ds.t.DATE)
    label_input(ds, ws, 7, S["labour_sales_plan"], in_span=1, fmt=ds.t.EUR)
    tp = t["target_pct"]
    ds.note(ws, 8, S["labour_note_tmpl"].format(pct=int(tp*100)), span=8)
    ds.thead(ws, 9, [t["role_label"]] + S["labour_head"])
    days = DataValidation(type="list", formula1=f'"{S["dv_days"]}"', allow_blank=True)
    ws.add_data_validation(days)
    N = 20
    for i in range(N):
        r = 10 + i; ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        for col in (2, 3, 5, 6, 7): ds.input_cell(ws, r, col)
        ds.input_cell(ws, r, 4); days.add(ws.cell(r, 4))
        ds.input_cell(ws, r, 8, ds.t.EUR)
        ds.calc_cell(ws, r, 9, ds.t.EUR).value = f'=IF(OR(G{r}="",H{r}=""),"",G{r}*H{r})'
    last = 9 + N
    ds.calc_cell(ws, 30, 9, ds.t.EUR, bold=True).value = f"=SUM(I10:I{last})"
    ws.cell(30, 2, S["labour_total"]).font = ds.font(9.5, bold=True, color=ds.t.primary)
    ds.calc_cell(ws, 31, 9, ds.t.PCT, bold=True).value = "=IF(C7=\"\",\"\",I30/C7)"
    ws.cell(31, 2, S["labour_pct"]).font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.conditional_formatting.add("I31", CellIsRule(operator="greaterThan", formula=[str(tp)], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.freeze_panes = "B10"
    return {"pct": "I31", "cost": "I30"}


def build_takings(ws, ds, t, name, S):
    ds.canvas(ws, [13, 14, 13, 13, 14, 15, 13, 16], tab=ds.t.primary)
    ds.title(ws, t["title"], t["subtitle"], span=8)
    label_input(ds, ws, 5, S["business_name"], in_span=3)
    label_input(ds, ws, 6, S["month"], in_span=2)
    ds.note(ws, 7, S["takings_note_tmpl"].format(src=t['source_label']), span=8)
    ds.thead(ws, 8, S["takings_head"])
    N = 31
    for i in range(N):
        r = 9 + i; ds.trow(ws, r, 8, zebra_on=(i % 2 == 1))
        ds.input_cell(ws, r, 2, ds.t.DATE)
        for col in (3, 4, 5, 6): ds.input_cell(ws, r, col, ds.t.EUR)
        ds.calc_cell(ws, r, 7, ds.t.EUR).value = f'=IF(AND(D{r}="",E{r}=""),"",D{r}+E{r})'
        ds.calc_cell(ws, r, 8, ds.t.EUR).value = f'=IF(OR(C{r}="",G{r}=""),"",G{r}-C{r})'
        ds.input_cell(ws, r, 9)
    last = 8 + N
    ws.conditional_formatting.add(f"H9:H{last}", CellIsRule(operator="greaterThan", formula=["1"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"H9:H{last}", CellIsRule(operator="lessThan", formula=["-1"], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ds.calc_cell(ws, 40, 3, ds.t.EUR, bold=True).value = f"=SUM(C9:C{last})"
    ws.cell(40, 2, S["total_year"]).font = ds.font(9.5, bold=True, color=ds.t.primary)
    ds.calc_cell(ws, 40, 8, ds.t.EUR, bold=True).value = f"=SUM(H9:H{last})"
    ws.freeze_panes = "B9"
    return {"variance_total": "H40"}


def build_training(ws, ds, t, name, S):
    topics = t["topics"]; ncols = 2 + len(topics) + 1
    ds.canvas(ws, [22, 14] + [13] * len(topics) + [14], tab=ds.t.ink)
    ds.title(ws, t["title"], t["subtitle"], span=ncols)
    label_input(ds, ws, 5, S["business_name"], in_span=3)
    ds.note(ws, 6, S["training_note"], span=ncols)
    ds.thead(ws, 7, [S["training_head_employee"], S["training_head_start"]] + topics + [S["training_head_retrain"]], height=34)
    N = 16
    for i in range(N):
        r = 8 + i; ds.trow(ws, r, ncols, zebra_on=(i % 2 == 1))
        ds.input_cell(ws, r, 2)
        for j in range(3, 1 + ncols): ds.input_cell(ws, r, j, ds.t.DATE)
    ds.note(ws, 8 + N + 1, S["training_legal_note"], span=ncols, tone="good")
    ws.freeze_panes = "C8"
    return {}


OPS = {"LEDGER_12M": build_ledger, "MARGIN": build_margin, "STOCK": build_stock,
       "LABOUR": build_labour, "TAKINGS": build_takings, "TRAINING": build_training}


def build_dashboard(ws, ds, t, anchors, S):
    ds.canvas(ws, [13] * 11, tab=ds.t.primary)
    ds.title(ws, t["title"], t["subtitle"], span=11)
    positions = [2, 6, 10]
    kpi_cells = []
    for idx, (label, src) in enumerate(t["kpis"]):
        mod, _, metric = src.partition(".")
        cell = anchors.get(mod, {}).get(metric)
        if not cell:
            continue
        top = 5 + (idx // 3) * 5
        left = positions[idx % 3]
        fmt = ds.t.EUR0 if metric in EUR_METRICS else (ds.t.PCT if metric in PCT_METRICS else None)
        accent = ds.t.accent if idx % 3 == 1 else (ds.t.ink if idx % 3 == 2 else ds.t.primary)
        val = ds.kpi(ws, top, left, label, f"=IFERROR({cell},0)", fmt=fmt, accent=accent)
        kpi_cells.append((val.coordinate, metric))
        # drill-down (Datarails pattern): the tile label links to its source cell
        lab = ws.cell(top + 1, left)
        lab.hyperlink = "#" + cell
        lab.value = (lab.value or label.upper()) + "  ↗"
        if metric == "net":
            ws.conditional_formatting.add(val.coordinate, CellIsRule(operator="lessThan", formula=["0"], font=ds.font(20, True, ds.t.bad)))
        if metric == "pct":
            ws.conditional_formatting.add(val.coordinate, CellIsRule(operator="greaterThan", formula=["0.35"], font=ds.font(20, True, ds.t.bad)))
    rows_used = (len(t["kpis"]) + 2) // 3
    r = 5 + rows_used * 5
    r = ds.section(ws, r, S["dash_insights"], span=11)
    A = anchors
    insights = []
    if "LEDGER_12M" in A:
        L = A["LEDGER_12M"]
        # plan-aware variance narrative (Datarails FP&A Genius parity, offline)
        rv, rvp = L.get("rev_var"), L.get("rev_var_pct")
        if rv and rvp:
            insights.append(S["ins_rev_plan"].format(rev_plan=L["rev_plan"], rv=rv, rvp=rvp))
        insights.append(S["ins_net"].format(net=L["net"], rev=L["revenue_total"]))
    if "LABOUR" in A:
        insights.append(S["ins_labour"].format(pct=A["LABOUR"]["pct"]))
    if "STOCK" in A:
        insights.append(S["ins_stock"].format(loss=A["STOCK"]["loss_value"]))
    if "MARGIN" in A:
        insights.append(S["ins_margin"].format(bt=A["MARGIN"]["below_target"]))
    if "TAKINGS" in A:
        insights.append(S["ins_takings"].format(vt=A["TAKINGS"]["variance_total"]))
    insight_cells = []
    for i, formula in enumerate(insights):
        ds.trow(ws, r, 11, zebra_on=(i % 2 == 1), align="left")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        c = ws.cell(r, 2, formula); c.font = ds.font(10.5, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1); ws.row_dimensions[r].height = 24
        insight_cells.append(c.coordinate); r += 1
    ds.note(ws, r + 1, S["dash_auto_note"], span=11)
    return {"kpi_cells": kpi_cells, "insight_cells": insight_cells}


def build(spec_key):
    spec = REGISTRY[spec_key]
    errs = validate(spec)
    assert not errs, f"{spec_key} invalid: {errs}"
    S = get_strings(spec.language)   # Phase 13h: UI chrome for this language
    ds = DS(Theme(primary=spec.palette["primary"], accent=spec.palette["accent"], ink=spec.palette["ink"]))
    wb = Workbook(); wb.remove(wb.active)
    plan = spec.sheet_plan()
    names = {mtype: f"{num} · {title}" for num, mtype, title in plan}
    for nm in names.values():
        assert len(nm) <= 31, f"tab too long: {nm!r}"
    mod = {m.type: m for m in spec.modules}
    build_method(wb.create_sheet(names["METHOD"]), ds, mod["METHOD"].terms, spec, S)
    build_planner(wb.create_sheet(names["PLANNER"]), ds, mod["PLANNER"].terms, S)
    anchors = {}
    for num, mtype, title in plan:
        if mtype in FIXED:
            continue
        ws = wb.create_sheet(names[mtype])
        cells = OPS[mtype](ws, ds, mod[mtype].terms, names[mtype], S)
        anchors[mtype] = {metric: refc(names[mtype], cell) for metric, cell in cells.items()}
    dws = wb.create_sheet(names["DASHBOARD"])
    build_dashboard(dws, ds, mod["DASHBOARD"].terms, anchors, S)
    # legibility guarantee: every label/title/header fully readable, no manual fiddling
    for ws in wb.worksheets:
        ds.fit(ws)
    order = [names["METHOD"], names["PLANNER"], names["DASHBOARD"]] + \
            [names[mt] for _, mt, _ in plan if mt not in FIXED]
    wb._sheets = [wb[n] for n in order]
    wb.properties.title = f"{spec.vertical} ({spec.language.upper()})"
    wb.properties.creator = "ASSET-FORGE"
    out = os.path.join(PRODUCTS, f"pack_{spec_key}.xlsx")
    wb.save(out)
    print(f"✓ {spec_key}: {out}  ({len(wb.sheetnames)} sheets) — {' | '.join(wb.sheetnames)}")
    return out


if __name__ == "__main__":
    keys = sys.argv[1:] or list(REGISTRY)
    for k in keys:
        build(k)
