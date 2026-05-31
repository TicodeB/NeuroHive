#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 13b — faithful PNG preview renderer
=======================================================
soffice/LibreOffice is broken in this sandbox, so this draws a sheet straight
from the workbook's REAL cells (values, fills, fonts, merges, widths, heights)
with Pillow. It is faithful — no mock data — so formula cells (whose values are
only computed when Excel opens the file) render blank, i.e. the clean
empty-state of the template.

Run:  python3 scripts/render_preview.py
Out:  products/preview/<sheet>.png
"""
from __future__ import annotations
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries, coordinate_from_string
from PIL import Image, ImageDraw, ImageFont

HERE = os.path.dirname(os.path.abspath(__file__))
PRODUCTS = os.path.normpath(os.path.join(HERE, "..", "products"))
OUT = os.path.join(PRODUCTS, "preview")
os.makedirs(OUT, exist_ok=True)
FONT_DIR = "/usr/share/fonts/truetype/dejavu"
SCALE = 2  # supersample for crisp text

def font(size, bold=False, light=False):
    name = "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"
    return ImageFont.truetype(os.path.join(FONT_DIR, name), int(size * 1.32 * SCALE))

def hexcol(c, default=None):
    if c is None: return default
    rgb = getattr(c, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6: return default
    h = rgb[-6:]
    if h.upper() == "000000" and getattr(c, "type", None) == "theme": return default
    return "#" + h

def col_px(ws, idx):
    w = ws.column_dimensions[get_column_letter(idx)].width or 8.43
    return int(round(w * 7 + 5)) * SCALE

def row_px(ws, idx):
    h = ws.row_dimensions[idx].height or 15
    return int(round(h * 96 / 72)) * SCALE

def render(path, sheet, max_row, max_col, outfile):
    wb = load_workbook(path)
    ws = wb[sheet]
    xs = [0]
    for c in range(1, max_col + 1): xs.append(xs[-1] + col_px(ws, c))
    ys = [0]
    for r in range(1, max_row + 1): ys.append(ys[-1] + row_px(ws, r))
    W, H = xs[-1], ys[-1]
    img = Image.new("RGB", (W, H), "#FFFFFF")
    d = ImageDraw.Draw(img)

    # merged map: top-left -> (r2,c2); covered -> skip
    merged_tl, covered = {}, set()
    for m in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(m))
        merged_tl[(r1, c1)] = (r2, c2)
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if (rr, cc) != (r1, c1): covered.add((rr, cc))

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in covered: continue
            cell = ws.cell(r, c)
            r2, c2 = merged_tl.get((r, c), (r, c))
            x0, y0, x1, y1 = xs[c-1], ys[r-1], xs[min(c2, max_col)], ys[min(r2, max_row)]
            # fill
            fg = None
            if cell.fill is not None and cell.fill.patternType == "solid":
                fg = hexcol(cell.fill.fgColor)
            if fg: d.rectangle([x0, y0, x1-1, y1-1], fill=fg)
            # hairline bottom border (faithful to design)
            if cell.border and cell.border.bottom and cell.border.bottom.style:
                bc = hexcol(cell.border.bottom.color, "#DCE3ED")
                d.line([x0, y1-1, x1-1, y1-1], fill=bc, width=SCALE)
            # text (skip formula strings — they render blank, i.e. empty-state)
            v = cell.value
            if v is None or (isinstance(v, str) and v.startswith("=")): continue
            txt = str(v)
            fnt = font(cell.font.size or 11, bold=bool(cell.font.bold))
            col = hexcol(cell.font.color, "#1A2B45")
            ha = (cell.alignment.horizontal or "general")
            pad = 5 * SCALE
            bbox = d.textbbox((0, 0), txt, font=fnt)
            tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
            if ha == "center": tx = x0 + ((x1-x0)-tw)//2
            elif ha == "right": tx = x1 - tw - pad
            else: tx = x0 + pad + (cell.alignment.indent or 0)*4*SCALE
            ty = y0 + ((y1-y0)-th)//2 - bbox[1]
            d.text((tx, ty), txt, font=fnt, fill=col)

    if SCALE != 1:
        img = img.resize((W // SCALE, H // SCALE), Image.LANCZOS)
    img.save(outfile)
    print("✓", outfile, f"({img.width}x{img.height})")

if __name__ == "__main__":
    p = os.path.join(PRODUCTS, "P2_SK_Hospitality_Premium.xlsx")
    render(p, "00 · Metóda", 30, 7, os.path.join(OUT, "00_metoda.png"))
    render(p, "01 · Denný plán", 36, 5, os.path.join(OUT, "01_denny_plan.png"))
    render(p, "02 · Prehľad", 22, 12, os.path.join(OUT, "02_prehlad.png"))
    render(p, "04 · Marža", 20, 8, os.path.join(OUT, "04_marza.png"))
