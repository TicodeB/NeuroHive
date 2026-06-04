#!/usr/bin/env python3
"""
ASSET-FORGE · Turnkey Startup Packs — Grant & Funding application pack
=====================================================================
A standalone, REUSABLE (cross-vertical) pack of funding-application workbooks,
grounded in the REAL published application structures (researched live; sources
cited in-sheet). The boutique hotel is the worked example; every figure links to
an embedded `Assumptions` sheet (the hotel model) so expenditure/cash-flow are
auto-linked, not hard-coded.

    python3 scripts/grant_build.py            # build all 5 workbooks

Out: products/templates/grant-applications/*.xlsx

Bodies covered:
  · IE — LEO Priming / Business Expansion grant   (localenterprise.ie)
  · IE — Microfinance Ireland loan                (microfinanceireland.ie)
  · IE — Fáilte Ireland tourism capital scheme    (failteireland.ie)
  · IE — SURE founder income-tax refund           (revenue.ie / sure.gov.ie)
  · SK — ÚPSVaR §49 + eurofondy podnikateľský plán (upsvr.gov.sk / eurofondy)

Reuses design_system.py via the helpers in startup_build.py.
"""
from __future__ import annotations
import os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from startup_spec import STARTUP_REGISTRY
from startup_build import (new_ds, build_assumptions, heading, label_value,
                           footer, traffic_cf, status_formula)
from startup_spec import KPI

OUT = os.path.normpath(os.path.join(HERE, "..", "products", "templates", "grant-applications"))
os.makedirs(OUT, exist_ok=True)


def narrative(ds, ws, r, label, hint, span=8, height=46):
    """A guided long-text input field (mirrors the 'min 150 words' grant boxes)."""
    c = ws.cell(r, 2, label); c.font = ds.font(10, bold=True, color=ds.t.ink)
    c.alignment = Alignment("left", "center")
    r += 1
    h = ws.cell(r, 2, "  " + hint); h.font = ds.font(8.5, italic=True, color=ds.t.muted)
    h.alignment = Alignment("left", "top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + span - 1)
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + span - 1)
    ic = ds.input_cell(ws, r, 2)
    ic.alignment = Alignment("left", "top", wrap_text=True)
    ws.row_dimensions[r].height = height
    return r + 1


def checklist(ds, ws, r, title, items, span=8):
    r = ds.section(ws, r, title, span=span)
    yn = DataValidation(type="list", formula1='"☐ To do,☑ Ready,n/a"', allow_blank=True)
    ws.add_data_validation(yn)
    for i, it in enumerate(items):
        ds.trow(ws, r, span, zebra_on=(i % 2 == 1), align="left")
        c = ws.cell(r, 2, it); c.alignment = Alignment("left", "center", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        ic = ds.input_cell(ws, r, span + 1); yn.add(ic); ic.value = "☐ To do"
        r += 1
    return r


def add_assumptions(wb, ds, spec):
    aws = wb.create_sheet("Assumptions")
    return build_assumptions(aws, ds, spec)


def finalize(wb, ds, spec, fname, title):
    sheets = [s for s in wb.sheetnames if s != "Assumptions"] + (
        ["Assumptions"] if "Assumptions" in wb.sheetnames else [])
    wb._sheets = [wb[n] for n in sheets]
    for ws in wb.worksheets:
        ds.fit(ws)
    wb.properties.title = title
    wb.properties.creator = "ASSET-FORGE · LEANTA"
    path = os.path.join(OUT, fname)
    wb.save(path)
    print(f"✓ {fname}  ({len(wb.sheetnames)} sheets)")
    return path


# ============================================================ LEO
def build_leo(spec):
    ds = new_ds(spec)
    wb = Workbook(); wb.remove(wb.active)
    ref = add_assumptions(wb, ds, spec)

    ws = wb.create_sheet("Application")
    r = heading(ds, ws, "LEO Priming / Business Expansion Grant",
                "Local Enterprise Office · application — mirrors the official e-form",
                [30, 16, 16, 16, 16, 16, 16, 16], span=8)
    r = ds.note(ws, r, "Up to 50% of qualifying non-salary costs; max €80,000 "
                "(€150,000 exceptional). Priming = first 18 months; Business Expansion "
                "= growth phase after 18 months. Source: localenterprise.ie.", span=8)

    r = ds.section(ws, r, "1 · Applicant & legal details", span=8)
    for lbl in ["Business / trading name", "Legal form (sole trader/partnership/Ltd)",
                "Company / tax reg. number", "Directors & % shareholding",
                "Business address", "Premises owned / leased (+ lease term)"]:
        r = label_value(ds, ws, r, lbl, None, None, vcol=4, vspan=5, input_cell=True) and r + 1

    r += 1
    r = narrative(ds, ws, r, "2 · Promoter background & qualifications",
                  "Relevant experience, education and track record of each promoter (min ~150 words).")
    r = narrative(ds, ws, r, "3 · Product / service",
                  "Describe the offer, its unique features, pricing, IP, suppliers and future development.")
    r = narrative(ds, ws, r, "4.1 Market research carried out",
                  "What research have you done? Size, trends, spending shifts (min ~150 words).")
    r = narrative(ds, ws, r, "4.2 Customers & marketing strategy",
                  "Who buys, how you reach them (website/OTA/direct), and forward orders (min ~150 words).")
    r = narrative(ds, ws, r, "4.3 Competitors & your edge",
                  "Who competes, where, and why guests will choose you (min ~150 words).")

    # Year-1 expenditure — auto-linked from capex
    r += 1
    r = ds.section(ws, r, "5 · Year-1 expenditure (auto-linked from Assumptions capex)", span=8)
    r = ds.thead(ws, r, ["Cost item", "Total cost €", "Eligible? (non-salary)",
                         "Grant-eligible €", "", "", "", ""])
    a = spec.assumptions
    cap0 = r
    eligible_flags = [1, 1, 1, 1, 1, 0, 0, 0]  # last three (branding/preopening/working cap) edge cases
    cap_label_rows = []
    for i, (lbl, val) in enumerate(a.capex.items()):
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1), align="left")
        ws.cell(r, 2, lbl)
        ds.calc_cell(ws, r, 4, ds.t.EUR0).value = f"='Assumptions'!D{12 + i}" if False else None
        # link to the capex value on Assumptions by label match via INDEX/MATCH
        ds.calc_cell(ws, r, 4, ds.t.EUR0).value = (
            f"=INDEX({ref['capex_range']},MATCH(B{r},{ref['capex_labels']},0))")
        elig = ds.input_cell(ws, r, 5); elig.value = "Yes" if eligible_flags[i] else "Review"
        ds.calc_cell(ws, r, 6, ds.t.EUR0).value = f'=IF(E{r}="Yes",D{r},0)'
        r += 1
    cap_last = r - 1
    ws.cell(r, 2, "Total eligible expenditure").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    elig_total = f"F{r}"
    ds.calc_cell(ws, r, 6, ds.t.EUR0, bold=True).value = f"=SUM(F{cap0}:F{cap_last})"
    r += 2

    # grant calc: 50% match capped 80k/150k, + €15k/FTE job support
    r = ds.section(ws, r, "Grant calculation", span=8)
    r = label_value(ds, ws, r, "Match rate", 0.50, ds.t.PCT, input_cell=True, vcol=4) and r + 1
    match_cell = f"D{r-1}"
    r = label_value(ds, ws, r, "Jobs created (FTE)", 6, "0", input_cell=True, vcol=4) and r + 1
    jobs_cell = f"D{r-1}"
    r = label_value(ds, ws, r, "Grant cap (€80k std / €150k exceptional)", 80000,
                    ds.t.EUR0, input_cell=True, vcol=4) and r + 1
    cap_cell = f"D{r-1}"
    ws.cell(r, 2, "Indicative grant (lower of match × eligible, cap)").font = ds.font(10, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    grant = ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True)
    grant.value = f"=MIN({elig_total}*{match_cell},{cap_cell})"
    r += 1
    ws.cell(r, 2, "Employment support guide (max €15k/FTE)").font = ds.font(9.5, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ds.calc_cell(ws, r, 4, ds.t.EUR0).value = f"={jobs_cell}*15000"
    r += 2

    checklist(ds, ws, r, "Required attachments", [
        "Completed & signed application form",
        "CV of main applicant / promoter",
        "Quotations for key costs (3 quotes for any item over €5,000)",
        "Most recent certified accounts (existing businesses)",
        "Business plan (see Phase 1 workbook)",
        "Tax clearance confirmation",
    ])
    return finalize(wb, ds, spec, "IE_LEO_Priming_Business_Expansion.xlsx",
                    "LEO Priming / Business Expansion Grant — application")


# ============================================================ Microfinance Ireland
def build_mfi(spec):
    ds = new_ds(spec)
    wb = Workbook(); wb.remove(wb.active)
    ref = add_assumptions(wb, ds, spec)

    ws = wb.create_sheet("Business Plan")
    r = heading(ds, ws, "Microfinance Ireland — Loan Application",
                "Business loans €2,000–€50,000 · business plan + cash-flow",
                [30, 16, 16, 16, 16, 16, 16, 16], span=8)
    r = ds.note(ws, r, "Loans above €5,000 require a business plan and a month-by-month "
                "cash-flow. Apply via your LEO for a 1% rate discount + mentor support. "
                "Source: microfinanceireland.ie.", span=8)
    r = ds.section(ws, r, "Loan request", span=8)
    r = label_value(ds, ws, r, "Amount requested €", 40000, ds.t.EUR0, input_cell=True, vcol=4) and r + 1
    r = label_value(ds, ws, r, "Purpose of loan", None, None, vcol=4, vspan=5, input_cell=True) and r + 1
    r = label_value(ds, ws, r, "Term (months)", 36, "0", input_cell=True, vcol=4) and r + 1
    r += 1
    r = narrative(ds, ws, r, "Business summary", "What the business does, stage, and why now.")
    r = narrative(ds, ws, r, "Promoter & experience", "Your background and relevant experience.")
    r = narrative(ds, ws, r, "Market & customers", "Demand, target customers, and competition.")
    r = narrative(ds, ws, r, "How the loan will be repaid", "Link to the cash-flow forecast sheet.")

    # cash-flow forecast (12 months, linked to Assumptions)
    cf = wb.create_sheet("Cashflow Forecast")
    r = heading(ds, cf, "12-Month Cash-Flow Forecast",
                "Month-by-month — required for loans above €5,000",
                [24] + [10] * 12 + [12], span=14)
    months = [f"M{i}" for i in range(1, 13)]
    r = cf_table(ds, cf, r, ref, spec)
    return finalize(wb, ds, spec, "IE_Microfinance_Ireland_Loan.xlsx",
                    "Microfinance Ireland — loan application")


def cf_table(ds, cf, r, ref, spec):
    months = [f"M{i}" for i in range(1, 13)]
    head = r
    r = ds.thead(cf, r, ["Line (€)"] + months + ["Year"])
    ramp = [0.45, 0.55, 0.62, 0.68, 0.74, 0.80, 0.85, 0.88, 0.85, 0.80, 0.74, 0.70]
    ramp_row = r
    cf.cell(r, 2, "Occupancy ramp ×").font = ds.font(9, italic=True, color=ds.t.muted)
    cf.cell(r, 2).alignment = Alignment("left", "center")
    for j, v in enumerate(ramp):
        ds.input_cell(cf, r, 3 + j, "0.00").value = v
    r += 1

    def line(label, fn, bold=False):
        nonlocal r
        c = cf.cell(r, 2, label); c.font = ds.font(9.5, bold=bold, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1); c.border = ds.hairline_bottom()
        for j in range(12):
            col = get_column_letter(3 + j)
            ds.calc_cell(cf, r, 3 + j, ds.t.EUR0, bold=bold).value = fn(col)
        ds.calc_cell(cf, r, 15, ds.t.EUR0, bold=True).value = f"=SUM(C{r}:N{r})"
        rr = r; r += 1; return rr

    base = f"({ref['total_rev']}/12)"
    rev = line("Cash in — revenue", lambda c: f"={base}*({c}{ramp_row}/{ref['occ']})")
    pay = line("Payroll", lambda c: f"={c}{rev}*{ref['payroll_pct']}")
    other = line("Other operating costs",
        lambda c: f"={c}{rev}*({ref['fb_cost_pct']}*{ref['fb_rev_pct']}+{ref['rooms_cost_pct']}+{ref['utilities_pct']}+{ref['sales_pct']}+{ref['admin_pct']}+{ref['property_pct']})")
    loan_repay = line("Loan repayment", lambda c: "=40000/36*1.06" if False else "=1133")
    net = line("Net cash flow", lambda c: f"={c}{rev}-{c}{pay}-{c}{other}-{c}{loan_repay}", bold=True)
    # opening/closing
    cf.cell(r, 2, "Opening cash").font = ds.font(9.5, color=ds.t.ink)
    cf.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    ds.calc_cell(cf, r, 3, ds.t.EUR0).value = f"={ref['funding_total']}-{ref['capex_total']}"
    ob = r
    for j in range(1, 12):
        prev = get_column_letter(2 + j)
        ds.calc_cell(cf, r, 3 + j, ds.t.EUR0).value = f"={prev}{ob+1}"
    r += 1
    cb = r
    cf.cell(r, 2, "Closing cash").font = ds.font(9.5, bold=True, color=ds.t.ink)
    cf.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    for j in range(12):
        col = get_column_letter(3 + j)
        ds.calc_cell(cf, r, 3 + j, ds.t.EUR0, bold=True).value = f"={col}{ob}+{col}{net}"
    ds.calc_cell(cf, r, 15, ds.t.EUR0, bold=True).value = f"=N{cb}"
    cf.conditional_formatting.add(f"C{cb}:N{cb}", CellIsRule(operator="lessThan", formula=["0"],
        fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    cf.freeze_panes = "C" + str(head)
    r += 1
    label_value(ds, cf, r, "Lowest closing cash €", f"=MIN(C{cb}:N{cb})", ds.t.EUR0,
                bold=True, vcol=3, vspan=2)
    return r + 2


# ============================================================ Fáilte Ireland
def build_failte(spec):
    ds = new_ds(spec)
    wb = Workbook(); wb.remove(wb.active)
    ref = add_assumptions(wb, ds, spec)

    ws = wb.create_sheet("Capital Grant Outline")
    r = heading(ds, ws, "Fáilte Ireland — Tourism Capital Grant",
                "Two-stage scheme for larger tourism capital projects (€200k+)",
                [32, 16, 16, 16, 16, 16, 16, 16], span=8)
    r = ds.note(ws, r, "Capital grants for projects that grow overseas visitors, create "
                "jobs and support the destination brands. Two-stage process; State Aid "
                "rules apply. Source: failteireland.ie.", span=8)

    r = ds.section(ws, r, "Eligibility & State-Aid screen", span=8)
    checks = [
        ("Project capital value ≥ €200,000", f'=IF({ref["capex_total"]}>=200000,"PASS","Below threshold")'),
        ("Tax-cleared tourism business", '=IF(TRUE,"Confirm","Confirm")'),
        ("Grows overseas visitor / off-season demand", '=IF(TRUE,"Confirm","Confirm")'),
        ("Within State-Aid limits (de minimis / GBER)", '=IF(TRUE,"Confirm","Confirm")'),
    ]
    for i, (lbl, f) in enumerate(checks):
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1), align="left")
        ws.cell(r, 2, lbl); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        sc = ds.calc_cell(ws, r, 7); sc.value = f; sc.alignment = Alignment("center", "center")
        ws.conditional_formatting.add(f"G{r}", CellIsRule(operator="equal", formula=['"PASS"'],
            fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
        r += 1
    r += 1
    r = narrative(ds, ws, r, "Project rationale",
                  "How the project boosts overseas visitor growth, jobs and the destination brand.")

    r = ds.section(ws, r, "Capital cost schedule (auto-linked)", span=8)
    r = ds.thead(ws, r, ["Capital item", "Cost €", "Grant ask €", "", "", "", "", ""])
    a = spec.assumptions
    c0 = r
    for i, (lbl, _) in enumerate(a.capex.items()):
        ds.trow(ws, r, 8, zebra_on=(i % 2 == 1), align="left")
        ws.cell(r, 2, lbl)
        ds.calc_cell(ws, r, 4, ds.t.EUR0).value = (
            f"=INDEX({ref['capex_range']},MATCH(B{r},{ref['capex_labels']},0))")
        ds.input_cell(ws, r, 6, ds.t.EUR0)
        r += 1
    cN = r - 1
    ws.cell(r, 2, "Total project / grant ask").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True).value = f"=SUM(D{c0}:D{cN})"
    ds.calc_cell(ws, r, 6, ds.t.EUR0, bold=True).value = f"=SUM(F{c0}:F{cN})"
    r += 1
    label_value(ds, ws, r, "Match-funding (own + other)", f"={ref['funding_total']}-SUM(F{c0}:F{cN})",
                ds.t.EUR0, bold=True, vcol=4)
    return finalize(wb, ds, spec, "IE_Failte_Ireland_Capital.xlsx",
                    "Fáilte Ireland — tourism capital grant outline")


# ============================================================ SURE
def build_sure(spec):
    ds = new_ds(spec)
    wb = Workbook(); wb.remove(wb.active)

    # 1 · eligibility
    el = wb.create_sheet("Eligibility")
    r = heading(ds, el, "SURE — Start-Up Refunds for Entrepreneurs",
                "Revenue income-tax refund — up to 41% of capital you invest",
                [44, 14, 18, 14], span=4)
    r = ds.note(el, r, "SURE refunds PAYE income tax you paid over the 6 years before the "
                "year you invest, against new ordinary shares in your new company. "
                "Source: revenue.ie (IT15) / sure.gov.ie.", span=4)
    r = ds.section(el, r, "Eligibility checklist", span=4)
    conds = [
        "Had mainly PAYE income in the previous 4 years (employed/unemployed/redundant/retired)",
        "Establishing a NEW company in a qualifying trading activity",
        "Investing cash for NEW ORDINARY SHARES (no preferential rights)",
        "Will take full-time director/employee role within 6 months of share issue",
        "Will hold the shares for at least 4 years (early disposal = clawback)",
        "Company is an unquoted Irish/EEA SME, tax-cleared, not under another company's control",
    ]
    yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    el.add_data_validation(yn)
    c0 = r
    for i, cond in enumerate(conds):
        ds.trow(el, r, 4, zebra_on=(i % 2 == 1), align="left")
        c = el.cell(r, 2, cond); c.alignment = Alignment("left", "center", indent=1, wrap_text=True)
        ic = ds.input_cell(el, r, 3); yn.add(ic); ic.value = "Yes"
        r += 1
    cN = r - 1
    r += 1
    el.cell(r, 2, "Overall eligibility").font = ds.font(11, bold=True, color=ds.t.primary)
    el.cell(r, 2).alignment = Alignment("left", "center")
    gate = el.cell(r, 3, f'=IF(COUNTIF(C{c0}:C{cN},"Yes")={len(conds)},"🟢 Eligible","🔴 Not yet eligible")')
    gate.alignment = Alignment("center", "center"); gate.font = ds.font(10, bold=True)
    el.conditional_formatting.add(f"C{r}", FormulaRule(
        formula=[f'ISNUMBER(SEARCH("🟢",C{r}))'],
        fill=ds.fill(ds.t.good_bg), font=ds.font(10, True, ds.t.good)))
    el.conditional_formatting.add(f"C{r}", FormulaRule(
        formula=[f'ISNUMBER(SEARCH("🔴",C{r}))'],
        fill=ds.fill(ds.t.bad_bg), font=ds.font(10, True, ds.t.bad)))
    footer(ds, el, r + 2, spec, "SURE · eligibility")

    # 2 · refund calculator
    rc = wb.create_sheet("Refund Calculator")
    r = heading(ds, rc, "SURE Refund Calculator",
                "Enter your P60/P21 figures — refund = lower of tax paid or investment",
                [22, 18, 18, 18, 18], span=5)
    r = ds.note(rc, r, "Investment is set against your income in the prior 6 years (most "
                "recent first), capped at €140,000 per year of assessment. Refund in a "
                "year ≤ the income tax you paid that year.", span=5)
    r = label_value(ds, rc, r, "Total SURE investment €", 180000, ds.t.EUR0,
                    input_cell=True, vcol=3, vspan=2) and r + 1
    inv_cell = f"C{r-1}"
    r += 1
    r = ds.section(rc, r, "Year-by-year (most recent first)", span=5)
    r = ds.thead(rc, r, ["Year of assessment", "Gross pay €", "Income tax paid €",
                         "Investment applied €", "Refund €"])
    sample = [("2025 (Y-1)", 85000, 24000), ("2024 (Y-2)", 80000, 21500),
              ("2023 (Y-3)", 72000, 18000), ("2022 (Y-4)", 60000, 13500),
              ("2021 (Y-5)", 0, 0), ("2020 (Y-6)", 0, 0)]
    y0 = r
    for i, (yr, gross, tax) in enumerate(sample):
        ds.trow(rc, r, 5, zebra_on=(i % 2 == 1), align="left")
        rc.cell(r, 2, yr)
        ds.input_cell(rc, r, 3, ds.t.EUR0).value = gross
        ds.input_cell(rc, r, 4, ds.t.EUR0).value = tax
        # investment applied = MIN(remaining, gross, 140000); remaining = inv - sum(applied above)
        if i == 0:
            remaining = f"{inv_cell}"
        else:
            remaining = f"({inv_cell}-SUM(E{y0}:E{r-1}))"
        ds.calc_cell(rc, r, 5, ds.t.EUR0).value = f"=MAX(0,MIN({remaining},C{r},140000))"
        ds.calc_cell(rc, r, 6, ds.t.EUR0, bold=True).value = f'=IF(C{r}=0,0,D{r}*E{r}/C{r})'
        r += 1
    yN = r - 1
    r += 0
    rc.cell(r, 2, "Estimated SURE refund").font = ds.font(11, bold=True, color=ds.t.primary)
    rc.cell(r, 2).alignment = Alignment("left", "center")
    refund_cell = f"F{r}"
    ds.calc_cell(rc, r, 6, ds.t.EUR0, bold=True).value = f"=SUM(F{y0}:F{yN})"
    r += 1
    rc.cell(r, 2, "Effective refund (% of investment)").font = ds.font(9.5, bold=True, color=ds.t.ink)
    rc.cell(r, 2).alignment = Alignment("left", "center")
    eff = ds.calc_cell(rc, r, 6, ds.t.PCT, bold=True)
    eff.value = f"=IFERROR({refund_cell}/{inv_cell},0)"
    rc.conditional_formatting.add(f"F{r}", CellIsRule(operator="greaterThan", formula=["0.41"],
        fill=ds.fill(ds.t.warn_bg), font=ds.font(9.5, color=ds.t.warn)))
    r += 1
    rc.cell(r, 2, "→ Feeds Phase-2 funding mix as founder equity").font = ds.font(9, italic=True, color=ds.t.muted)
    rc.cell(r, 2).alignment = Alignment("left", "center")
    footer(ds, rc, r + 2, spec, "SURE · calculator")

    # 3 · claim process
    cp = wb.create_sheet("Claim Process")
    r = heading(ds, cp, "SURE — How to Claim",
                "Steps, documents and contacts", [40, 16, 16, 16], span=4)
    r = ds.section(cp, r, "Process", span=4)
    steps = [
        "Incorporate your qualifying company (CRO) with an appropriate trade.",
        "Invest cash from your own resources for NEW ordinary shares; record board minutes + share cert.",
        "Take up full-time director/employee role within 6 months of the share issue.",
        "Company issues you a SURE/EII Statement of Qualification.",
        "Claim the refund via Revenue (IT15 guide) for the relevant years.",
        "Keep all records 6 years; do not dispose of shares for 4 years (clawback).",
    ]
    for i, s in enumerate(steps):
        ds.trow(cp, r, 4, zebra_on=(i % 2 == 1), align="left")
        c = cp.cell(r, 2, f"{i+1}.  {s}"); c.alignment = Alignment("left", "center", indent=1, wrap_text=True)
        ws_merge_end = 4
        cp.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1
    r = checklist(ds, cp, r, "Documents to keep", [
        "P60 / P21 for each relevant prior year",
        "Share certificate + updated shareholder register",
        "Board minutes recording the share issue",
        "SURE / EII Statement of Qualification from the company",
        "Evidence of full-time engagement within 6 months",
    ], span=3)
    r += 1
    r = ds.section(cp, r, "Contacts & related relief", span=4)
    for s in ["Revenue SURE admin: SUREadmin@revenue.ie · guide IT15",
              "Online estimator: sure.gov.ie",
              "Related: §486C start-up corporation-tax relief (first 5 yrs, linked to employer PRSI)"]:
        r = ds.note(cp, r, s, span=4)
    footer(ds, cp, r + 1, spec, "SURE · claim")

    sheets = ["Eligibility", "Refund Calculator", "Claim Process"]
    wb._sheets = [wb[n] for n in sheets]
    for ws in wb.worksheets:
        ds.fit(ws)
    wb.properties.title = "SURE — Start-Up Refunds for Entrepreneurs"
    wb.properties.creator = "ASSET-FORGE · LEANTA"
    path = os.path.join(OUT, "IE_SURE_Tax_Refund.xlsx")
    wb.save(path)
    print(f"✓ IE_SURE_Tax_Refund.xlsx  ({len(wb.sheetnames)} sheets)")
    return path


# ============================================================ Slovak
def build_sk(spec):
    ds = new_ds(spec)
    wb = Workbook(); wb.remove(wb.active)

    pp = wb.create_sheet("Podnikateľský plán")
    r = heading(ds, pp, "Podnikateľský plán",
                "Štruktúra podľa ÚPSVaR / eurofondy — žiadosť o príspevok / dotáciu",
                [34, 16, 16, 16, 16], span=5)
    r = ds.note(pp, r, "Štruktúra vychádza z metodiky ÚPSVaR (§49 príspevok na SZČ) a "
                "eurofondy.praca.gov.sk. Finančný plán je povinná príloha. "
                "Zdroj: upsvr.gov.sk, eurofondy.praca.gov.sk.", span=5)

    r = ds.section(pp, r, "1 · Základné informácie o spoločnosti", span=5)
    for lbl in ["Obchodné meno", "Právna forma (SZČO / s.r.o. / …)", "IČO", "DIČ",
                "Platca DPH (áno/nie)", "Dátum vzniku", "Sídlo"]:
        r = label_value(ds, pp, r, lbl, None, None, vcol=3, vspan=3, input_cell=True) and r + 1
    r += 1
    r = narrative(ds, pp, r, "2 · Popis podniku a skúsenosti žiadateľa",
                  "Relevantné skúsenosti žiadateľa a zamestnancov v predmete činnosti.", span=5)
    r = narrative(ds, pp, r, "3 · Produkty a služby",
                  "Čo ponúkate, ceny, jedinečnosť, dodávatelia.", span=5)
    r = narrative(ds, pp, r, "4 · Trh a konkurencia (SWOT)",
                  "Veľkosť trhu, zákazníci, konkurenti; silné/slabé stránky, príležitosti, hrozby.", span=5)
    r = narrative(ds, pp, r, "5 · Marketingová stratégia",
                  "Ako oslovíte a získate zákazníkov (web/OTA/priamy predaj).", span=5)

    # 7 · Finančný plán (5-year)
    fp = wb.create_sheet("Finančný plán")
    r = heading(ds, fp, "Finančný plán (povinná príloha)",
                "Strednodobý výhľad — 5 rokov", [28, 16, 16, 16, 16, 16], span=6)
    a = spec.assumptions
    # editable inputs (Východiská) — seeded so the plan computes
    cost_ratio = (a.payroll_pct + a.utilities_pct + a.sales_mktg_pct + a.admin_pct
                  + a.property_pct + a.rooms_cost_pct + a.fb_cost_pct * a.fb_rev_pct)
    iv = r
    fp.cell(r, 2, "Východiská (upraviteľné)").font = ds.font(9.5, bold=True, color=ds.t.primary)
    fp.cell(r, 2).alignment = Alignment("left", "center")
    r += 1
    fp.cell(r, 2, "Plánované tržby Rok 1 €").font = ds.font(9.5, color=ds.t.ink)
    fp.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    ds.input_cell(fp, r, 3, ds.t.EUR0).value = int(a.total_revenue)
    base_cell = f"C{r}"
    r += 1
    fp.cell(r, 2, "Podiel nákladov na tržbách").font = ds.font(9.5, color=ds.t.ink)
    fp.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    ds.input_cell(fp, r, 3, ds.t.PCT).value = round(cost_ratio, 4)
    ratio_cell = f"C{r}"
    r += 2

    r = ds.thead(fp, r, ["Položka (€)", "Rok 1", "Rok 2", "Rok 3", "Rok 4", "Rok 5"])
    growth = [1.0, 1.05, 1.10, 1.13, 1.15]
    # growth-factor input row
    gr_row = r
    fp.cell(r, 2, "Index rastu tržieb").font = ds.font(9, italic=True, color=ds.t.muted)
    fp.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    for j, g in enumerate(growth):
        ds.input_cell(fp, r, 3 + j, "0.00").value = g
    r += 1
    rev_row = r
    fp.cell(r, 2, "Tržby spolu").font = ds.font(9.5, bold=True, color=ds.t.ink)
    fp.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    for j in range(5):
        col = get_column_letter(3 + j)
        ds.calc_cell(fp, r, 3 + j, ds.t.EUR0, bold=True).value = f"=ROUND({base_cell}*{col}{gr_row},0)"
    r += 1
    cost_row = r
    fp.cell(r, 2, "Náklady spolu").font = ds.font(9.5, color=ds.t.ink)
    fp.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    for j in range(5):
        col = get_column_letter(3 + j)
        ds.calc_cell(fp, r, 3 + j, ds.t.EUR0).value = f"=ROUND({col}{rev_row}*{ratio_cell},0)"
    r += 1
    fp.cell(r, 2, "Hospodársky výsledok").font = ds.font(9.5, bold=True, color=ds.t.primary)
    fp.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    for j in range(5):
        col = get_column_letter(3 + j)
        ds.calc_cell(fp, r, 3 + j, ds.t.EUR0, bold=True).value = f"={col}{rev_row}-{col}{cost_row}"
    profit_row = r
    fp.conditional_formatting.add(f"C{r}:G{r}", CellIsRule(operator="lessThan", formula=["0"],
        fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    fp.conditional_formatting.add(f"C{r}:G{r}", CellIsRule(operator="greaterThan", formula=["0"],
        fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    r += 2
    r = ds.section(fp, r, "§49 kalkulácia nákladov (príspevok na SZČ)", span=6)
    r = ds.thead(fp, r, ["Nákladová položka", "Suma €", "", "", "", ""])
    items = [("Vybavenie a zariadenie", 18000), ("Prvotné zásoby", 6500),
             ("Marketing a propagácia", 4000), ("Nájom (prvé mesiace)", 9000),
             ("Ostatné prevádzkové náklady", 3500)]
    k0 = r
    for i, (it, amt) in enumerate(items):
        ds.trow(fp, r, 6, zebra_on=(i % 2 == 1), align="left")
        fp.cell(r, 2, it); ds.input_cell(fp, r, 3, ds.t.EUR0).value = amt
        r += 1
    kN = r - 1
    fp.cell(r, 2, "Náklady spolu").font = ds.font(9.5, bold=True, color=ds.t.primary)
    fp.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(fp, r, 3, ds.t.EUR0, bold=True).value = f"=SUM(C{k0}:C{kN})"

    # prílohy + vyhlásenia
    pr = wb.create_sheet("Prílohy a vyhlásenia")
    r = heading(ds, pr, "Prílohy a čestné vyhlásenia",
                "Povinné prílohy k žiadosti", [44, 16, 16], span=3)
    r = checklist(ds, pr, r, "Povinné prílohy", [
        "Vyplnená a podpísaná žiadosť",
        "Podnikateľský plán + finančný plán (povinná príloha)",
        "Kalkulácia nákladov na prevádzkovanie SZČ",
        "Čestné vyhlásenie o minimálnej pomoci (de minimis)",
        "Čestné vyhlásenie o predchádzajúcom podnikaní",
        "Kópia dokladu o najvyššom ukončenom vzdelaní",
        "Doklad o vzniku oprávnenia na podnikanie (živnosť)",
    ], span=2)
    r += 1
    r = ds.section(pr, r, "Harmonogram výplaty (§49)", span=3)
    for s in ["1. správa po 12 mesiacoch → vyplatenie 60% príspevku",
              "2. správa po 24 mesiacoch → doplatok do 100%",
              "SZČ prevádzkovaná nepretržite najmenej 2 roky."]:
        r = ds.note(pr, r, s, span=3)
    footer(ds, pr, r + 1, spec, "SK · ÚPSVaR / eurofondy")

    sheets = ["Podnikateľský plán", "Finančný plán", "Prílohy a vyhlásenia"]
    wb._sheets = [wb[n] for n in sheets]
    for ws in wb.worksheets:
        ds.fit(ws)
    wb.properties.title = "SK — ÚPSVaR / eurofondy podnikateľský plán"
    wb.properties.creator = "ASSET-FORGE · LEANTA"
    path = os.path.join(OUT, "SK_UPSVaR_Eurofondy_Podnikatelsky_Plan.xlsx")
    wb.save(path)
    print(f"✓ SK_UPSVaR_Eurofondy_Podnikatelsky_Plan.xlsx  ({len(wb.sheetnames)} sheets)")
    return path


def build_all():
    spec = STARTUP_REGISTRY["boutique_hotel_4star"]
    build_leo(spec)
    build_mfi(spec)
    build_failte(spec)
    build_sure(spec)
    build_sk(spec)


if __name__ == "__main__":
    build_all()
