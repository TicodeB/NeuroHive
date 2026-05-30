#!/usr/bin/env python3
"""
ASSET-FORGE · Phase 9 (A) — Second flagship
P2 — Hospitality Operations & GP Bundle  (the daily-money layer above P1's floor)

Bilingual (EN / SK) workbook bundling the operations/money assets plus the
Staff Training & Induction Matrix (added at owner request — manual-handling /
induction training must be evidenced before a worker hits the floor):

  id 8  Cashflow & P&L Tracker
  id 7  Recipe & Menu GP Costing Calculator
  id 6  Stock & Wastage Tracker
  id 9  Staff Rota & Labour-Cost Scheduler
  id 11 Daily Takings & Till Reconciliation
  id 10 Staff Training & Induction Matrix      <-- added (people module)

EU conventions: € comma thousands, DD/MM/YYYY, metric. Price €49 / Lemon Squeezy.
Run:  python3 scripts/build_p2_operations_bundle.py
Out:  products/P2_Hospitality_Operations_GP_Bundle.xlsx
      products/P2_DEMO_Hospitality_Operations_GP_Bundle.xlsx
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PRODUCTS = os.path.normpath(os.path.join(HERE, "..", "products"))
os.makedirs(PRODUCTS, exist_ok=True)

VERSION = "v1.0"
BUILD_DATE = "30/05/2026"

NAVY = "1F3A5F"; TEAL = "2A7D7B"; SAND = "F4EEE2"; LBLUE = "DCE6F1"
LGREY = "F2F2F2"; AMBER = "FFF2CC"; GREEN = "E2EFDA"; RED = "F8CBAD"; WHITE = "FFFFFF"
EUR = '#,##0.00\\ "€"'   # EU comma thousands, € suffix
PCT = "0.0%"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=11, b=False, color="000000", italic=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=italic)
def fill(c): return PatternFill("solid", fgColor=c)
def style_header(cell, bg=NAVY, fg=WHITE, sz=10):
    cell.font = f(sz, True, fg); cell.fill = fill(bg)
    cell.alignment = Alignment("center", "center", wrap_text=True); cell.border = BORDER
def set_widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
def header_row(ws, row, headers, bg=NAVY):
    for j, h in enumerate(headers, 1): style_header(ws.cell(row, j, h), bg=bg)
    ws.row_dimensions[row].height = 38
def title_block(ws, en, sk, sub_en, sub_sk, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(1, 1, f"{en}  |  {sk}")
    t.font = f(15, True, WHITE); t.fill = fill(NAVY)
    t.alignment = Alignment("left", "center", indent=1); ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(2, 1, f"{sub_en}  ·  {sub_sk}")
    s.font = f(10, False, NAVY, italic=True); s.fill = fill(SAND)
    s.alignment = Alignment("left", "center", indent=1, wrap_text=True); ws.row_dimensions[2].height = 26
def biz_field(ws, row, en, sk, span=8):
    c = ws.cell(row, 1, f"{en} / {sk}:"); c.font = f(10, True, NAVY)
    c.alignment = Alignment("right", "center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span)
    v = ws.cell(row, 2, ""); v.fill = fill(AMBER); v.border = BORDER
def note(ws, row, en, sk, span=8, bg=LBLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, f"ℹ  {en}\n    {sk}")
    c.font = f(9, False, NAVY, italic=True); c.fill = fill(bg)
    c.alignment = Alignment("left", "top", wrap_text=True, indent=1); ws.row_dimensions[row].height = 30
def zebra(ws, r0, n, ncols, dates=(), money=(), pcts=()):
    for i in range(n):
        r = r0 + i
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c, ""); cell.border = BORDER; cell.font = f(9)
            if i % 2: cell.fill = fill(LGREY)
            if c in dates: cell.number_format = "DD/MM/YYYY"
            if c in money: cell.number_format = EUR
            if c in pcts: cell.number_format = PCT
        ws.row_dimensions[r].height = 18


# ---------------------------------------------------------------- Start Here
def build_cover(wb):
    ws = wb.active; ws.title = "00 · Start Here · Začnite tu"
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 34, 34, 20, 20, 20, 14, 14])
    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, "HOSPITALITY OPERATIONS & GP BUNDLE")
    t.font = f(20, True, WHITE); t.fill = fill(NAVY); t.alignment = Alignment("center", "center")
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A2:H2")
    t2 = ws.cell(2, 1, "Prevádzka a hrubá marža pre gastro — balík nástrojov")
    t2.font = f(13, True, NAVY); t2.fill = fill(SAND); t2.alignment = Alignment("center", "center")
    ws.row_dimensions[2].height = 26
    ws.merge_cells("A3:H3")
    t3 = ws.cell(3, 1, "See your margin, labour % and cash before they bite · Vidíte maržu, mzdy a hotovosť skôr, než zabolia")
    t3.font = f(10, False, TEAL, italic=True); t3.alignment = Alignment("center", "center")
    rows = [
        ("", ""),
        ("WHAT'S INSIDE / ČO OBSAHUJE", "header"),
        ("01 · Cashflow & P&L Tracker", "Sledovač cash flow a výkazu ziskov a strát"),
        ("02 · Recipe & Menu GP Costing Calculator", "Kalkulačka nákladov a hrubej marže receptúr"),
        ("03 · Stock & Wastage Tracker", "Sledovač zásob a strát"),
        ("04 · Staff Rota & Labour-Cost Scheduler", "Plánovač zmien a mzdových nákladov"),
        ("05 · Daily Takings & Till Reconciliation", "Hárok denných tržieb a uzávierky pokladne"),
        ("06 · Staff Training & Induction Matrix", "Matica školení a zaškolenia zamestnancov"),
        ("", ""),
        ("WHY THE TRAINING MATRIX IS HERE / PREČO JE TU MATICA ŠKOLENÍ", "header"),
        ("No worker on the floor without manual-handling & induction training —",
         "Žiadny pracovník na zmenu bez školenia o manipulácii s bremenami a zaškolenia —"),
        ("Sheet 06 is your dated proof of who was trained, on what, and when refreshers fall due.",
         "Hárok 06 je dôkaz: kto, na čo a kedy bol školený a kedy je preškolenie."),
        ("", ""),
        ("HOW TO USE / AKO POUŽÍVAŤ", "header"),
        ("Amber = you fill · Blue = auto-calculated · Grey/white = entries.",
         "Oranžová = vyplníte · Modrá = automaticky · Sivá/biela = záznamy."),
        ("Sits on top of P1 Compliance Pack — buy both as the Hospitality Pro Bundle.",
         "Nadväzuje na P1 Compliance Pack — kúpte spolu ako Hospitality Pro Bundle."),
    ]
    r = 5
    for en, sk in rows:
        if sk == "header":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = ws.cell(r, 1, en); c.font = f(12, True, WHITE); c.fill = fill(TEAL)
            c.alignment = Alignment("left", "center", indent=1); ws.row_dimensions[r].height = 24
        elif en == "" and sk == "":
            ws.row_dimensions[r].height = 6
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
            ce = ws.cell(r, 1, en); ce.font = f(10, True, NAVY)
            ce.alignment = Alignment("left", "center", indent=1, wrap_text=True)
            cs = ws.cell(r, 5, sk); cs.font = f(10, False, "404040", italic=True)
            cs.alignment = Alignment("left", "center", indent=1, wrap_text=True)
            ws.row_dimensions[r].height = 18
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    fc = ws.cell(r, 1, f"ASSET-FORGE · {VERSION} · {BUILD_DATE} · EU (Ireland) · "
                       "Template only — not financial or legal advice. · Šablóna — nie je finančné ani právne poradenstvo.")
    fc.font = f(8, False, "808080", italic=True); fc.alignment = Alignment("center", "center", wrap_text=True)
    ws.row_dimensions[r].height = 26; ws.sheet_properties.tabColor = NAVY


# ---------------------------------------------------------------- 01 Cashflow & P&L
def build_cashflow(wb):
    ws = wb.create_sheet("01 · Cashflow & P&L")
    ws.sheet_view.showGridLines = False
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    ncols = 1 + 12 + 1
    set_widths(ws, [30] + [12] * 12 + [13])
    title_block(ws, "Cashflow & P&L Tracker", "Sledovač cash flow a výkazu ziskov a strát",
                "12-month rolling cash + monthly profit & loss for an owner-run venue",
                "12-mesačný cash flow a mesačný výkaz ziskov a strát", span=ncols)
    biz_field(ws, 3, "Business name", "Názov prevádzky", span=ncols)
    biz_field(ws, 4, "Year (YYYY)", "Rok", span=ncols)
    note(ws, 6, "Enter actuals in the amber cells. Totals, GP, net profit and closing cash calculate automatically.",
         "Vyplňte oranžové bunky. Súčty, marža, zisk a koncová hotovosť sa počítajú automaticky.", span=ncols)

    style_header(ws.cell(7, 1, "Line / Položka"))
    for j, m in enumerate(months, 2): style_header(ws.cell(7, j, m))
    style_header(ws.cell(7, 14, "Year / Rok"))
    ws.row_dimensions[7].height = 22

    def section(r, label):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1, label); c.font = f(10, True, WHITE); c.fill = fill(TEAL)
        c.alignment = Alignment("left", "center", indent=1); ws.row_dimensions[r].height = 18

    def line(r, en, sk, total=True, bold=False, formula=None):
        c = ws.cell(r, 1, f"{en} / {sk}"); c.font = f(9, bold, NAVY if bold else "000000")
        c.alignment = Alignment("left", "center", indent=1); c.border = BORDER
        for j in range(2, 14):
            cell = ws.cell(r, j, "")
            cell.border = BORDER; cell.number_format = EUR; cell.font = f(9)
            if formula:
                cell.value = formula(get_column_letter(j))
                cell.fill = fill(LBLUE)
            else:
                cell.fill = fill(AMBER)
        tot = ws.cell(r, 14, f"=SUM(B{r}:M{r})")
        tot.border = BORDER; tot.number_format = EUR; tot.font = f(9, True); tot.fill = fill(LBLUE)
        ws.row_dimensions[r].height = 16

    r = 8
    section(r, "REVENUE / TRŽBY"); r += 1
    rev_rows = []
    for en, sk in [("Food sales", "Tržby z jedál"), ("Drink sales", "Tržby z nápojov"), ("Other income", "Ostatné príjmy")]:
        line(r, en, sk); rev_rows.append(r); r += 1
    rev_total = r
    line(r, "Total revenue", "Tržby spolu", bold=True,
         formula=lambda col: f"=SUM({col}{rev_rows[0]}:{col}{rev_rows[-1]})"); r += 1

    section(r, "COST OF SALES / NÁKLADY NA PREDAJ"); r += 1
    cos_rows = []
    for en, sk in [("Food cost", "Náklady na jedlo"), ("Drink cost", "Náklady na nápoje")]:
        line(r, en, sk); cos_rows.append(r); r += 1
    cos_total = r
    line(r, "Total cost of sales", "Náklady na predaj spolu", bold=True,
         formula=lambda col: f"=SUM({col}{cos_rows[0]}:{col}{cos_rows[-1]})"); r += 1
    gp_row = r
    line(r, "GROSS PROFIT", "HRUBÁ MARŽA", bold=True,
         formula=lambda col: f"={col}{rev_total}-{col}{cos_total}"); r += 1

    section(r, "OVERHEADS / RÉŽIE"); r += 1
    oh_rows = []
    for en, sk in [("Wages & labour", "Mzdy a personál"), ("Rent & rates", "Nájom a poplatky"),
                   ("Utilities", "Energie"), ("Marketing", "Marketing"), ("Other overheads", "Ostatné réžie")]:
        line(r, en, sk); oh_rows.append(r); r += 1
    oh_total = r
    line(r, "Total overheads", "Réžie spolu", bold=True,
         formula=lambda col: f"=SUM({col}{oh_rows[0]}:{col}{oh_rows[-1]})"); r += 1
    np_row = r
    line(r, "NET PROFIT", "ČISTÝ ZISK", bold=True,
         formula=lambda col: f"={col}{gp_row}-{col}{oh_total}"); r += 1

    section(r, "CASH POSITION / HOTOVOSŤ"); r += 1
    # opening cash: Jan amber, others = prior closing
    ob_row = r
    ws.cell(r, 1, "Opening cash / Počiatočná hotovosť").font = f(9); ws.cell(r, 1).border = BORDER
    ws.cell(r, 1).alignment = Alignment("left", "center", indent=1)
    ws.cell(r, 2).fill = fill(AMBER); ws.cell(r, 2).border = BORDER; ws.cell(r, 2).number_format = EUR
    cb_row = ob_row + 1
    for j in range(3, 14):
        prev = get_column_letter(j - 1)
        cell = ws.cell(r, j, f"={prev}{cb_row}"); cell.fill = fill(LBLUE)
        cell.border = BORDER; cell.number_format = EUR; cell.font = f(9)
    ws.cell(r, 14).value = f"=B{r}"; ws.cell(r, 14).border = BORDER; ws.cell(r, 14).number_format = EUR
    ws.row_dimensions[r].height = 16; r += 1
    ws.cell(r, 1, "Closing cash / Koncová hotovosť").font = f(9, True); ws.cell(r, 1).border = BORDER
    ws.cell(r, 1).alignment = Alignment("left", "center", indent=1)
    for j in range(2, 14):
        col = get_column_letter(j)
        cell = ws.cell(r, j, f"={col}{ob_row}+{col}{np_row}"); cell.fill = fill(LBLUE)
        cell.border = BORDER; cell.number_format = EUR; cell.font = f(9, True)
    ws.cell(r, 14).value = f"=M{r}"; ws.cell(r, 14).border = BORDER; ws.cell(r, 14).number_format = EUR
    ws.row_dimensions[r].height = 16
    ws.conditional_formatting.add(f"B{r}:M{r}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED), font=f(9, True, "9C0006")))
    ws.sheet_properties.tabColor = "2E75B6"; ws.freeze_panes = "B8"


# ---------------------------------------------------------------- 02 GP Costing
def build_gp(wb):
    ws = wb.create_sheet("02 · GP Costing")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [30, 14, 12, 14, 14, 14, 12, 16])
    title_block(ws, "Recipe & Menu GP Costing Calculator", "Kalkulačka nákladov a hrubej marže receptúr",
                "Plate cost, gross-profit % and suggested price per dish or drink",
                "Náklady na porciu, hrubá marža (%) a odporúčaná cena")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    note(ws, 5, "Enter ingredient cost and your target GP%. The suggested ex-VAT price is calculated; GP% on your actual menu price is shown alongside.",
         "Zadajte náklady a cieľovú maržu. Odporúčaná cena (bez DPH) sa vypočíta; vedľa vidíte maržu pri vašej reálnej cene.")
    header_row(ws, 6, [
        "Dish / drink\nJedlo / nápoj", "Portion size\nVeľkosť porcie", "Ingredient cost €\nNáklady €",
        "Target GP %\nCieľová marža %", "Suggested price € (ex-VAT)\nOdpor. cena €", "Menu price € (ex-VAT)\nCena v menu €",
        "Actual GP %\nReálna marža %", "Flag\nUpozornenie",
    ])
    seed = [
        ("Brown soda bread basket", "120 g", 0.45, 0.70),
        ("Seafood chowder", "350 ml", 2.10, 0.68),
        ("Beef burger & chips", "300 g", 3.40, 0.68),
        ("House cappuccino", "240 ml", 0.55, 0.85),
        ("Pint craft lager", "568 ml", 1.95, 0.55),
    ]
    r = 7; N = 22
    for i in range(N):
        row = seed[i] if i < len(seed) else ("", "", "", "")
        ws.cell(r, 1, row[0]).fill = fill(WHITE if row[0] else AMBER)
        ws.cell(r, 2, row[1])
        ws.cell(r, 3, row[2] if row[2] != "" else "").number_format = EUR
        ws.cell(r, 4, row[3] if row[3] != "" else "").number_format = PCT
        # suggested price = cost / (1 - target GP)
        ws.cell(r, 5).value = f'=IF(OR(C{r}="",D{r}=""),"",C{r}/(1-D{r}))'
        ws.cell(r, 5).number_format = EUR
        # actual GP% = (menu price - cost)/menu price
        ws.cell(r, 7).value = f'=IF(OR(F{r}="",C{r}=""),"",(F{r}-C{r})/F{r})'
        ws.cell(r, 7).number_format = PCT
        # flag low margin
        ws.cell(r, 8).value = f'=IF(OR(G{r}="",D{r}=""),"",IF(G{r}<D{r},"LOW","OK"))'
        for j in range(1, 9):
            c = ws.cell(r, j); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment("center" if j in (2,3,4,5,6,7,8) else "left", "center", wrap_text=True)
        ws.cell(r, 3).fill = fill(AMBER); ws.cell(r, 4).fill = fill(AMBER); ws.cell(r, 6).fill = fill(AMBER)
        ws.cell(r, 5).fill = fill(LBLUE); ws.cell(r, 7).fill = fill(LBLUE)
        ws.row_dimensions[r].height = 18
        r += 1
    ws.conditional_formatting.add(f"H7:H{6+N}",
        CellIsRule(operator="equal", formula=['"LOW"'], fill=fill(RED), font=f(9, True, "9C0006")))
    ws.conditional_formatting.add(f"H7:H{6+N}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=fill(GREEN)))
    ws.sheet_properties.tabColor = "548235"; ws.freeze_panes = "A7"


# ---------------------------------------------------------------- 03 Stock & Wastage
def build_stock(wb):
    ws = wb.create_sheet("03 · Stock & Wastage")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [28, 12, 14, 14, 14, 14, 14, 16])
    title_block(ws, "Stock & Wastage Tracker", "Sledovač zásob a strát",
                "Opening + purchases − closing = usage; wastage value highlights where GP leaks",
                "Počiatočné + nákup − koncové = spotreba; hodnota strát ukáže únik marže")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Period (DD/MM/YYYY)", "Obdobie")
    note(ws, 6, "Usage and wastage value calculate automatically. High wastage rows turn red so you can act on shrinkage and over-pour.",
         "Spotreba a hodnota strát sa počítajú automaticky. Riadky s vysokými stratami zčervenajú.")
    header_row(ws, 7, [
        "Item\nPoložka", "Unit\nJednotka", "Opening\nPočiatočné", "Purchases\nNákup",
        "Closing\nKoncové", "Wastage qty\nStraty (množ.)", "Unit cost €\nCena/jedn. €", "Wastage value €\nHodnota strát €",
    ])
    seed = [
        ("Draught lager", "L", 80, 220, 60, 4),
        ("House red wine", "btl", 24, 60, 30, 1),
        ("Fresh fish", "kg", 5, 30, 2, 1.5),
        ("Rib-eye steak", "kg", 8, 25, 6, 0.5),
        ("Milk", "L", 20, 120, 25, 3),
    ]
    r = 7  # header at 7, data from 8
    header_row(ws, 7, [
        "Item\nPoložka", "Unit\nJednotka", "Opening\nPočiatočné", "Purchases\nNákup",
        "Closing\nKoncové", "Wastage qty\nStraty (množ.)", "Unit cost €\nCena/jedn. €", "Wastage value €\nHodnota strát €",
    ])
    r = 8; N = 18
    for i in range(N):
        row = seed[i] if i < len(seed) else ("", "", "", "", "", "")
        ws.cell(r, 1, row[0]).fill = fill(WHITE if row[0] else AMBER)
        ws.cell(r, 2, row[1])
        for idx, col in [(2, 3), (3, 4), (4, 5), (5, 6)]:
            ws.cell(r, col, row[idx] if (len(row) > idx and row[idx] != "") else "")
        ws.cell(r, 7, row[5] if (len(row) > 5 and row[5] != "") else "").number_format = EUR
        ws.cell(r, 8).value = f'=IF(OR(F{r}="",G{r}=""),"",F{r}*G{r})'
        ws.cell(r, 8).number_format = EUR
        for j in range(1, 9):
            c = ws.cell(r, j); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment("center" if j > 1 else "left", "center", wrap_text=True)
            if j in (3, 4, 5, 6): c.fill = fill(AMBER)
            if j == 7: c.fill = fill(AMBER)
        ws.cell(r, 8).fill = fill(LBLUE)
        ws.row_dimensions[r].height = 18
        r += 1
    ws.conditional_formatting.add(f"H8:H{7+N}",
        CellIsRule(operator="greaterThan", formula=["50"], fill=fill(RED)))
    ws.sheet_properties.tabColor = "BF8F00"; ws.freeze_panes = "A8"


# ---------------------------------------------------------------- 04 Rota & Labour
def build_rota(wb):
    ws = wb.create_sheet("04 · Rota & Labour")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 12, 12, 12, 14, 16, 14])
    title_block(ws, "Staff Rota & Labour-Cost Scheduler", "Plánovač zmien a mzdových nákladov",
                "Schedule shifts and instantly see labour cost as a % of forecast sales",
                "Naplánujte zmeny a hneď vidíte podiel mzdových nákladov na tržbách")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Week commencing (DD/MM/YYYY)", "Týždeň od")
    # forecast sales input
    ws.cell(5, 1, "Forecast sales € / Plán tržieb €:").font = f(10, True, NAVY)
    ws.cell(5, 1).alignment = Alignment("right", "center")
    fs = ws.cell(5, 2, ""); fs.fill = fill(AMBER); fs.border = BORDER; fs.number_format = EUR
    note(ws, 6, "Enter each shift's hours and hourly rate. Labour cost = hours × rate. Target labour % for hospitality is typically 25–35% of sales.",
         "Zadajte hodiny a hodinovú sadzbu. Mzdové náklady = hodiny × sadzba. Cieľ pre gastro býva 25–35 % tržieb.")
    header_row(ws, 7, [
        "Employee\nZamestnanec", "Role\nPozícia", "Day\nDeň", "Start\nOd", "End\nDo",
        "Hours\nHodiny", "Rate €/h\nSadzba €/h", "Cost €\nNáklad €",
    ])
    days = DataValidation(type="list", formula1='"Mon,Tue,Wed,Thu,Fri,Sat,Sun"', allow_blank=True)
    ws.add_data_validation(days)
    r = 8; N = 20
    for i in range(N):
        for j in range(1, 9):
            c = ws.cell(r, j, ""); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment("center" if j > 2 else "left", "center")
            if i % 2: c.fill = fill(LGREY)
        days.add(ws.cell(r, 3))
        for col in (6, 7): ws.cell(r, col).fill = fill(AMBER)
        ws.cell(r, 8).value = f'=IF(OR(F{r}="",G{r}=""),"",F{r}*G{r})'
        ws.cell(r, 8).number_format = EUR; ws.cell(r, 7).number_format = EUR
        ws.row_dimensions[r].height = 16
        r += 1
    # totals
    tot = r
    ws.cell(r, 1, "TOTAL / SPOLU").font = f(10, True, NAVY); ws.cell(r, 1).border = BORDER
    ws.cell(r, 6).value = f"=SUM(F8:F{r-1})"; ws.cell(r, 6).font = f(10, True); ws.cell(r, 6).border = BORDER
    ws.cell(r, 8).value = f"=SUM(H8:H{r-1})"; ws.cell(r, 8).font = f(10, True); ws.cell(r, 8).border = BORDER
    ws.cell(r, 8).number_format = EUR; r += 1
    ws.cell(r, 1, "LABOUR % OF SALES / PODIEL MIEZD %").font = f(10, True, NAVY); ws.cell(r, 1).border = BORDER
    lp = ws.cell(r, 8, f'=IF(B5="","",H{tot}/B5)'); lp.number_format = PCT; lp.font = f(10, True); lp.border = BORDER
    lp.fill = fill(LBLUE)
    ws.conditional_formatting.add(f"H{r}:H{r}",
        CellIsRule(operator="greaterThan", formula=["0.35"], fill=fill(RED), font=f(10, True, "9C0006")))
    ws.conditional_formatting.add(f"H{r}:H{r}",
        CellIsRule(operator="lessThanOrEqual", formula=["0.35"], fill=fill(GREEN)))
    ws.sheet_properties.tabColor = "7030A0"; ws.freeze_panes = "A8"


# ---------------------------------------------------------------- 05 Daily Takings
def build_takings(wb):
    ws = wb.create_sheet("05 · Daily Takings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 14, 14, 14, 14, 16, 14, 16])
    title_block(ws, "Daily Takings & Till Reconciliation", "Hárok denných tržieb a uzávierky pokladne",
                "Compares the till Z-read against counted cash + card and flags any variance",
                "Porovná uzávierku pokladne (Z) s hotovosťou a kartami a upozorní na rozdiel")
    biz_field(ws, 3, "Business name", "Názov prevádzky")
    biz_field(ws, 4, "Month (MM/YYYY)", "Mesiac")
    note(ws, 6, "Variance = (cash + card) − Z-read. Anything not near zero is highlighted — investigate before it becomes a habit.",
         "Rozdiel = (hotovosť + karty) − uzávierka. Hodnoty mimo nuly sa zvýraznia — preverte ich.")
    header_row(ws, 7, [
        "Date\nDátum", "Z-read €\nUzávierka €", "Cash €\nHotovosť €", "Card €\nKarty €",
        "Banked €\nVklad €", "Counted total €\nSpolu spočítané €", "Variance €\nRozdiel €", "Checked by\nKontroloval",
    ])
    r = 8; N = 31
    for i in range(N):
        for j in range(1, 9):
            c = ws.cell(r, j, ""); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment("center" if j < 8 else "left", "center")
            if i % 2: c.fill = fill(LGREY)
        ws.cell(r, 1).number_format = "DD/MM/YYYY"
        for col in (2, 3, 4, 5): ws.cell(r, col).fill = fill(AMBER); ws.cell(r, col).number_format = EUR
        ws.cell(r, 6).value = f'=IF(AND(C{r}="",D{r}=""),"",C{r}+D{r})'; ws.cell(r, 6).number_format = EUR
        ws.cell(r, 6).fill = fill(LBLUE)
        ws.cell(r, 7).value = f'=IF(OR(B{r}="",F{r}=""),"",F{r}-B{r})'; ws.cell(r, 7).number_format = EUR
        ws.cell(r, 7).fill = fill(LBLUE)
        ws.row_dimensions[r].height = 16
        r += 1
    rng = f"G8:G{7+N}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["1"], fill=fill(RED)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-1"], fill=fill(RED)))
    ws.sheet_properties.tabColor = "2E75B6"; ws.freeze_panes = "A8"


# ---------------------------------------------------------------- 06 Training & Induction (added)
def build_training(wb):
    ws = wb.create_sheet("06 · Training & Induction")
    ws.sheet_view.showGridLines = False
    topics = [
        ("Induction", "Zaškolenie"),
        ("Manual handling", "Manipulácia s bremenami"),
        ("Food hygiene", "Hygiena potravín"),
        ("Allergen awareness", "Alergény"),
        ("Fire safety", "Požiarna ochrana"),
        ("H&S / first aid", "BOZP / prvá pomoc"),
    ]
    ncols = 2 + len(topics) + 1
    set_widths(ws, [22, 18] + [13] * len(topics) + [14])
    title_block(ws, "Staff Training & Induction Matrix", "Matica školení a zaškolenia zamestnancov",
                "Dated proof each worker was trained BEFORE the floor — incl. manual handling & induction",
                "Dôkaz, že pracovník bol školený PRED zmenou — vrátane manipulácie a zaškolenia", span=ncols)
    biz_field(ws, 3, "Business name", "Názov prevádzky", span=ncols)
    note(ws, 5, "Enter the date each topic was completed (DD/MM/YYYY). Blank cells flag a gap before that person can work that area. Set a refresher cadence in the last column.",
         "Zadajte dátum absolvovania (DD/MM/YYYY). Prázdna bunka = chýbajúce školenie pred prácou. Termín preškolenia v poslednom stĺpci.", span=ncols)
    headers = ["Employee\nZamestnanec", "Start date\nNástup"] + [f"{en}\n{sk}" for en, sk in topics] + ["Refresher due\nPreškolenie"]
    header_row(ws, 6, headers, bg=TEAL)
    seed = ["", "", "", "", "", "", "", ""]
    r = 7; N = 16
    for i in range(N):
        for j in range(1, ncols + 1):
            c = ws.cell(r, j, ""); c.border = BORDER; c.font = f(9)
            c.alignment = Alignment("center" if j > 1 else "left", "center")
            if i % 2: c.fill = fill(LGREY)
        ws.cell(r, 1).fill = fill(AMBER)
        ws.cell(r, 2).number_format = "DD/MM/YYYY"
        for j in range(3, ncols):  # topic date cells
            ws.cell(r, j).number_format = "DD/MM/YYYY"
        ws.cell(r, ncols).number_format = "DD/MM/YYYY"
        ws.row_dimensions[r].height = 18
        r += 1
    # blank topic cell => amber "missing"
    rng = f"C7:{get_column_letter(ncols-1)}{6+N}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['""'], fill=fill(AMBER)))
    note(ws, r + 1, "Reg. note: under the Safety, Health & Welfare at Work Act 2005 employees must be trained for the work they do — keep this matrix current and dated.",
         "Pozn.: podľa zákona o BOZP 2005 musia byť zamestnanci školení na svoju prácu — maticu udržiavajte aktuálnu a s dátumami.", span=ncols, bg=GREEN)
    ws.sheet_properties.tabColor = "C55A11"; ws.freeze_panes = "C7"


def build_workbook():
    wb = Workbook()
    build_cover(wb)
    build_cashflow(wb)
    build_gp(wb)
    build_stock(wb)
    build_rota(wb)
    build_takings(wb)
    build_training(wb)
    wb.properties.title = "Hospitality Operations & GP Bundle (EN/SK)"
    wb.properties.creator = "ASSET-FORGE"
    wb.properties.subject = "Hospitality operations — cashflow, GP costing, stock, rota, takings, training"
    return wb


def add_demo_watermark(wb):
    for ws in wb.worksheets:
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        c = ws.cell(1, 1, "DEMO — PREVIEW ONLY · NOT FOR RESALE · ukážka — buy the full bundle to unlock editing")
        c.font = f(10, True, "FFFFFF"); c.fill = fill("C00000")
        c.alignment = Alignment("center", "center"); ws.row_dimensions[1].height = 22
        ws.protection.sheet = True; ws.protection.password = "demo"; ws.protection.enable()
    wb.properties.title = "Hospitality Operations & GP Bundle — DEMO (EN/SK)"
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    full = os.path.join(PRODUCTS, "P2_Hospitality_Operations_GP_Bundle.xlsx")
    wb.save(full)
    print(f"✓ Built flagship: {full}  ({len(wb.sheetnames)} sheets)")
    demo = add_demo_watermark(build_workbook())
    demo_path = os.path.join(PRODUCTS, "P2_DEMO_Hospitality_Operations_GP_Bundle.xlsx")
    demo.save(demo_path)
    print(f"✓ Built watermarked demo: {demo_path}")
    print("  Sheets:", " | ".join(wb.sheetnames))
