#!/usr/bin/env python3
"""
ASSET-FORGE · Turnkey Startup Packs — phase workbook builders
=============================================================
Emits one premium, formula-linked `.xlsx` per phase (00→06) for a
`StartupPackSpec` (see `startup_spec.py`), reusing the `design_system.py` layer.

    python3 scripts/startup_build.py                       # all phases, all packs
    python3 scripts/startup_build.py boutique_hotel_4star  # one pack
    python3 scripts/startup_build.py boutique_hotel_4star 05  # one phase

Out: products/industries/<industry_dir>/NN_*.xlsx

Design rules honoured:
  · every workbook is SELF-CONTAINED — each has its own orange-input `Assumptions`
    sheet and all formulas link within the file (no fragile cross-file links).
  · all dashboard/figure cells are LINKED formulas, never hard-coded.
  · sample data is included so every calc proves out.
  · 🟢🟡🔴 traffic lights driven by the KPI thresholds in the spec.
  · `ds.fit(ws)` runs on every sheet so nothing clips.
"""
from __future__ import annotations
import os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from design_system import Theme, DS
from startup_spec import STARTUP_REGISTRY, validate, KPI

VERSION, BUILD_DATE = "v1.0", "04/06/2026"

SOURCES = [
    "ADR/occupancy/RevPAR benchmarks: Fáilte Ireland & STR Ireland hotel "
    "performance reports (4-star/boutique ranges).",
    "Grant/funding figures: Local Enterprise Office (localenterprise.ie), "
    "Microfinance Ireland, SBCI, Fáilte Ireland investment schemes.",
    "Cost ratios (payroll ~30%, F&B cost ~30%): USALI hotel P&L conventions.",
    "Figures are seed assumptions — replace the orange cells with your own quotes "
    "and verified local data before relying on them.",
]


# --------------------------------------------------------------- helpers
def new_ds(spec):
    return DS(Theme(primary=spec.palette["primary"], accent=spec.palette["accent"],
                    ink=spec.palette["ink"]))


def fmt_for(ds, unit):
    return {"pct": ds.t.PCT, "eur": ds.t.EUR0, "num": "0.0"}[unit]


def heading(ds, ws, ws_title, ws_sub, widths, tab=None, span=10):
    ds.canvas(ws, widths, tab=tab or ds.t.primary)
    return ds.title(ws, ws_title, ws_sub, span=span)


def label_value(ds, ws, row, label, value, fmt=None, bold=False, col=2, vcol=4,
                vspan=2, input_cell=False):
    """A label (col) + a value/merged value (vcol..). Returns the value cell coord."""
    c = ws.cell(row, col, label)
    c.font = ds.font(10, bold=bold, color=ds.t.ink)
    c.alignment = Alignment("left", "center")
    if vspan > 1:
        ws.merge_cells(start_row=row, start_column=vcol, end_row=row,
                       end_column=vcol + vspan - 1)
    if input_cell:
        vc = ds.input_cell(ws, row, vcol, fmt)
    else:
        vc = ds.calc_cell(ws, row, vcol, fmt, bold=bold)
    if value is not None:
        vc.value = value
    ws.row_dimensions[row].height = 19
    return f"{get_column_letter(vcol)}{row}"


def status_formula(kpi: KPI, valcell: str) -> str:
    """Excel formula → '🟢 On target' / '🟡 Watch' / '🔴 Action'."""
    g, w = kpi.good, kpi.warn
    if kpi.higher_better:
        return (f'=IF({valcell}="","",IF({valcell}>={g},"🟢 On target",'
                f'IF({valcell}>={w},"🟡 Watch","🔴 Action")))')
    return (f'=IF({valcell}="","",IF({valcell}<={g},"🟢 On target",'
            f'IF({valcell}<={w},"🟡 Watch","🔴 Action")))')


def traffic_cf(ws, ds, cell_range):
    """Colour a status-text range green/amber/red by its marker word."""
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("On target",{cell_range.split(":")[0]}))'],
        fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("Watch",{cell_range.split(":")[0]}))'],
        fill=ds.fill(ds.t.warn_bg), font=ds.font(9.5, color=ds.t.warn)))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("Action",{cell_range.split(":")[0]}))'],
        fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, bold=True, color=ds.t.bad)))


def footer(ds, ws, row, spec, extra=""):
    ds.footer(ws, row, f"ASSET-FORGE · LEANTA Turnkey Pack · {spec.vertical} · "
                       f"{VERSION} · {BUILD_DATE}{(' · ' + extra) if extra else ''}")


# ----------------------------------------------------- the Assumptions sheet
def build_assumptions(ws, ds, spec):
    """Single source of truth — orange inputs + linked calcs. Returns a dict of
    cell references other sheets link to (e.g. {'adr': \"'Assumptions'!C7\"}).
    """
    a = spec.assumptions
    name = ws.title
    r = heading(ds, ws, "Assumptions & Inputs",
                "Single source of truth — edit the orange cells; everything links here",
                [34, 4, 18, 14, 26], tab=ds.t.ink, span=5)
    ref = {}

    def put(label, value, fmt, key, input_cell=True, note=""):
        nonlocal r
        c = ws.cell(r, 2, label); c.font = ds.font(10, color=ds.t.ink)
        c.alignment = Alignment("left", "center")
        if input_cell:
            vc = ds.input_cell(ws, r, 4, fmt)
        else:
            vc = ds.calc_cell(ws, r, 4, fmt, bold=True)
        vc.value = value
        if note:
            n = ws.cell(r, 6, note); n.font = ds.font(8.5, italic=True, color=ds.t.muted)
            n.alignment = Alignment("left", "center")
        ref[key] = f"'{name}'!D{r}"
        ws.row_dimensions[r].height = 18
        r += 1

    r = ds.section(ws, r, "Demand & rate", span=5)
    put("Rooms available", a.rooms, "0", "rooms")
    put("ADR — average daily rate", a.adr, ds.t.EUR, "adr", note="4★ boutique IE €110–180")
    put("Occupancy", a.occupancy, ds.t.PCT, "occ", note="annualised; 65–80% typical")
    put("RevPAR (ADR × occupancy)", f"={ref['adr']}*{ref['occ']}", ds.t.EUR, "revpar",
        input_cell=False)
    put("F&B + other revenue (% of rooms)", a.fb_rev_pct, ds.t.PCT, "fb_rev_pct")
    put("Rooms revenue / yr", f"={ref['rooms']}*365*{ref['revpar']}", ds.t.EUR0,
        "rooms_rev", input_cell=False)
    put("Total revenue / yr",
        f"={ref['rooms_rev']}*(1+{ref['fb_rev_pct']})", ds.t.EUR0, "total_rev",
        input_cell=False)

    r += 1
    r = ds.section(ws, r, "Cost ratios", span=5)
    put("Payroll (% of total revenue)", a.payroll_pct, ds.t.PCT, "payroll_pct")
    put("F&B cost (% of F&B revenue)", a.fb_cost_pct, ds.t.PCT, "fb_cost_pct")
    put("Rooms cost (% of rooms revenue)", a.rooms_cost_pct, ds.t.PCT, "rooms_cost_pct")
    put("Utilities (% of total revenue)", a.utilities_pct, ds.t.PCT, "utilities_pct")
    put("Sales & marketing (% of revenue)", a.sales_mktg_pct, ds.t.PCT, "sales_pct")
    put("Admin & general (% of revenue)", a.admin_pct, ds.t.PCT, "admin_pct")
    put("Property: rent/rates/insurance (%)", a.property_pct, ds.t.PCT, "property_pct")

    r += 1
    r = ds.section(ws, r, "Capital expenditure (fit-out led — leased property)", span=5)
    cap0 = r
    for label, val in a.capex.items():
        c = ws.cell(r, 2, label); c.font = ds.font(10, color=ds.t.ink)
        c.alignment = Alignment("left", "center")
        ds.input_cell(ws, r, 4, ds.t.EUR0).value = val
        ws.row_dimensions[r].height = 18; r += 1
    cap_last = r - 1
    c = ws.cell(r, 2, "Total capex required"); c.font = ds.font(10, bold=True, color=ds.t.primary)
    c.alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True).value = f"=SUM(D{cap0}:D{cap_last})"
    ref["capex_total"] = f"'{name}'!D{r}"
    ref["capex_range"] = f"'{name}'!D{cap0}:D{cap_last}"
    ref["capex_labels"] = f"'{name}'!B{cap0}:B{cap_last}"
    r += 2

    r = ds.section(ws, r, "Funding mix", span=5)
    fund0 = r
    for label, val in a.funding.items():
        c = ws.cell(r, 2, label); c.font = ds.font(10, color=ds.t.ink)
        c.alignment = Alignment("left", "center")
        ds.input_cell(ws, r, 4, ds.t.EUR0).value = val
        pc = ds.calc_cell(ws, r, 6, ds.t.PCT)
        pc.value = f"=IF($D${r}=\"\",\"\",D{r}/SUM($D${fund0}:$D${fund0 + len(a.funding) - 1}))"
        ws.row_dimensions[r].height = 18; r += 1
    fund_last = r - 1
    c = ws.cell(r, 2, "Total funding raised"); c.font = ds.font(10, bold=True, color=ds.t.primary)
    c.alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True).value = f"=SUM(D{fund0}:D{fund_last})"
    ref["funding_total"] = f"'{name}'!D{r}"
    ref["funding_range"] = f"'{name}'!D{fund0}:D{fund_last}"
    ref["funding_labels"] = f"'{name}'!B{fund0}:B{fund_last}"
    r += 1
    # funding gap check
    c = ws.cell(r, 2, "Funding vs capex (surplus/gap)")
    c.font = ds.font(10, bold=True, color=ds.t.ink); c.alignment = Alignment("left", "center")
    gap = ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True)
    gap.value = f"={ref['funding_total']}-{ref['capex_total']}"
    ws.conditional_formatting.add(f"D{r}", CellIsRule(
        operator="lessThan", formula=["0"], fill=ds.fill(ds.t.bad_bg),
        font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"D{r}", CellIsRule(
        operator="greaterThanOrEqual", formula=["0"], fill=ds.fill(ds.t.good_bg),
        font=ds.font(9.5, color=ds.t.good)))
    r += 2

    r = ds.section(ws, r, "Sources & benchmark notes", span=5)
    for s in SOURCES:
        r = ds.note(ws, r, s, span=5)
    footer(ds, ws, r + 1, spec)
    return ref


# ============================================================ PHASE 0
def build_market_validation(wb, ds, spec, ref):
    a = spec.assumptions
    ws = wb.create_sheet("Market Validation")
    r = heading(ds, ws, "Market Validation",
                "Phase 0 · Prove demand before you commit capital", [30, 16, 16, 16, 16, 18])

    # --- demand / TAM-SAM-SOM ---
    r = ds.section(ws, r, "Demand sizing (catchment)", span=6)
    r = ds.thead(ws, r, ["Market layer", "Definition", "Annual room-nights",
                         "Capturable %", "Room-nights", "Revenue @ ADR"])
    rows = [
        ("TAM — total area visitors", "All overnight visitors in catchment", 480000),
        ("SAM — 4★ boutique segment", "Mid-upscale leisure + business", 96000),
        ("SOM — realistic year-1 share", "What this hotel can win", 6300),
    ]
    cap = [0.0, 0.0, 1.0]
    start = r
    for i, (layer, defn, nights) in enumerate(rows):
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, layer); ws.cell(r, 3, defn)
        ds.input_cell(ws, r, 4, "#,##0").value = nights
        ds.input_cell(ws, r, 5, ds.t.PCT).value = (nights / rows[0][2]) if i else 1.0
        ds.calc_cell(ws, r, 6, "#,##0").value = f"=D{r}*E{r}"
        ds.calc_cell(ws, r, 7, ds.t.EUR0).value = f"=F{r}*{ref['adr']}"
        r += 1
    som_rev = f"G{r-1}"
    capacity = f"={ref['rooms']}*365*{ref['occ']}"
    r += 1
    r = label_value(ds, ws, r, "Your capacity (rooms×365×occ) room-nights",
                    capacity, "#,##0", bold=True, vcol=4) and r + 1
    r = ds.note(ws, r, "SOM room-nights should be ≤ your capacity. If SOM > capacity "
                "you are capacity-constrained (good problem); if far below, demand is thin.",
                span=6)

    # --- competitor matrix ---
    r += 1
    r = ds.section(ws, r, "Competitor matrix", span=6)
    r = ds.thead(ws, r, ["Competitor", "Rooms", "Est. ADR €", "Rating /5",
                         "Key strength", "Your edge vs them"])
    comps = [
        ("The Harbour Townhouse", 28, 145, 4.3, "Sea views, spa"),
        ("City Mews Boutique", 19, 128, 4.5, "Design-led, central"),
        ("The Old Mill Hotel", 32, 119, 4.0, "Free parking, events"),
        ("Riverside Guesthouse", 12, 99, 4.4, "Owner-hosted, breakfast"),
    ]
    cstart = r
    for i, (nm, rm, adr, rate, strength) in enumerate(comps):
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, nm)
        ds.input_cell(ws, r, 3, "0").value = rm
        ds.input_cell(ws, r, 4, ds.t.EUR0).value = adr
        ds.input_cell(ws, r, 5, "0.0").value = rate
        ws.cell(r, 6, strength)
        ds.input_cell(ws, r, 7)
        r += 1
    cend = r - 1
    r += 1
    r = label_value(ds, ws, r, "Competitor avg ADR", f"=AVERAGE(D{cstart}:D{cend})",
                    ds.t.EUR, bold=True, vcol=4) and r + 1
    r = label_value(ds, ws, r, "Your planned ADR", f"={ref['adr']}", ds.t.EUR,
                    bold=True, vcol=4) and r + 1
    pos = ds.calc_cell(ws, r, 4, ds.t.PCT, bold=True)
    ws.cell(r, 2, "Your ADR vs market").font = ds.font(10, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    pos.value = f"=IFERROR({ref['adr']}/AVERAGE(D{cstart}:D{cend})-1,\"\")"
    r += 2

    # --- location scoring ---
    r = ds.section(ws, r, "Location scoring (weighted)", span=6)
    r = ds.thead(ws, r, ["Factor", "Weight", "Score /10", "Weighted", "", ""])
    factors = [("Footfall / visibility", 0.20, 8), ("Transport access", 0.15, 7),
               ("Nearby demand drivers", 0.20, 9), ("Competition density", 0.15, 6),
               ("Lease cost vs ADR", 0.15, 7), ("Planning / fit-out ease", 0.15, 8)]
    lstart = r
    for i, (f, w, sc) in enumerate(factors):
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, f)
        ds.input_cell(ws, r, 3, ds.t.PCT).value = w
        ds.input_cell(ws, r, 4, "0").value = sc
        ds.calc_cell(ws, r, 5, "0.0").value = f"=C{r}*D{r}"
        r += 1
    lend = r - 1
    sc_cell = f"E{r}"
    ws.cell(r, 2, "Location score / 10").font = ds.font(10, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 5, "0.0", bold=True).value = f"=SUM(E{lstart}:E{lend})"
    st = ws.cell(r, 6, status_formula(KPI("loc", "loc", "num", 7, 7, 5, True, 7), sc_cell))
    st.font = ds.font(9.5); st.alignment = Alignment("center", "center")
    traffic_cf(ws, ds, f"F{r}:F{r}")
    r += 2

    # --- go / no-go ---
    r = ds.section(ws, r, "Phase 0 go / no-go gate", span=6)
    gate = [
        ("SOM room-nights within capacity", f'=IF(F{start+2}<= ({ref["rooms"]}*365),"PASS","REVIEW")'),
        ("ADR competitive (within ±15% of market)", f'=IF(ABS({pos.coordinate})<=0.15,"PASS","REVIEW")'),
        ("Location score ≥ 7/10", f'=IF({sc_cell}>=7,"PASS","REVIEW")'),
    ]
    for i, (label, formula) in enumerate(gate):
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        sc = ds.calc_cell(ws, r, 6); sc.value = formula
        sc.alignment = Alignment("center", "center")
        ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal",
            formula=['"PASS"'], fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
        ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal",
            formula=['"REVIEW"'], fill=ds.fill(ds.t.warn_bg), font=ds.font(9.5, color=ds.t.warn)))
        r += 1
    footer(ds, ws, r + 1, spec, "Phase 0 of 6")
    return ws


# ============================================================ PHASE 1
def build_business_plan(wb, ds, spec, ref):
    a = spec.assumptions
    ws = wb.create_sheet("Business Plan")
    r = heading(ds, ws, "Business Plan",
                "Phase 1 · Service definition + 3-scenario annual P&L", [34, 16, 16, 16, 18])

    r = ds.section(ws, r, "Executive summary", span=5)
    for line in [
        f"A {a.rooms}-room 4★ boutique hotel; ADR €{a.adr:.0f}, target occupancy "
        f"{a.occupancy:.0%}, RevPAR €{a.revpar:.0f}.",
        "Positioning: design-led, service-rich, direct-booking-first to beat OTA "
        "commission and lift repeat stays.",
        "Funded by a blended stack (founder equity incl. SURE refund, LEO grant, "
        "term loan, MFI) — see Phase 2.",
    ]:
        r = ds.note(ws, r, line, span=5)
    r += 1

    # service definition
    r = ds.section(ws, r, "Service definition", span=5)
    r = ds.thead(ws, r, ["Revenue stream", "What it is", "Pricing basis", "", ""])
    svc = [("Rooms", "24 keys, 3 room grades", "ADR + dynamic"),
           ("Food & beverage", "Breakfast, bar, small-plates", "Per cover / menu"),
           ("Events & meetings", "Small functions, private dining", "Day-delegate / hire"),
           ("Ancillary", "Parking, late checkout, retail", "Add-on")]
    for i, (s, d, p) in enumerate(svc):
        ds.trow(ws, r, 5, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, s); ws.cell(r, 3, d); ws.cell(r, 4, p)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        r += 1
    r += 1

    # 3-scenario P&L
    r = ds.section(ws, r, "3-scenario annual P&L (linked to Assumptions)", span=5)
    r = ds.note(ws, r, "Low / Base / High flex occupancy & ADR via the multipliers row. "
                "Every figure links to the Assumptions sheet.", span=5)
    hdr = r
    r = ds.thead(ws, r, ["Line (€/yr)", "Low", "Base", "High", "Driver"])
    # multiplier row
    occ_mult = {"C": 0.85, "D": 1.0, "E": 1.12}
    adr_mult = {"C": 0.93, "D": 1.0, "E": 1.08}
    mrow = r
    ws.cell(r, 2, "Occupancy ×  /  ADR ×").font = ds.font(9.5, italic=True, color=ds.t.muted)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    for col, om in occ_mult.items():
        ds.input_cell(ws, r, {"C": 3, "D": 4, "E": 5}[col], "0.00").value = om
    ws.cell(r, 6, "occupancy flex").font = ds.font(9, italic=True, color=ds.t.muted)
    r += 1
    arow = r
    ws.cell(r, 2, "ADR multiplier").font = ds.font(9.5, italic=True, color=ds.t.muted)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    for col, am in adr_mult.items():
        ds.input_cell(ws, r, {"C": 3, "D": 4, "E": 5}[col], "0.00").value = am
    ws.cell(r, 6, "rate flex").font = ds.font(9, italic=True, color=ds.t.muted)
    r += 1

    def prow(label, formula_for, bold=False, driver="", section=False):
        nonlocal r
        if section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            c = ws.cell(r, 2, label); c.font = ds.font(9.5, bold=True, color=ds.t.primary)
            c.fill = ds.fill(ds.t.band); c.alignment = Alignment("left", "center", indent=1)
            ws.row_dimensions[r].height = 16; r += 1; return None
        c = ws.cell(r, 2, label); c.font = ds.font(9.5, bold=bold, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1); c.border = ds.hairline_bottom()
        cells = {}
        for col in ("C", "D", "E"):
            cc = ds.calc_cell(ws, r, {"C": 3, "D": 4, "E": 5}[col], ds.t.EUR0, bold=bold)
            cc.value = formula_for(col)
            cells[col] = f"{col}{r}"
        if driver:
            d = ws.cell(r, 6, driver); d.font = ds.font(8.5, italic=True, color=ds.t.muted)
            d.alignment = Alignment("left", "center")
        ws.row_dimensions[r].height = 16
        rr = r; r += 1
        return rr

    prow("Revenue", None, section=True)
    rooms_rev = prow("Rooms revenue",
        lambda c: f"={ref['rooms']}*365*({ref['adr']}*{c}{arow})*({ref['occ']}*{c}{mrow})",
        driver="rooms×365×ADR×occ")
    fb_rev = prow("F&B + other revenue",
        lambda c: f"=C{rooms_rev}/0+0".replace("C", c) if False else f"={c}{rooms_rev}*{ref['fb_rev_pct']}",
        driver="% of rooms")
    tot_rev = prow("Total revenue",
        lambda c: f"={c}{rooms_rev}+{c}{fb_rev}", bold=True)
    prow("Costs", None, section=True)
    payroll = prow("Payroll", lambda c: f"={c}{tot_rev}*{ref['payroll_pct']}", driver="% of revenue")
    fbcost = prow("F&B cost of sales", lambda c: f"={c}{fb_rev}*{ref['fb_cost_pct']}", driver="% of F&B rev")
    roomscost = prow("Rooms cost (hskpg etc.)", lambda c: f"={c}{rooms_rev}*{ref['rooms_cost_pct']}", driver="% of rooms rev")
    util = prow("Utilities", lambda c: f"={c}{tot_rev}*{ref['utilities_pct']}", driver="% of revenue")
    mktg = prow("Sales & marketing", lambda c: f"={c}{tot_rev}*{ref['sales_pct']}", driver="% of revenue")
    admin = prow("Admin & general", lambda c: f"={c}{tot_rev}*{ref['admin_pct']}", driver="% of revenue")
    prop = prow("Property (rent/rates/ins.)", lambda c: f"={c}{tot_rev}*{ref['property_pct']}", driver="% of revenue")
    totcost = prow("Total costs",
        lambda c: f"=SUM({c}{payroll}:{c}{prop})", bold=True)
    gop = prow("Gross operating profit",
        lambda c: f"={c}{tot_rev}-{c}{totcost}", bold=True)
    gopm = prow("GOP margin",
        lambda c: f"=IFERROR({c}{gop}/{c}{tot_rev},0)", bold=True, driver="GOP ÷ revenue")
    # format GOP margin row as pct
    for col in (3, 4, 5):
        ws.cell(gopm, col).number_format = ds.t.PCT
    r += 1

    # go/no-go gate on base case
    r = ds.section(ws, r, "Phase 1 go / no-go gate (base case)", span=5)
    gp = ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True)
    ws.cell(r, 2, "Base GOP ≥ €250k?").font = ds.font(10, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    gp.value = f"=D{gop}"
    g2 = ds.calc_cell(ws, r, 5)
    g2.value = f'=IF(D{gop}>=250000,"PASS","REVIEW")'
    g2.alignment = Alignment("center", "center")
    ws.conditional_formatting.add(f"E{r}", CellIsRule(operator="equal", formula=['"PASS"'],
        fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    ws.conditional_formatting.add(f"E{r}", CellIsRule(operator="equal", formula=['"REVIEW"'],
        fill=ds.fill(ds.t.warn_bg), font=ds.font(9.5, color=ds.t.warn)))
    r += 2

    # chart: scenario GOP
    chart = BarChart(); chart.type = "col"; chart.title = "GOP by scenario"
    chart.height = 6; chart.width = 12
    data = Reference(ws, min_col=3, max_col=5, min_row=gop, max_row=gop)
    cats = Reference(ws, min_col=3, max_col=5, min_row=hdr, max_row=hdr)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"B{r}")
    footer(ds, ws, r + 13, spec, "Phase 1 of 6")
    return ws


# ============================================================ PHASE 2
def build_capital_raising(wb, ds, spec, ref):
    a = spec.assumptions
    ws = wb.create_sheet("Capital & Funding")
    r = heading(ds, ws, "Capital Raising",
                "Phase 2 · Capex, funding mix, debt service & 12-month cash-flow",
                [30, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12], span=14)

    # uses & sources summary (links to Assumptions)
    r = ds.section(ws, r, "Uses vs sources", span=6)
    r = label_value(ds, ws, r, "Total capex (uses)", f"={ref['capex_total']}",
                    ds.t.EUR0, bold=True, vcol=4) and r + 1
    r = label_value(ds, ws, r, "Total funding (sources)", f"={ref['funding_total']}",
                    ds.t.EUR0, bold=True, vcol=4) and r + 1
    gap_row = r
    ws.cell(r, 2, "Surplus / (gap)").font = ds.font(10, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 4, ds.t.EUR0, bold=True).value = f"={ref['funding_total']}-{ref['capex_total']}"
    ws.conditional_formatting.add(f"D{r}", CellIsRule(operator="lessThan", formula=["0"],
        fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"D{r}", CellIsRule(operator="greaterThanOrEqual", formula=["0"],
        fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    r += 2

    # debt service schedule
    r = ds.section(ws, r, "Debt service (term loan)", span=6)
    loan_amt = 300000; rate = 0.075; years = 7
    di = r
    r = label_value(ds, ws, r, "Loan principal €", loan_amt, ds.t.EUR0, input_cell=True, vcol=4) and r + 1
    pr_cell = f"D{di}"
    r = label_value(ds, ws, r, "Interest rate (APR)", rate, ds.t.PCT, input_cell=True, vcol=4) and r + 1
    rt_cell = f"D{di+1}"
    r = label_value(ds, ws, r, "Term (years)", years, "0", input_cell=True, vcol=4) and r + 1
    tm_cell = f"D{di+2}"
    ws.cell(r, 2, "Monthly repayment €").font = ds.font(10, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    pmt = ds.calc_cell(ws, r, 4, ds.t.EUR, bold=True)
    pmt.value = f"=IFERROR(-PMT({rt_cell}/12,{tm_cell}*12,{pr_cell}),0)"
    pmt_cell = f"D{r}"
    r = r + 1
    ws.cell(r, 2, "Annual debt service €").font = ds.font(10, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 4, ds.t.EUR0).value = f"={pmt_cell}*12"
    r += 2

    # 12-month cash-flow with ramp
    r = ds.section(ws, r, "12-month opening cash-flow (with occupancy ramp)", span=14)
    r = ds.note(ws, r, "Revenue ramps from a soft-open fraction up to steady state. "
                "Costs scale with revenue; debt service is constant. Closing cash carries forward.",
                span=14)
    months = ["M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"]
    head = r
    r = ds.thead(ws, r, ["Line (€)"] + months + ["Year"])
    ramp = [0.45, 0.55, 0.62, 0.68, 0.74, 0.80, 0.85, 0.88, 0.85, 0.80, 0.74, 0.70]
    ramp_row = r
    ws.cell(r, 2, "Occupancy ramp ×").font = ds.font(9, italic=True, color=ds.t.muted)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    for j, v in enumerate(ramp):
        ds.input_cell(ws, r, 3 + j, "0.00").value = v
    r += 1

    def cf_line(label, per_month, bold=False, total=True):
        nonlocal r
        c = ws.cell(r, 2, label); c.font = ds.font(9.5, bold=bold, color=ds.t.ink)
        c.alignment = Alignment("left", "center", indent=1); c.border = ds.hairline_bottom()
        for j in range(12):
            col = get_column_letter(3 + j)
            ds.calc_cell(ws, r, 3 + j, ds.t.EUR0, bold=bold).value = per_month(col)
        if total:
            ds.calc_cell(ws, r, 15, ds.t.EUR0, bold=True).value = f"=SUM(C{r}:N{r})"
        ws.row_dimensions[r].height = 15
        rr = r; r += 1; return rr

    base_month_rev = f"({ref['total_rev']}/12)"
    rev = cf_line("Revenue", lambda col: f"={base_month_rev}*({col}{ramp_row}/{ref['occ']})")
    payroll = cf_line("Payroll", lambda col: f"={col}{rev}*{ref['payroll_pct']}")
    other = cf_line("Other operating costs",
        lambda col: f"={col}{rev}*({ref['fb_cost_pct']}*{ref['fb_rev_pct']}+{ref['rooms_cost_pct']}+{ref['utilities_pct']}+{ref['sales_pct']}+{ref['admin_pct']}+{ref['property_pct']})")
    debt = cf_line("Debt service", lambda col: f"={pmt_cell}")
    net = cf_line("Net cash flow", lambda col: f"={col}{rev}-{col}{payroll}-{col}{other}-{col}{debt}", bold=True)
    # opening/closing cash
    ws.cell(r, 2, "Opening cash").font = ds.font(9.5, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    ds.calc_cell(ws, r, 3, ds.t.EUR0).value = f"={ref['funding_total']}-{ref['capex_total']}"
    ob = r
    for j in range(1, 12):
        col = get_column_letter(3 + j); prev = get_column_letter(2 + j)
        ds.calc_cell(ws, r, 3 + j, ds.t.EUR0).value = f"={prev}{ob+1}"
    r += 1
    cb = r
    ws.cell(r, 2, "Closing cash").font = ds.font(9.5, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center", indent=1)
    for j in range(12):
        col = get_column_letter(3 + j)
        ds.calc_cell(ws, r, 3 + j, ds.t.EUR0, bold=True).value = f"={col}{ob}+{col}{net}"
    ds.calc_cell(ws, r, 15, ds.t.EUR0, bold=True).value = f"=N{cb}"
    ws.conditional_formatting.add(f"C{cb}:N{cb}", CellIsRule(operator="lessThan", formula=["0"],
        fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    r += 2

    # runway + DSCR
    r = ds.section(ws, r, "Runway & cover", span=6)
    r = label_value(ds, ws, r, "Lowest monthly closing cash €",
                    f"=MIN(C{cb}:N{cb})", ds.t.EUR0, bold=True, vcol=4) and r + 1
    r = label_value(ds, ws, r, "Min cash buffer status",
                    f'=IF(MIN(C{cb}:N{cb})>=0,"🟢 Positive all year","🔴 Goes negative — raise more")',
                    None, bold=True, vcol=4, vspan=4) and r + 1
    ws.cell(r, 2, "DSCR (GOP ÷ debt service)").font = ds.font(10, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    dscr = ds.calc_cell(ws, r, 4, "0.00", bold=True)
    dscr.value = f"=IFERROR((O{rev}-O{payroll}-O{other})/(O{debt}),0)"
    ws.conditional_formatting.add(f"D{r}", CellIsRule(operator="greaterThanOrEqual",
        formula=["1.25"], fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    ws.conditional_formatting.add(f"D{r}", CellIsRule(operator="lessThan",
        formula=["1.25"], fill=ds.fill(ds.t.warn_bg), font=ds.font(9.5, color=ds.t.warn)))
    r += 2

    # funding mix chart
    chart = BarChart(); chart.type = "bar"; chart.title = "Funding mix (€)"
    chart.height = 6; chart.width = 12
    # build a small local table for the chart from Assumptions refs is cross-sheet;
    # instead reference the funding range on Assumptions directly
    footer(ds, ws, r, spec, "Phase 2 of 6")
    return ws


# ============================================================ PHASE 3
def build_procurement(wb, ds, spec, ref):
    ws = wb.create_sheet("Procurement")
    r = heading(ds, ws, "Procurement",
                "Phase 3 · Vendors, fit-out budget, RFQ compare & purchase orders",
                [26, 20, 14, 14, 14, 14, 16])

    # vendor master
    r = ds.section(ws, r, "Vendor master", span=7)
    r = ds.thead(ws, r, ["Vendor", "Category", "Lead time (wk)", "Payment terms",
                         "Rating /5", "Contact", "Approved?"])
    vendors = [("Murphy FF&E Ltd", "Furniture & fit-out", 8, "30% / 70%", 4.5),
               ("Atlantic Linen Co", "Linen & textiles", 3, "Net 30", 4.2),
               ("KitchenPro", "Kitchen equipment", 6, "50% deposit", 4.0),
               ("BrightIT Systems", "PMS/POS/network", 4, "Net 14", 4.6),
               ("GreenClean Supplies", "Housekeeping consumables", 1, "Net 30", 4.1)]
    yn = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
    ws.add_data_validation(yn)
    for i, (nm, cat, lt, pay, rate) in enumerate(vendors):
        ds.trow(ws, r, 7, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, nm); ws.cell(r, 3, cat)
        ds.input_cell(ws, r, 4, "0").value = lt
        ws.cell(r, 5, pay); ds.input_cell(ws, r, 6, "0.0").value = rate
        ds.input_cell(ws, r, 7); ds.input_cell(ws, r, 8); yn.add(ws.cell(r, 8))
        ws.cell(r, 8).value = "Yes" if i < 3 else "Pending"
        r += 1
    r += 1

    # RFQ comparison (auto best price)
    r = ds.section(ws, r, "RFQ comparison — auto-flags lowest quote", span=7)
    r = ds.thead(ws, r, ["Item", "Quote A €", "Quote B €", "Quote C €",
                         "Lowest €", "Best vendor", "Saving vs A"])
    items = [("36 bedroom furniture sets", 64000, 59500, 61200),
             ("Commercial kitchen line", 52000, 55000, 49900),
             ("PMS + POS (3-yr)", 18000, 21000, 17400),
             ("Bed linen (par 3)", 14500, 13900, 15200)]
    qs = r
    for i, (it, a_, b_, c_) in enumerate(items):
        ds.trow(ws, r, 7, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, it)
        ds.input_cell(ws, r, 3, ds.t.EUR0).value = a_
        ds.input_cell(ws, r, 4, ds.t.EUR0).value = b_
        ds.input_cell(ws, r, 5, ds.t.EUR0).value = c_
        ds.calc_cell(ws, r, 6, ds.t.EUR0, bold=True).value = f"=MIN(C{r}:E{r})"
        bv = ds.calc_cell(ws, r, 7)
        bv.value = f'=CHOOSE(MATCH(MIN(C{r}:E{r}),C{r}:E{r},0),"Quote A","Quote B","Quote C")'
        bv.alignment = Alignment("center", "center")
        ds.calc_cell(ws, r, 8, ds.t.EUR0).value = f"=C{r}-F{r}"
        r += 1
    qe = r - 1
    ws.cell(r, 2, "Total lowest-quote spend").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 6, ds.t.EUR0, bold=True).value = f"=SUM(F{qs}:F{qe})"
    ds.calc_cell(ws, r, 8, ds.t.EUR0, bold=True).value = f"=SUM(H{qs}:H{qe})"
    r += 2

    # purchase order register
    r = ds.section(ws, r, "Purchase order register", span=7)
    r = ds.thead(ws, r, ["PO #", "Vendor", "Description", "Order date", "Due date",
                         "Amount €", "Status"])
    stat = DataValidation(type="list", formula1='"Draft,Sent,Confirmed,Delivered,Paid"', allow_blank=True)
    ws.add_data_validation(stat)
    pos = r
    seed_pos = [("PO-001", "Murphy FF&E Ltd", "Bedroom furniture x24"),
                ("PO-002", "KitchenPro", "Kitchen line install"),
                ("PO-003", "BrightIT Systems", "PMS + POS rollout")]
    for i in range(14):
        ds.trow(ws, r, 7, zebra_on=(i % 2 == 1))
        if i < len(seed_pos):
            ws.cell(r, 2, seed_pos[i][0]); ws.cell(r, 3, seed_pos[i][1]); ws.cell(r, 4, seed_pos[i][2])
        else:
            ds.input_cell(ws, r, 2); ds.input_cell(ws, r, 3); ds.input_cell(ws, r, 4)
        ds.input_cell(ws, r, 5, ds.t.DATE); ds.input_cell(ws, r, 6, ds.t.DATE)
        ds.input_cell(ws, r, 7, ds.t.EUR0)
        ds.input_cell(ws, r, 8); stat.add(ws.cell(r, 8))
        if i < len(seed_pos):
            ws.cell(r, 7).value = [59500, 49900, 17400][i]
            ws.cell(r, 8).value = ["Confirmed", "Sent", "Draft"][i]
        r += 1
    poe = r - 1
    ws.cell(r, 2, "Committed spend (PO register)").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 7, ds.t.EUR0, bold=True).value = f"=SUM(G{pos}:G{poe})"
    r += 1
    ws.cell(r, 2, "Capex budget (Assumptions)").font = ds.font(9.5, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 7, ds.t.EUR0).value = f"={ref['capex_total']}"
    r += 1
    ws.cell(r, 2, "Budget remaining").font = ds.font(9.5, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 7, ds.t.EUR0, bold=True).value = f"={ref['capex_total']}-SUM(G{pos}:G{poe})"
    ws.freeze_panes = "B5"
    footer(ds, ws, r + 2, spec, "Phase 3 of 6")
    return ws


# ============================================================ PHASE 4
def build_team_building(wb, ds, spec, ref):
    a = spec.assumptions
    ws = wb.create_sheet("Team & Payroll")
    r = heading(ds, ws, "Team Building",
                "Phase 4 · Org design, hiring tracker & payroll budget (PRSI/pension)",
                [26, 14, 12, 14, 14, 14, 16])

    # staffing plan / payroll budget
    r = ds.section(ws, r, "Staffing plan & payroll budget", span=7)
    r = ds.note(ws, r, "Employer PRSI ~11.05% and pension ~5% are added on gross. Total "
                "payroll should land near the Assumptions payroll % of revenue.", span=7)
    r = ds.thead(ws, r, ["Role", "Dept", "FTE", "Gross €/yr", "PRSI+pension",
                         "Total cost €", "Notes"])
    roles = [("General Manager", "Leadership", 1, 55000),
             ("Front Office Manager", "Front office", 1, 38000),
             ("Receptionist", "Front office", 2.5, 30000),
             ("Head Housekeeper", "Housekeeping", 1, 34000),
             ("Room Attendant", "Housekeeping", 3, 27500),
             ("Head Chef", "F&B", 1, 48000),
             ("Chef de Partie / KP", "F&B", 2, 31000),
             ("F&B / Bar Attendant", "F&B", 2.5, 28000),
             ("Maintenance", "Facilities", 1, 36000)]
    loadrow = r
    ws.cell(r, 2, "On-cost loading (PRSI+pension)").font = ds.font(9, italic=True, color=ds.t.muted)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.input_cell(ws, r, 6, ds.t.PCT).value = 0.16
    load = f"F{loadrow}"
    r += 1
    rs = r
    for i, (role, dept, fte, gross) in enumerate(roles):
        ds.trow(ws, r, 7, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, role); ws.cell(r, 3, dept)
        ds.input_cell(ws, r, 4, "0.0").value = fte
        ds.input_cell(ws, r, 5, ds.t.EUR0).value = gross
        ds.calc_cell(ws, r, 6, ds.t.EUR0).value = f"=D{r}*E{r}*{load}"
        ds.calc_cell(ws, r, 7, ds.t.EUR0, bold=True).value = f"=D{r}*E{r}*(1+{load})"
        ds.input_cell(ws, r, 8)
        r += 1
    re = r - 1
    ws.cell(r, 2, "Total FTE / payroll").font = ds.font(9.5, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    ds.calc_cell(ws, r, 4, "0.0", bold=True).value = f"=SUM(D{rs}:D{re})"
    pay_cell = f"G{r}"
    ds.calc_cell(ws, r, 7, ds.t.EUR0, bold=True).value = f"=SUM(G{rs}:G{re})"
    r += 1
    ws.cell(r, 2, "Payroll as % of revenue").font = ds.font(9.5, bold=True, color=ds.t.ink)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    pctc = ds.calc_cell(ws, r, 7, ds.t.PCT, bold=True)
    pctc.value = f"=IFERROR({pay_cell}/{ref['total_rev']},0)"
    ws.conditional_formatting.add(f"G{r}", CellIsRule(operator="greaterThan",
        formula=[str(a.payroll_pct + 0.03)], fill=ds.fill(ds.t.bad_bg), font=ds.font(9.5, True, ds.t.bad)))
    ws.conditional_formatting.add(f"G{r}", CellIsRule(operator="lessThanOrEqual",
        formula=[str(a.payroll_pct + 0.03)], fill=ds.fill(ds.t.good_bg), font=ds.font(9.5, color=ds.t.good)))
    r += 2

    # recruitment tracker
    r = ds.section(ws, r, "Recruitment tracker", span=7)
    r = ds.thead(ws, r, ["Role", "Stage", "Candidate", "Interview score /10",
                         "Offer status", "Start date", "Notes"])
    stage = DataValidation(type="list", formula1='"Open,Advertised,Screening,Interview,Offer,Hired"', allow_blank=True)
    offer = DataValidation(type="list", formula1='"—,Offered,Accepted,Declined"', allow_blank=True)
    ws.add_data_validation(stage); ws.add_data_validation(offer)
    for i in range(10):
        ds.trow(ws, r, 7, zebra_on=(i % 2 == 1))
        if i < len(roles):
            ws.cell(r, 2, roles[i][0])
        else:
            ds.input_cell(ws, r, 2)
        ds.input_cell(ws, r, 3); stage.add(ws.cell(r, 3))
        ds.input_cell(ws, r, 4)
        ds.input_cell(ws, r, 5, "0.0")
        ds.input_cell(ws, r, 6); offer.add(ws.cell(r, 6))
        ds.input_cell(ws, r, 7, ds.t.DATE); ds.input_cell(ws, r, 8)
        r += 1
    ws.freeze_panes = "B5"
    footer(ds, ws, r + 1, spec, "Phase 4 of 6")
    return ws


# ============================================================ PHASE 5
def build_operations(wb, ds, spec, ref):
    a = spec.assumptions
    # --- daily data entry sheet (drives the dashboard) ---
    de = wb.create_sheet("Daily Data Entry")
    r = heading(ds, de, "Daily Data Entry",
                "Phase 5 · Log the day — the dashboard reads from here",
                [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12], span=11)
    cols = ["Date", "Rooms sold", "Rooms avail", "Rooms rev €", "F&B rev €",
            "F&B cost €", "Direct bookings", "Total bookings", "Repeat guests",
            "NPS responses"]
    r = ds.thead(de, r, cols)
    ds0 = r
    import datetime as _dt
    sample = [
        # rooms_sold, avail, rooms_rev, fb_rev, fb_cost, direct, total_bk, repeat, nps
        (18, 24, 2484, 690, 221, 7, 17, 5, 60),
        (20, 24, 2760, 760, 243, 8, 18, 6, 58),
        (17, 24, 2295, 640, 205, 6, 16, 4, 62),
        (21, 24, 2898, 800, 256, 9, 19, 7, 55),
        (22, 24, 3036, 840, 269, 8, 20, 6, 59),
        (19, 24, 2622, 720, 230, 7, 18, 5, 61),
        (16, 24, 2160, 600, 192, 6, 15, 4, 57),
    ]
    for i, row in enumerate(sample):
        rr = ds0 + i
        ds.trow(de, rr, 11, zebra_on=(i % 2 == 1))
        dc = ds.input_cell(de, rr, 2, ds.t.DATE); dc.value = _dt.date(2026, 6, 1 + i)
        for j, v in enumerate(row):
            ds.input_cell(de, rr, 3 + j, ds.t.EUR0 if j in (2, 3, 4) else "0").value = v
    # blank rows for more entry
    for i in range(len(sample), 31):
        rr = ds0 + i
        ds.trow(de, rr, 11, zebra_on=(i % 2 == 1))
        ds.input_cell(de, rr, 2, ds.t.DATE)
        for j in range(9):
            ds.input_cell(de, rr, 3 + j, ds.t.EUR0 if j in (2, 3, 4) else "0")
    de_last = ds0 + 30
    de.freeze_panes = "B" + str(ds0)
    dn = de.title
    # aggregate references used by the dashboard
    agg = {
        "rooms_sold": f"SUM('{dn}'!C{ds0}:C{de_last})",
        "rooms_avail": f"SUM('{dn}'!D{ds0}:D{de_last})",
        "rooms_rev": f"SUM('{dn}'!E{ds0}:E{de_last})",
        "fb_rev": f"SUM('{dn}'!F{ds0}:F{de_last})",
        "fb_cost": f"SUM('{dn}'!G{ds0}:G{de_last})",
        "direct": f"SUM('{dn}'!I{ds0}:I{de_last})",
        "total_bk": f"SUM('{dn}'!J{ds0}:J{de_last})",
        "repeat": f"SUM('{dn}'!K{ds0}:K{de_last})",
        "nps": f"AVERAGE('{dn}'!L{ds0}:L{de_last})",
    }
    footer(ds, de, de_last + 2, spec, "Phase 5 · data entry")

    # --- the dashboard ---
    ws = wb.create_sheet("KPI Dashboard")
    r = heading(ds, ws, "Operations KPI Dashboard",
                "Phase 5 · 15 live hotel metrics vs target with 🟢🟡🔴 status",
                [13] * 11, span=11)

    # compute each KPI value from the data-entry aggregates (some need other inputs)
    # supplementary single-input cells live just below the table for metrics not in the log
    # Build the KPI table first
    r = ds.section(ws, r, "Metrics vs target", span=11)
    head = r
    r = ds.thead(ws, r, ["KPI", "Actual", "Target", "Status", "How it's measured", "", "", "", "", ""])
    # formulas for the 15 KPIs (linked to data-entry where possible)
    occ_v = f"IFERROR({agg['rooms_sold']}/{agg['rooms_avail']},0)"
    adr_v = f"IFERROR({agg['rooms_rev']}/{agg['rooms_sold']},0)"
    revpar_v = f"IFERROR({agg['rooms_rev']}/{agg['rooms_avail']},0)"
    fbcost_v = f"IFERROR({agg['fb_cost']}/{agg['fb_rev']},0)"
    direct_v = f"IFERROR({agg['direct']}/{agg['total_bk']},0)"
    repeat_v = f"IFERROR({agg['repeat']}/{agg['total_bk']},0)"
    nps_v = f"IFERROR({agg['nps']},0)"
    # GOPPAR ≈ (rooms_rev+fb_rev - costs)/rooms_avail ; approximate operating cost via ratios
    goppar_v = (f"IFERROR((({agg['rooms_rev']}+{agg['fb_rev']})*(1-{ref['payroll_pct']}"
                f"-{ref['utilities_pct']}-{ref['sales_pct']}-{ref['admin_pct']}-{ref['property_pct']})"
                f"-{agg['fb_cost']})/{agg['rooms_avail']},0)")
    # the rest come from manual KPI inputs (orange) below the table
    formulas = {
        "occupancy": occ_v, "adr": adr_v, "revpar": revpar_v, "goppar": goppar_v,
        "nps": nps_v, "payroll_pct": None, "fb_cost_pct": fbcost_v, "repeat_pct": repeat_v,
        "alos": None, "direct_pct": direct_v, "turnaround": None, "maint_resp": None,
        "staff_turnover": None, "complaints": None, "collection_pct": None,
    }
    table_start = r
    input_rows = {}
    for i, kpi in enumerate(spec.kpis):
        ds.trow(ws, r, 11, zebra_on=(i % 2 == 1), align="left")
        ws.cell(r, 2, kpi.label)
        fmt = fmt_for(ds, kpi.unit)
        act = ds.calc_cell(ws, r, 3, fmt, bold=True)
        if formulas[kpi.key]:
            act.value = f"={formulas[kpi.key]}"
        else:
            # manual metric → orange input seeded with sample
            ds.input_cell(ws, r, 3, fmt).value = kpi.sample
            input_rows[kpi.key] = r
        ds.calc_cell(ws, r, 4, fmt).value = kpi.target
        st = ws.cell(r, 5, status_formula(kpi, f"C{r}"))
        st.font = ds.font(9.5); st.alignment = Alignment("center", "center")
        st.border = ds.hairline_bottom()
        note = ws.cell(r, 6, kpi.note); note.font = ds.font(9, color=ds.t.muted)
        note.alignment = Alignment("left", "center")
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=11)
        r += 1
    table_end = r - 1
    traffic_cf(ws, ds, f"E{table_start}:E{table_end}")
    r += 1

    # summary tiles (hero KPIs) above were skipped; add a compact scorecard
    r = ds.section(ws, r, "Headline scorecard", span=11)
    heroes = ["occupancy", "adr", "revpar", "goppar"]
    positions = [2, 5, 8, 11]
    kpi_by = {k.key: (idx, k) for idx, k in enumerate(spec.kpis)}
    tile_top = r
    for n, key in enumerate(heroes):
        idx, kpi = kpi_by[key]
        src = f"C{table_start + idx}"
        fmt = fmt_for(ds, kpi.unit)
        accent = [ds.t.primary, ds.t.accent, ds.t.ink, ds.t.primary][n]
        ds.kpi(ws, tile_top, positions[n], kpi.label, f"=IFERROR({src},0)", fmt=fmt,
               width=2, accent=accent)
    r = tile_top + 5

    # count of red/amber/green
    r = ds.section(ws, r, "Status summary", span=11)
    for label, marker, tone in [("🟢 On target", "On target", "good"),
                                 ("🟡 Watch", "Watch", "warn"),
                                 ("🔴 Action", "Action", "bad")]:
        ws.cell(r, 2, label).font = ds.font(10, bold=True, color=ds.t.ink)
        ws.cell(r, 2).alignment = Alignment("left", "center")
        cc = ds.calc_cell(ws, r, 4, "0", bold=True)
        cc.value = f'=COUNTIF(E{table_start}:E{table_end},"*{marker}*")'
        r += 1
    # trend chart from the daily entry (occupancy proxy: rooms sold / avail per day)
    chart = LineChart(); chart.title = "Daily rooms sold (sample)"; chart.height = 6; chart.width = 16
    data = Reference(de, min_col=3, max_col=3, min_row=ds0, max_row=ds0 + 6)
    chart.add_data(data, titles_from_data=False); chart.legend = None
    ws.add_chart(chart, f"B{r+1}")
    footer(ds, ws, r + 14, spec, "Phase 5 of 6")
    return ws


# ============================================================ PHASE 6
def build_launch(wb, ds, spec, ref):
    ws = wb.create_sheet("Launch 100 Days")
    r = heading(ds, ws, "Launch & First 100 Days",
                "Phase 6 · Soft open → grand open → 100-day scorecard",
                [28, 16, 14, 14, 16, 16], span=6)

    # soft-open / grand-open plan
    r = ds.section(ws, r, "Launch timeline", span=6)
    r = ds.thead(ws, r, ["Milestone", "Day", "Owner", "Status", "Success measure", ""])
    miles = [("Friends & family soft open", "-7", "GM", "Snag list cleared"),
             ("Staff dry-run service", "-3", "FOM", "Full guest journey rehearsed"),
             ("Soft opening (50% inventory)", "1", "GM", "≥60% occ, no critical faults"),
             ("Grand opening event", "14", "Marketing", "Press + 100 guests"),
             ("Full inventory live", "21", "GM", "All 24 rooms sellable"),
             ("100-day review", "100", "GM", "Scorecard green")]
    sd = DataValidation(type="list", formula1='"Not started,In progress,Done,Blocked"', allow_blank=True)
    ws.add_data_validation(sd)
    for i, (m, d, o, meas) in enumerate(miles):
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, m); ds.input_cell(ws, r, 3).value = d
        ws.cell(r, 4, o); ds.input_cell(ws, r, 5); sd.add(ws.cell(r, 5))
        ws.cell(r, 5).value = "Not started"
        ws.cell(r, 6, meas)
        r += 1
    r += 1

    # 100-day scorecard (links to targets)
    r = ds.section(ws, r, "100-day scorecard", span=6)
    r = ds.thead(ws, r, ["KPI", "Day-100 actual", "Target", "Status", "", ""])
    score_kpis = ["occupancy", "adr", "revpar", "nps", "direct_pct", "repeat_pct"]
    kmap = {k.key: k for k in spec.kpis}
    ss = r
    for i, key in enumerate(score_kpis):
        kpi = kmap[key]; fmt = fmt_for(ds, kpi.unit)
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, kpi.label)
        ds.input_cell(ws, r, 3, fmt).value = kpi.sample
        ds.calc_cell(ws, r, 4, fmt).value = kpi.target
        st = ws.cell(r, 5, status_formula(kpi, f"C{r}"))
        st.alignment = Alignment("center", "center"); st.font = ds.font(9.5)
        st.border = ds.hairline_bottom()
        r += 1
    se = r - 1
    traffic_cf(ws, ds, f"E{ss}:E{se}")
    r += 1

    # contingency triggers + go/no-go
    r = ds.section(ws, r, "Contingency triggers", span=6)
    r = ds.thead(ws, r, ["If this happens…", "Trigger level", "Then do this", "", "", ""])
    cont = [("Occupancy < 50% for 2 wks", "Red", "Tactical OTA push + rate review"),
            ("Cash buffer < €25k", "Red", "Draw contingency / defer capex"),
            ("NPS < 30", "Amber", "Service huddle + recovery training"),
            ("Staff turnover > 45%", "Amber", "Pay/benefits review + exit interviews")]
    for i, (cond, lvl, act) in enumerate(cont):
        ds.trow(ws, r, 6, zebra_on=(i % 2 == 1))
        ws.cell(r, 2, cond); ws.cell(r, 3, lvl)
        ws.cell(r, 4, act); ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        r += 1
    r += 1
    ws.cell(r, 2, "Scale-up go / no-go").font = ds.font(10, bold=True, color=ds.t.primary)
    ws.cell(r, 2).alignment = Alignment("left", "center")
    g = ds.calc_cell(ws, r, 4)
    g.value = f'=IF(COUNTIF(E{ss}:E{se},"*On target*")>=4,"GO — scale","HOLD — stabilise")'
    g.alignment = Alignment("center", "center")
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    footer(ds, ws, r + 2, spec, "Phase 6 of 6")
    return ws


# --------------------------------------------------------------- orchestrator
PHASE_BUILDERS = {
    "market_validation": build_market_validation,
    "business_plan": build_business_plan,
    "capital_raising": build_capital_raising,
    "procurement": build_procurement,
    "team_building": build_team_building,
    "operations": build_operations,
    "launch_100days": build_launch,
}


def out_dir(spec):
    d = os.path.normpath(os.path.join(HERE, "..", "products", "industries", spec.industry_dir))
    os.makedirs(d, exist_ok=True)
    return d


def build_phase(spec, phase):
    ds = new_ds(spec)
    wb = Workbook(); wb.remove(wb.active)
    aws = wb.create_sheet("Assumptions")
    ref = build_assumptions(aws, ds, spec)
    PHASE_BUILDERS[phase.key](wb, ds, spec, ref)
    # order: phase sheet(s) first, Assumptions last
    sheets = [s for s in wb.sheetnames if s != "Assumptions"] + ["Assumptions"]
    wb._sheets = [wb[n] for n in sheets]
    for ws in wb.worksheets:
        ds.fit(ws)
    wb.properties.title = f"{spec.vertical} · {phase.title}"
    wb.properties.creator = "ASSET-FORGE · LEANTA"
    path = os.path.join(out_dir(spec), phase.filename)
    wb.save(path)
    return path


def build_pack(pack_key, only_phase=None):
    spec = STARTUP_REGISTRY[pack_key]
    errs = validate(spec)
    assert not errs, f"{pack_key} invalid: {errs}"
    built = []
    for phase in spec.phases:
        if only_phase and phase.num != only_phase:
            continue
        path = build_phase(spec, phase)
        built.append(path)
        print(f"✓ {pack_key} {phase.num} {phase.title}: {os.path.relpath(path, HERE)}")
    return built


if __name__ == "__main__":
    args = sys.argv[1:]
    keys = [a for a in args if a in STARTUP_REGISTRY] or list(STARTUP_REGISTRY)
    phase = next((a for a in args if a.isdigit() or (len(a) == 2 and a.isdigit())), None)
    for k in keys:
        build_pack(k, only_phase=phase)
